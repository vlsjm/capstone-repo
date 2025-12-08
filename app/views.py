from django.views.decorators.http import require_POST, require_http_methods
from django.http import HttpResponseRedirect
from django.urls import reverse
import re
from .permissions import has_admin_permission, admin_permission_required, AdminPermissionMixin

def redirect_with_tab(request, view_name):
    """
    Helper function to redirect while preserving the current tab from the referer URL.
    
    Args:
        request: The HTTP request object
        view_name: The name of the view to redirect to
    
    Returns:
        HttpResponseRedirect to the view, optionally with tab parameter preserved
    """
    referer = request.META.get('HTTP_REFERER', '')
    if 'tab=' in referer:
        # Extract the tab parameter from the referer
        tab_match = re.search(r'tab=([^&]+)', referer)
        if tab_match:
            tab = tab_match.group(1)
            return redirect(f"{reverse(view_name)}?tab={tab}")
    return redirect(view_name)

@require_POST
def update_damage_status(request, pk):
    report = get_object_or_404(DamageReport, pk=pk)
    action = request.POST.get('update_action')
    remarks = report.remarks or ''
    if action == 'unserviceable':
        report.remarks = f"Classified as Unserviceable. {remarks}"
        report.save()
    elif action == 'needs_repair':
        report.remarks = f"Classified as Needs Repair. {remarks}"
        report.save()
    elif action == 'back_in_use':
        report.status = 'reviewed'
        report.remarks = f"Marked as Back in Use. {remarks}"
        report.save()
    return HttpResponseRedirect(reverse('damaged_items_management'))

@require_POST
@admin_permission_required('manage_lost_items')
def mark_lost_item_found(request, pk):
    """Mark a lost item as found and restore property availability"""
    from .models import LostItem
    lost_item = get_object_or_404(LostItem, pk=pk)
    
    # Restore property condition to 'In good condition' and mark as available
    lost_item.item.condition = 'In good condition'
    lost_item.item.availability = 'available'
    lost_item.item.save(update_fields=['condition', 'availability'])
    
    # Update lost item status to 'found'
    lost_item.status = 'found'
    lost_item.remarks = f"Item marked as found by {request.user.username}"
    lost_item.save()
    
    # Log the activity
    ActivityLog.log_activity(
        user=request.user,
        action='update',
        model_name='LostItem',
        object_repr=str(lost_item.item.property_name),
        description=f"Marked lost property '{lost_item.item.property_name}' as found and restored to available status"
    )
    
    messages.success(request, f"{lost_item.item.property_name} has been marked as found and is now available.")
    return HttpResponseRedirect(reverse('damaged_items_management') + '?tab=lost-items')

@admin_permission_required('manage_lost_items')
def mark_property_as_lost(request, property_id):
    """
    View for admins to verify and mark a property as lost.
    This is typically used after reviewing a lost item report to confirm the item is indeed lost.
    Updates the property condition to 'Lost' and availability to 'not_available'.
    """
    # Check if user is authenticated and has admin permissions
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'message': 'Authentication required'}, status=401)
    
    try:
        if not request.user.userprofile.role == 'ADMIN':
            return JsonResponse({'success': False, 'message': 'Admin access required'}, status=403)
    except:
        return JsonResponse({'success': False, 'message': 'Access denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST method required'}, status=405)
    
    property_obj = get_object_or_404(Property, id=property_id)
    
    # Store old condition for logging
    old_condition = property_obj.condition
    
    # Update property condition to Lost
    property_obj.condition = 'Lost'
    property_obj.availability = 'not_available'
    property_obj.save(update_fields=['condition', 'availability'])
    
    # Log the activity
    ActivityLog.log_activity(
        user=request.user,
        action='update',
        model_name='Property',
        object_repr=str(property_obj.property_name),
        description=f"Admin verified and marked property '{property_obj.property_name}' as Lost (previous condition: {old_condition})"
    )
    
    messages.success(request, f"{property_obj.property_name} has been marked as Lost.")
    
    # Return JSON response for AJAX calls, or redirect to damaged items page
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'message': f'{property_obj.property_name} has been marked as Lost.'
        })
    
    # Redirect to damaged items management page with lost-items tab
    return HttpResponseRedirect(reverse('damaged_items_management') + '?tab=lost-items')

@require_POST
@admin_permission_required('manage_lost_items')
def delete_lost_item(request, pk):
    """Delete a lost item report"""
    from .models import LostItem
    lost_item = get_object_or_404(LostItem, pk=pk)
    item_name = lost_item.item.property_name
    
    # Optionally restore property availability
    lost_item.item.availability = 'available'
    lost_item.item.save(update_fields=['availability'])
    
    lost_item.delete()
    
    messages.success(request, f"Lost item report for {item_name} has been deleted.")
    return HttpResponseRedirect(reverse('damaged_items_management') + '?tab=lost-items')

@require_POST
def update_property_condition(request, pk):
    property_obj = get_object_or_404(Property, pk=pk)
    new_condition = request.POST.get('condition')
    if new_condition in ['Unserviceable', 'Needing repair', 'Working', 'In good condition']:
        property_obj.condition = new_condition
        if new_condition in ['Unserviceable', 'Needing repair']:
            property_obj.availability = 'not_available'
        else:
            property_obj.availability = 'available'
        property_obj.save()
    return HttpResponseRedirect(reverse('damaged_items_management'))
# Damaged Items Management Page (Admin)
from django.views.generic import TemplateView

from django.contrib.auth.mixins import PermissionRequiredMixin

class DamagedItemsManagementView(PermissionRequiredMixin, TemplateView):
    permission_required = 'app.view_admin_module'
    template_name = 'app/damaged_items_management.html'

    def _get_reported_date_for_property(self, prop, condition):
        """Get the date when a property was set to a specific condition"""
        from .models import PropertyHistory
        from django.utils import timezone
        from datetime import datetime
        
        # Try to find the most recent history entry where condition was changed to the specified condition
        history = PropertyHistory.objects.filter(
            property=prop,
            field_name='condition',
            new_value=condition
        ).order_by('-timestamp').first()
        
        if history:
            return history.timestamp
        
        # Fallback to minimum aware datetime if no history found (will sort to end)
        return timezone.make_aware(datetime.min.replace(year=1900))

    def get_context_data(self, **kwargs):
        from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
        from django.utils import timezone
        context = super().get_context_data(**kwargs)
        
        # Get properties with unserviceable condition
        unserviceable_properties = Property.objects.filter(condition__iexact='Unserviceable').order_by('-id')
        
        # Get properties needing repair
        needs_repair_properties = Property.objects.filter(condition__iexact='Needing repair').order_by('-id')
        
        # Add reported_date to each property by fetching from PropertyHistory
        for prop in unserviceable_properties:
            prop.reported_date = self._get_reported_date_for_property(prop, 'Unserviceable')
        
        for prop in needs_repair_properties:
            prop.reported_date = self._get_reported_date_for_property(prop, 'Needing repair')
        
        # Sort by reported_date (most recent first)
        from datetime import datetime
        min_aware_datetime = timezone.make_aware(datetime.min.replace(year=1900))
        unserviceable_properties = sorted(unserviceable_properties, key=lambda x: x.reported_date if x.reported_date else min_aware_datetime, reverse=True)
        needs_repair_properties = sorted(needs_repair_properties, key=lambda x: x.reported_date if x.reported_date else min_aware_datetime, reverse=True)
        
        # Get lost items (only show items with 'lost' status, exclude 'found' items)
        from .models import LostItem
        lost_items = LostItem.objects.filter(status='lost').select_related('item', 'user').order_by('-report_date')
        
        # Get URL parameters for building pagination links
        url_params = ""
        
        # Paginate each tab separately with 15 items per page
        paginate_by = 15
        
        # Get page numbers for each tab
        unserviceable_page = self.request.GET.get('unserviceable_page', 1)
        needs_repair_page = self.request.GET.get('needs_repair_page', 1)
        lost_items_page = self.request.GET.get('lost_items_page', 1)
        
        # Paginate unserviceable items
        unserviceable_paginator = Paginator(unserviceable_properties, paginate_by)
        try:
            unserviceable_page_obj = unserviceable_paginator.page(unserviceable_page)
        except PageNotAnInteger:
            unserviceable_page_obj = unserviceable_paginator.page(1)
        except EmptyPage:
            unserviceable_page_obj = unserviceable_paginator.page(unserviceable_paginator.num_pages)
        
        # Paginate needs repair items
        needs_repair_paginator = Paginator(needs_repair_properties, paginate_by)
        try:
            needs_repair_page_obj = needs_repair_paginator.page(needs_repair_page)
        except PageNotAnInteger:
            needs_repair_page_obj = needs_repair_paginator.page(1)
        except EmptyPage:
            needs_repair_page_obj = needs_repair_paginator.page(needs_repair_paginator.num_pages)
        
        # Paginate lost items
        lost_items_paginator = Paginator(lost_items, paginate_by)
        try:
            lost_items_page_obj = lost_items_paginator.page(lost_items_page)
        except PageNotAnInteger:
            lost_items_page_obj = lost_items_paginator.page(1)
        except EmptyPage:
            lost_items_page_obj = lost_items_paginator.page(lost_items_paginator.num_pages)
        
        # Also include damage reports that have been classified (for backward compatibility)
        unserviceable_reports = DamageReport.objects.filter(status='resolved', remarks__icontains='Unserviceable')
        needs_repair_reports = DamageReport.objects.filter(status='resolved', remarks__icontains='Needs Repair')
        
        # Add paginated objects to context
        context['unserviceable_items'] = unserviceable_page_obj
        context['needs_repair_items'] = needs_repair_page_obj
        context['lost_items'] = lost_items_page_obj
        
        # Add total counts for badges
        context['unserviceable_count'] = len(unserviceable_properties)
        context['needs_repair_count'] = len(needs_repair_properties)
        context['lost_items_count'] = lost_items.count()
        
        # Add legacy reports for backward compatibility
        context['unserviceable_reports'] = unserviceable_reports
        context['needs_repair_reports'] = needs_repair_reports
        context['url_params'] = url_params
        
        return context

def export_unserviceable_items(request):
    """Export unserviceable items to Excel"""
    # Check permissions
    if not request.user.is_authenticated or not request.user.userprofile.role == 'ADMIN':
        return HttpResponse('Unauthorized', status=403)
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.utils import timezone
    
    # Get unserviceable properties
    properties = Property.objects.filter(condition__iexact='Unserviceable').select_related('category').order_by('-id')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Unserviceable Items"
    
    # Header styling
    header_fill = PatternFill(start_color="152d64", end_color="152d64", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:I1')
    ws['A1'] = 'Unserviceable Items Report'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:I2')
    ws['A2'] = f'Generated on: {timezone.now().strftime("%B %d, %Y %I:%M %p")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Property No.', 'Item Name', 'Category', 'Description', 'Location', 
               'Accountable Person', 'Year Acquired', 'Unit Value', 'Quantity']
    ws.append([])  # Empty row
    ws.append(headers)
    
    # Style headers
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Add data
    for prop in properties:
        ws.append([
            prop.property_number or 'N/A',
            prop.property_name,
            prop.category.name if prop.category else 'N/A',
            prop.description or 'N/A',
            prop.location or 'N/A',
            prop.accountable_person or 'N/A',
            prop.year_acquired.strftime('%Y-%m-%d') if prop.year_acquired else 'N/A',
            float(prop.unit_value) if prop.unit_value else 0,
            prop.quantity or 0
        ])
    
    # Style data rows
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    # Adjust column widths
    column_widths = [15, 25, 20, 30, 20, 25, 15, 12, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Unserviceable_Items_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response

def export_needs_repair_items(request):
    """Export items needing repair to Excel"""
    # Check permissions
    if not request.user.is_authenticated or not request.user.userprofile.role == 'ADMIN':
        return HttpResponse('Unauthorized', status=403)
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.utils import timezone
    
    # Get properties needing repair
    properties = Property.objects.filter(condition__iexact='Needing repair').select_related('category').order_by('-id')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Items Needing Repair"
    
    # Header styling
    header_fill = PatternFill(start_color="152d64", end_color="152d64", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:I1')
    ws['A1'] = 'Items Needing Repair Report'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:I2')
    ws['A2'] = f'Generated on: {timezone.now().strftime("%B %d, %Y %I:%M %p")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Property No.', 'Item Name', 'Category', 'Description', 'Location', 
               'Accountable Person', 'Year Acquired', 'Unit Value', 'Quantity']
    ws.append([])  # Empty row
    ws.append(headers)
    
    # Style headers
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Add data
    for prop in properties:
        ws.append([
            prop.property_number or 'N/A',
            prop.property_name,
            prop.category.name if prop.category else 'N/A',
            prop.description or 'N/A',
            prop.location or 'N/A',
            prop.accountable_person or 'N/A',
            prop.year_acquired.strftime('%Y-%m-%d') if prop.year_acquired else 'N/A',
            float(prop.unit_value) if prop.unit_value else 0,
            prop.quantity or 0
        ])
    
    # Style data rows
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    # Adjust column widths
    column_widths = [15, 25, 20, 30, 20, 25, 15, 12, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Items_Needing_Repair_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response

def export_lost_items(request):
    """Export lost items to Excel"""
    # Check permissions
    if not request.user.is_authenticated or not request.user.userprofile.role == 'ADMIN':
        return HttpResponse('Unauthorized', status=403)
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.utils import timezone
    from .models import LostItem
    
    # Get lost items
    lost_items = LostItem.objects.filter(status='lost').select_related('item', 'item__category', 'user').order_by('-report_date')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Lost Items"
    
    # Header styling
    header_fill = PatternFill(start_color="152d64", end_color="152d64", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:L1')
    ws['A1'] = 'Lost Items Report'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:L2')
    ws['A2'] = f'Generated on: {timezone.now().strftime("%B %d, %Y %I:%M %p")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Property No.', 'Item Name', 'Category', 'Reported By', 'Report Date',
               'Last Seen Location', 'Last Seen Date', 'Description', 'Condition Status',
               'Unit Value', 'Quantity', 'Remarks']
    ws.append([])  # Empty row
    ws.append(headers)
    
    # Style headers
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Add data
    for lost_item in lost_items:
        prop = lost_item.item
        condition_status = 'Verified Lost' if prop.condition == 'Lost' else 'Pending Verification'
        
        ws.append([
            prop.property_number or 'N/A',
            prop.property_name,
            prop.category.name if prop.category else 'N/A',
            lost_item.user.username,
            lost_item.report_date.strftime('%Y-%m-%d %I:%M %p'),
            lost_item.last_seen_location or 'N/A',
            lost_item.last_seen_date.strftime('%Y-%m-%d') if lost_item.last_seen_date else 'N/A',
            lost_item.description[:100] if lost_item.description else 'N/A',
            condition_status,
            float(prop.unit_value) if prop.unit_value else 0,
            prop.quantity or 0,
            lost_item.remarks or 'N/A'
        ])
    
    # Style data rows
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    # Adjust column widths
    column_widths = [15, 25, 20, 15, 20, 20, 15, 35, 18, 12, 10, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Lost_Items_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response

from collections import defaultdict
from django.contrib import messages
from django.contrib.auth.models import User, Group
from django.contrib.auth.views import LoginView, LogoutView, PasswordChangeView, PasswordChangeDoneView
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import TemplateView, ListView
from django.core.exceptions import ValidationError
from django.db.models import Count, Q, F
from django.views import View
from django.contrib.auth import login, authenticate, logout
from .models import(
    Supply, Property, BorrowRequest, BorrowRequestBatch, BorrowRequestItem,
    SupplyRequest, DamageReport, LostItem, Reservation, ReservationBatch, ReservationItem,
    ActivityLog, UserProfile, Notification, AdminPermission,
    SupplyQuantity, SupplyHistory, PropertyHistory,
    Department, PropertyCategory, SupplyCategory, SupplySubcategory,
    SupplyRequestBatch, SupplyRequestItem, BadStockReport, UserSession
)
from .forms import PropertyForm, PropertyNumberChangeForm, SupplyForm, UserProfileForm, UserRegistrationForm, DepartmentForm, SupplyRequestBatchForm, SupplyRequestItemForm, SupplyRequestBatchForm, SupplyRequestItemForm, BadStockReportForm
from django.contrib.auth.forms import AuthenticationForm
from datetime import timedelta, date, datetime
import json
from django.utils import timezone
from django.utils.timezone import now
from django.http import JsonResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import PermissionRequiredMixin, LoginRequiredMixin
from django.contrib.auth.decorators import permission_required
from django.db import models
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.utils import timezone
from datetime import timedelta
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
import logging
import os
from .utils import generate_barcode
from openpyxl import Workbook, load_workbook
from django.http import HttpResponse
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from io import BytesIO
# Create a logger
logger = logging.getLogger(__name__)

#user create user
@admin_permission_required('manage_users')
def create_user(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            try:
                # Create user
                initial_password = form.cleaned_data['password']
                user = User.objects.create_user(
                    username=form.cleaned_data['username'],
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name'],
                    email=form.cleaned_data['email'],
                    password=initial_password,
                )
                user.is_staff = True
                user.save()

                role = form.cleaned_data['role']
            
                # Add user to group
                try:
                    group = Group.objects.get(name=role)
                    user.groups.add(group)
                except Group.DoesNotExist:
                    messages.warning(request, f'Group {role} does not exist. User created without group assignment.')

                # Create user profile
                # Set must_change_password to True for USER and ADMIN roles (new users must change password on first login)
                must_change_pwd = (role == 'USER' or role == 'ADMIN')
                
                profile = UserProfile.objects.create(
                    user=user,
                    role=role,
                    department=form.cleaned_data['department'],
                    phone=form.cleaned_data['phone'],
                    must_change_password=must_change_pwd,
                )

                # Send welcome email
                try:
                    context = {
                        'user': user,
                        'user_profile': profile,
                        'initial_password': initial_password,
                    }
                    html_message = render_to_string('app/email/account_created.html', context)
                    plain_message = strip_tags(html_message)
                    
                    send_mail(
                        subject='Welcome to ResourceHive - Your Account Has Been Created',
                        message=plain_message,
                        from_email=settings.EMAIL_HOST_USER,
                        recipient_list=[user.email],
                        html_message=html_message,
                        fail_silently=False,
                    )
                    messages.success(request, f'Account created successfully for {user.username} and welcome email sent.')
                except Exception as e:
                    messages.warning(request, f'Account created successfully but failed to send welcome email.')

                # Log the activity
                ActivityLog.log_activity(
                    user=request.user,
                    action='create',
                    model_name='User',
                    object_repr=user.username,
                    description=f"Created new user account for {user.username} with role {role}"
                )

                # Check if custom permissions configuration was requested
                configure_permissions = request.POST.get('configure_permissions') == 'true'
                
                # If AJAX request (for permissions flow), return JSON
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'user_id': profile.id,
                        'username': user.username,
                        'configure_permissions': configure_permissions
                    })
                
                # Store flag for opening permissions modal after redirect
                if configure_permissions and role == 'ADMIN':
                    request.session['open_permissions_for'] = profile.id
                    request.session['permissions_user_created'] = user.username
                
                return redirect('user_profile_list')
                
            except Exception as e:
                # If AJAX request, return JSON error
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'error': str(e)
                    })
                
                messages.error(request, f'Error creating account: {str(e)}')
                # Store form data in session to preserve it
                request.session['form_data'] = request.POST.dict()
                request.session['show_modal'] = True
                return redirect('manage_users')
        else:
            # If AJAX request, return JSON with errors
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                form_errors_dict = {}
                for field, errors in form.errors.items():
                    form_errors_dict[field] = list(errors)
                return JsonResponse({
                    'success': False,
                    'error': 'Please correct the errors below.',
                    'errors': form_errors_dict
                })
            
            # Store form data and errors in session as simple dict
            request.session['form_data'] = request.POST.dict()
            # Convert errors to simple list of strings per field
            form_errors_dict = {}
            for field, errors in form.errors.items():
                form_errors_dict[field] = list(errors)
            request.session['form_errors'] = form_errors_dict
            request.session['show_modal'] = True
            return redirect('manage_users')
    
    return redirect('manage_users')

def create_department(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            Department.objects.create(name=name)
            messages.success(request, "Department added.")
    return redirect('manage_users')

@permission_required('app.view_admin_module')
def edit_department(request, dept_id):
    if request.method == 'POST':
        department = get_object_or_404(Department, id=dept_id)
        new_name = request.POST.get('name')
        if new_name:
            department.name = new_name
            department.save()
            return JsonResponse({"success": True})
    return JsonResponse({"success": False})

@permission_required('app.view_admin_module')
def delete_department(request, dept_id):
    if request.method == 'POST':
        department = get_object_or_404(Department, id=dept_id)
        # Check if there are any users in this department
        if UserProfile.objects.filter(department=department).exists():
            return JsonResponse({
                "success": False,
                "error": "Cannot delete department that has users assigned to it. Please reassign or remove users first."
            })
        department.delete()
        return JsonResponse({"success": True})
    return JsonResponse({"success": False})

@permission_required('app.view_admin_module')
@require_POST
def toggle_user_status(request, user_id):
    """Toggle user active/inactive status"""
    try:
        user = get_object_or_404(User, id=user_id)
        user_profile = get_object_or_404(UserProfile, user=user)
        
        # Prevent admin from deactivating themselves
        if user == request.user:
            messages.error(request, "You cannot deactivate your own account.")
            return redirect('manage_users')
        
        # Toggle the is_active status
        user.is_active = not user.is_active
        
        if not user.is_active:
            # User is being deactivated - check for auto-reactivation date
            auto_enable_date = request.POST.get('auto_enable_date')
            if auto_enable_date:
                try:
                    from datetime import datetime
                    # Parse the datetime-local input format
                    enable_datetime = datetime.strptime(auto_enable_date, '%Y-%m-%dT%H:%M')
                    # Make it timezone aware
                    from django.utils import timezone as tz
                    enable_datetime_aware = tz.make_aware(enable_datetime)
                    
                    # Ensure the date is in the future
                    if enable_datetime_aware > timezone.now():
                        user_profile.auto_enable_at = enable_datetime_aware
                        user_profile.save()
                        status_text = f"deactivated (will auto-reactivate on {enable_datetime_aware.strftime('%b %d, %Y at %I:%M %p')})"
                    else:
                        messages.warning(request, "Auto-reactivation date must be in the future. Account deactivated without auto-reactivation.")
                        user_profile.auto_enable_at = None
                        user_profile.save()
                        status_text = "deactivated"
                except ValueError:
                    messages.warning(request, "Invalid date format. Account deactivated without auto-reactivation.")
                    user_profile.auto_enable_at = None
                    user_profile.save()
                    status_text = "deactivated"
            else:
                user_profile.auto_enable_at = None
                user_profile.save()
                status_text = "deactivated"
        else:
            # User is being reactivated - clear auto_enable_at
            user_profile.auto_enable_at = None
            user_profile.save()
            status_text = "activated"
        
        user.save()
        
        # Log the activity
        ActivityLog.log_activity(
            user=request.user,
            action='activate' if user.is_active else 'update',
            model_name='User',
            object_repr=user.username,
            description=f"{status_text.capitalize()} user account for {user.username}"
        )
        
        messages.success(request, f"User {user.username} has been {status_text}.")
        return redirect('manage_users')
        
    except Exception as e:
        messages.error(request, f"Error updating user status: {str(e)}")
        return redirect('manage_users')

@login_required
@require_POST
def mark_all_notifications_as_read(request):
    try:
        data = json.loads(request.body)
        ids = data.get('notification_ids', [])
        Notification.objects.filter(id__in=ids, user=request.user, is_read=False).update(is_read=True)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_POST
def clear_all_notifications(request):
    try:
        Notification.objects.filter(user=request.user).delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_POST
def mark_notification_as_read_ajax(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            notification_id = data.get('notification_id')
        except json.JSONDecodeError:
            notification_id = request.POST.get('notification_id')
        
        if not notification_id:
            return JsonResponse({'error': 'No notification ID provided'}, status=400)

        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        notification.is_read = True
        notification.save()
        return JsonResponse({'success': True})
    return JsonResponse({'error': 'Invalid request method'}, status=400)


from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from .models import Reservation, DamageReport, BorrowRequest, SupplyRequest, Notification

def reservation_detail(request, pk):
    """Detail view for individual reservation (redirects to batch view if part of batch)"""
    reservation = get_object_or_404(Reservation, pk=pk)
    
    # If this reservation is part of a batch, redirect to batch detail
    if reservation.batch_id:
        return redirect('reservation_batch_detail', batch_id=reservation.batch_id)
    
    # Legacy single reservation view
    if request.method == 'POST':
        action = request.POST.get('action')
        remarks = request.POST.get('remarks')
        reservation.remarks = remarks
        
        if action == 'approve':
            # Check for overlapping reservations
            overlapping_reservations = Reservation.objects.filter(
                item=reservation.item,
                status__in=['approved', 'active'],
                needed_date__lte=reservation.return_date,
                return_date__gte=reservation.needed_date
            ).exclude(pk=reservation.pk)
            
            # Calculate total quantity reserved for the overlapping period
            total_reserved = sum(r.quantity for r in overlapping_reservations)
            available_quantity = reservation.item.quantity - total_reserved
            
            if available_quantity >= reservation.quantity:
                reservation.status = 'approved'
                reservation.approved_date = timezone.now()
                Notification.objects.create(
                    user=reservation.user,
                    message=f"Your reservation for {reservation.item.property_name} has been approved.",
                    remarks=remarks
                )
                reservation.save()
                messages.success(request, 'Reservation approved successfully.')
            else:
                messages.error(request, f'Cannot approve reservation. Only {available_quantity} items available for the requested period.')
                return render(request, 'app/reservation_detail.html', {'reservation': reservation})
                
        elif action == 'reject':
            reservation.status = 'rejected'
            reservation.approved_date = timezone.now()
            Notification.objects.create(
                user=reservation.user,
                message=f"Your reservation for {reservation.item.property_name} has been rejected.",
                remarks=remarks
            )
            reservation.save()
            messages.success(request, 'Reservation rejected successfully.')
            
        return redirect('user_reservations')
    return render(request, 'app/reservation_detail.html', {'reservation': reservation})


@login_required
@permission_required('app.view_admin_module', raise_exception=True)
def reservation_batch_detail(request, batch_id):
    """View and manage a batch reservation request"""
    # Check and update reservation batch statuses (triggers active reservations)
    ReservationBatch.check_and_update_batches()
    
    # Get the reservation batch
    batch = get_object_or_404(ReservationBatch.objects.select_related('user').prefetch_related('items__property'), id=batch_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        remarks = request.POST.get('remarks', '').strip()
        
        # Handle void action for approved reservations
        if action == 'void_request' and batch.status == 'approved':
            # Check permission
            from .permissions import has_admin_permission
            if not has_admin_permission(request.user, 'void_request'):
                messages.error(request, 'You do not have permission to void requests.')
                return redirect('reservation_batch_detail', batch_id=batch_id)
            
            # Validate that remarks is provided
            if not remarks:
                messages.error(request, 'Please provide a reason for voiding this reservation.')
                return redirect('reservation_batch_detail', batch_id=batch_id)
            
            # Release reserved quantities from approved items
            approved_items = batch.items.filter(status='approved')
            for item in approved_items:
                # Release the reserved quantity
                item.property.reserved_quantity = max(0, item.property.reserved_quantity - item.quantity)
                item.property.save()
            
            # Update ALL items to voided status (not just approved ones)
            batch.items.all().update(status='voided')
            
            # Change batch status to voided
            batch.status = 'voided'
            batch.remarks = remarks
            batch.save()
            
            # Create notification for user
            Notification.objects.create(
                user=batch.user,
                message=f"Your reservation batch #{batch.id} has been voided by admin.",
                remarks=remarks
            )
            
            # Send email notification to the requester
            if batch.user.email:
                from django.core.mail import send_mail
                from django.conf import settings
                
                subject = f'Reservation Request #{batch.id} - Voided'
                message = f"""Dear {batch.user.get_full_name() or batch.user.username},

Your reservation request (Batch #{batch.id}) has been voided by the administrator.

Reason for voiding:
{remarks}

All reserved quantities have been released back to the inventory.

If you have any questions, please contact the administrator.

Best regards,
Resource Hive Management System
"""
                try:
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        [batch.user.email],
                        fail_silently=False,
                    )
                except Exception as e:
                    # Log the error but don't fail the request
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f"Failed to send void notification email: {str(e)}")
            
            # Log activity
            ActivityLog.log_activity(
                user=request.user,
                action='update',
                model_name='ReservationBatch',
                object_repr=f"Reservation Batch #{batch.id}",
                description=f"Voided approved reservation, released reserved quantities. Reason: {remarks[:100]}"
            )
            
            messages.success(request, 'Reservation voided successfully. Reserved quantities have been released and the requester has been notified.')
            return redirect('reservation_batch_detail', batch_id=batch_id)
    
    # Refresh the batch object to get updated item statuses
    batch = ReservationBatch.objects.select_related('user').prefetch_related('items__property').get(id=batch_id)
    
    # Get all items in this batch
    items = batch.items.all().order_by('id')
    
    # Calculate status counts
    status_counts = {}
    for item in items:
        status_counts[item.status] = status_counts.get(item.status, 0) + 1
    
    from .permissions import has_admin_permission
    context = {
        'batch': batch,
        'items': items,
        'batch_id': batch.id,
        'batch_number': batch.id,
        'batch_status': batch.status,
        'total_items': batch.total_items,
        'total_quantity': batch.total_quantity,
        'status_counts': status_counts,
        'can_void_request': has_admin_permission(request.user, 'void_request'),
    }
    
    return render(request, 'app/reservation_batch_detail.html', context)


@login_required
@permission_required('app.view_admin_module', raise_exception=True)
@admin_permission_required('approve_reservation')
def approve_reservation_item(request, batch_id, item_id):
    """Approve a single item in a reservation batch"""
    if request.method == 'POST':
        batch = get_object_or_404(ReservationBatch, id=batch_id)
        item = get_object_or_404(ReservationItem, id=item_id, batch_request=batch)
        remarks = request.POST.get('remarks', '')
        scroll_position = request.POST.get('scroll_position', '0')
        
        # Get approved quantity from POST (default to original quantity if not provided)
        approved_quantity = request.POST.get('approved_quantity', item.quantity)
        try:
            approved_quantity = int(approved_quantity)
        except (ValueError, TypeError):
            approved_quantity = item.quantity
        
        # Validate approved quantity
        if approved_quantity < 1:
            messages.error(request, 'Approved quantity must be at least 1.')
            response = redirect('reservation_batch_detail', batch_id=batch_id)
            response['Location'] += f'#scroll-{scroll_position}'
            return response
        
        if approved_quantity > item.quantity:
            messages.error(request, f'Approved quantity cannot exceed requested quantity of {item.quantity}.')
            response = redirect('reservation_batch_detail', batch_id=batch_id)
            response['Location'] += f'#scroll-{scroll_position}'
            return response
        
        # Check available quantity (uses reserved_quantity from Property model)
        if item.property.available_quantity >= approved_quantity:
            # Update quantity to approved quantity
            original_quantity = item.quantity
            item.quantity = approved_quantity
            item.status = 'approved'
            
            # Add remarks about quantity adjustment if different
            if approved_quantity < original_quantity:
                adjustment_note = f"Approved quantity: {approved_quantity} (Originally requested: {original_quantity})"
                item.remarks = f"{adjustment_note}\n{remarks}" if remarks else adjustment_note
            else:
                item.remarks = remarks
            
            item.save()
            
            # Update batch approved_date if first approval
            if not batch.approved_date:
                batch.approved_date = timezone.now()
            
            # Check if all items are now approved/rejected
            all_items = batch.items.all()
            if all(i.status in ['approved', 'rejected'] for i in all_items):
                # If some approved and some rejected
                if any(i.status == 'approved' for i in all_items) and any(i.status == 'rejected' for i in all_items):
                    batch.status = 'partially_approved'
                # If all approved
                elif all(i.status == 'approved' for i in all_items):
                    batch.status = 'approved'
                # If all rejected
                elif all(i.status == 'rejected' for i in all_items):
                    batch.status = 'rejected'
                batch.save()
            
            # Notify user about approval (handled by ReservationItem save method if needed)
            # For individual item approval, send a notification
            approval_msg = f"Item {item.property.property_name} in reservation batch #{batch_id} has been approved."
            if approved_quantity < original_quantity:
                approval_msg += f" Approved quantity: {approved_quantity} (Requested: {original_quantity})"
            
            Notification.objects.create(
                user=batch.user,
                message=approval_msg,
                remarks=remarks if remarks else None
            )
            
            # Log activity
            ActivityLog.log_activity(
                user=request.user,
                action='approve',
                model_name='ReservationItem',
                object_repr=f"{item.property.property_name} (Batch #{batch_id})",
                description=f"Approved reservation item #{item_id} in batch #{batch_id} - Quantity: {approved_quantity}"
            )
            
            success_msg = f'Reservation for {item.property.property_name} approved successfully.'
            if approved_quantity < original_quantity:
                success_msg += f' (Approved: {approved_quantity} out of {original_quantity} requested)'
            messages.success(request, success_msg)
        else:
            messages.error(request, f'Cannot approve reservation. Only {item.property.available_quantity} units of {item.property.property_name} available (trying to approve {approved_quantity}).')
        
        response = redirect('reservation_batch_detail', batch_id=batch_id)
        response['Location'] += f'#scroll-{scroll_position}'
        return response
    
    return redirect('user_reservations')


@login_required
@permission_required('app.view_admin_module', raise_exception=True)
@admin_permission_required('approve_reservation')
def reject_reservation_item(request, batch_id, item_id):
    """Reject a single item in a reservation batch"""
    if request.method == 'POST':
        batch = get_object_or_404(ReservationBatch, id=batch_id)
        item = get_object_or_404(ReservationItem, id=item_id, batch_request=batch)
        remarks = request.POST.get('remarks', '')
        scroll_position = request.POST.get('scroll_position', '0')
        
        item.status = 'rejected'
        item.remarks = remarks
        item.save()
        
        # Update batch status if needed
        all_items = batch.items.all()
        if all(i.status in ['approved', 'rejected'] for i in all_items):
            # If some approved and some rejected
            if any(i.status == 'approved' for i in all_items) and any(i.status == 'rejected' for i in all_items):
                batch.status = 'partially_approved'
            # If all rejected
            elif all(i.status == 'rejected' for i in all_items):
                batch.status = 'rejected'
                batch.approved_date = timezone.now()
            batch.save()
        
        # Notify user about rejection
        Notification.objects.create(
            user=batch.user,
            message=f"Item {item.property.property_name} in reservation batch #{batch_id} has been rejected.",
            remarks=remarks if remarks else None
        )
        
        # Log activity
        ActivityLog.log_activity(
            user=request.user,
            action='reject',
            model_name='ReservationItem',
            object_repr=f"{item.property.property_name} (Batch #{batch_id})",
            description=f"Rejected reservation item #{item_id} in batch #{batch_id}"
        )
        
        messages.success(request, f'Reservation for {item.property.property_name} rejected.')
        response = redirect('reservation_batch_detail', batch_id=batch_id)
        response['Location'] += f'#scroll-{scroll_position}'
        return response
    
    return redirect('user_reservations')


@login_required
@permission_required('app.view_admin_module', raise_exception=True)
@admin_permission_required('approve_reservation')
def approve_reservation_batch(request, batch_id):
    """Approve all pending items in a reservation batch"""
    if request.method == 'POST':
        batch = get_object_or_404(ReservationBatch, id=batch_id)
        items = batch.items.filter(status='pending')
        remarks = request.POST.get('remarks', '')
        
        if not items.exists():
            messages.warning(request, 'No pending items to approve in this batch.')
            return redirect('reservation_batch_detail', batch_id=batch_id)
        
        approved_count = 0
        failed_items = []
        
        for item in items:
            # Check available quantity (uses reserved_quantity from Property model)
            if item.property.available_quantity >= item.quantity:
                item.status = 'approved'
                item.remarks = remarks
                item.save()
                approved_count += 1
            else:
                failed_items.append(f"{item.property.property_name} (only {item.property.available_quantity} available)")
        
        # Update batch status
        if approved_count > 0:
            if not batch.approved_date:
                batch.approved_date = timezone.now()
            
            # Check final status
            all_items = batch.items.all()
            if all(i.status == 'approved' for i in all_items):
                batch.status = 'approved'
            elif any(i.status == 'approved' for i in all_items):
                batch.status = 'partially_approved'
            batch.save()
            
            # Notify user once for the batch
            Notification.objects.create(
                user=batch.user,
                message=f"Your batch reservation request #{batch_id} has been approved ({approved_count} items).",
                remarks=remarks if remarks else None
            )
            
            # Log activity
            ActivityLog.log_activity(
                user=request.user,
                action='approve',
                model_name='ReservationBatch',
                object_repr=f"Batch #{batch_id}",
                description=f"Approved {approved_count} items in reservation batch"
            )
            
            messages.success(request, f'{approved_count} item(s) approved successfully.')
        
        if failed_items:
            messages.warning(request, f'Could not approve: {", ".join(failed_items)}')
        
        return redirect('reservation_batch_detail', batch_id=batch_id)
    
    return redirect('user_reservations')


@login_required
@permission_required('app.view_admin_module', raise_exception=True)
@admin_permission_required('approve_reservation')
def reject_reservation_batch(request, batch_id):
    """Reject all pending items in a reservation batch"""
    if request.method == 'POST':
        batch = get_object_or_404(ReservationBatch, id=batch_id)
        items = batch.items.filter(status='pending')
        remarks = request.POST.get('remarks', '')
        
        if not items.exists():
            messages.warning(request, 'No pending items to reject in this batch.')
            return redirect('reservation_batch_detail', batch_id=batch_id)
        
        count = items.count()
        
        # Update all pending items
        for item in items:
            item.status = 'rejected'
            item.remarks = remarks
            item.save()
        
        # Update batch status
        batch.status = 'rejected'
        batch.approved_date = timezone.now()
        batch.remarks = remarks
        batch.save()
        
        # Notify user once for the batch
        Notification.objects.create(
            user=batch.user,
            message=f"Your batch reservation request #{batch_id} has been rejected ({count} items).",
            remarks=remarks if remarks else None
        )
        
        # Log activity
        ActivityLog.log_activity(
            user=request.user,
            action='reject',
            model_name='ReservationBatch',
            object_repr=f"Batch #{batch_id}",
            description=f"Rejected {count} items in reservation batch"
        )
        
        messages.success(request, f'{count} item(s) rejected.')
        return redirect('reservation_batch_detail', batch_id=batch_id)
    
    return redirect('user_reservations')


def damage_report_detail(request, pk):
    report_obj = get_object_or_404(DamageReport, pk=pk)
    if request.method == 'POST':
        action = request.POST.get('action')
        remarks = request.POST.get('remarks')
        
        if action == 'resolve':
            classification = request.POST.get('classification')
            
            if classification == 'unserviceable':
                report_obj.status = 'resolved'
                report_obj.remarks = f"Classified as Unserviceable. {remarks}"
                # Update property condition
                report_obj.item.condition = 'Unserviceable'
                report_obj.item.availability = 'not_available'
                report_obj.item.save()
                report_obj.save()
                Notification.objects.create(
                    user=report_obj.user,
                    message=f"Your damage report for {report_obj.item.property_name} has been resolved - Classified as Unserviceable.",
                    remarks=remarks
                )
            elif classification == 'needs_repair':
                report_obj.status = 'resolved'
                report_obj.remarks = f"Classified as Needs Repair. {remarks}"
                # Update property condition
                report_obj.item.condition = 'Needing repair'
                report_obj.item.availability = 'not_available'
                report_obj.item.save()
                report_obj.save()
                Notification.objects.create(
                    user=report_obj.user,
                    message=f"Your damage report for {report_obj.item.property_name} has been resolved - Classified as Needs Repair.",
                    remarks=remarks
                )
            elif classification == 'good_condition':
                report_obj.status = 'resolved'
                report_obj.remarks = f"Classified as Good Condition. {remarks}"
                # Update property condition
                report_obj.item.condition = 'In good condition'
                report_obj.item.availability = 'available'
                report_obj.item.save()
                report_obj.save()
                Notification.objects.create(
                    user=report_obj.user,
                    message=f"Your damage report for {report_obj.item.property_name} has been resolved - Item is in Good Condition.",
                    remarks=remarks
                )
            
            messages.success(request, f'Damage report #{report_obj.id} has been classified and resolved successfully.')
            return redirect('user_damage_reports')
    
    return render(request, 'app/report_details.html', {'report_obj': report_obj})


def borrow_request_details(request, pk):
    request_obj = get_object_or_404(BorrowRequest, pk=pk)
    if request.method == 'POST':
        action = request.POST.get('action')
        remarks = request.POST.get('remarks')
        request_obj.remarks = remarks

        try:
            if action == 'approve':
                request_obj.status = 'approved'
                request_obj.approved_date = timezone.now()
                request_obj.save()  # This will trigger validation
                messages.success(request, 'Borrow request approved successfully.')
                Notification.objects.create(
                    user=request_obj.user,
                    message=f"Your borrow request for {request_obj.property.property_name} has been approved.",
                    remarks=remarks
                )
            elif action == 'decline':
                request_obj.status = 'declined'
                request_obj.approved_date = timezone.now()
                request_obj.save()
                messages.success(request, 'Borrow request declined successfully.')
                Notification.objects.create(
                    user=request_obj.user,
                    message=f"Your borrow request for {request_obj.property.property_name} has been declined.",
                    remarks=remarks
                )
            elif action == 'return':
                request_obj.status = 'returned'
                request_obj.actual_return_date = timezone.now().date()
                request_obj.save()
                messages.success(request, 'Borrow request marked as returned successfully.')
                Notification.objects.create(
                    user=request_obj.user,
                    message=f"Your borrow request for {request_obj.property.property_name} has been marked as returned.",
                    remarks=remarks
                )
            elif action == 'overdue':
                request_obj.status = 'overdue'
                request_obj.save()
                messages.warning(request, 'Borrow request marked as overdue.')
                Notification.objects.create(
                    user=request_obj.user,
                    message=f"Your borrow request for {request_obj.property.property_name} is overdue.",
                    remarks=remarks
                )
        except ValidationError as e:
            if 'quantity' in e.message_dict:
                messages.error(request, e.message_dict['quantity'][0])
            else:
                messages.error(request, "An error occurred while processing the request.")
            return render(request, 'app/borrow_request_details.html', {'borrow_obj': request_obj})
            
        return redirect('user_borrow_requests')
    return render(request, 'app/borrow_request_details.html', {'borrow_obj': request_obj})


def borrow_request_status_update(request, pk):
    request_obj = get_object_or_404(BorrowRequest, pk=pk)
    status = request.POST.get('status')
    remarks = request.POST.get('remarks')  # optional if you want to use it here

    if status == 'returned':
        request_obj.status = 'returned'
        request_obj.actual_return_date = timezone.now().date()
        Notification.objects.create(
            user=request_obj.user,
            message=f"Your borrow request for {request_obj.property.property_name} has been marked as returned.",
            remarks=remarks
        )
    elif status == 'overdue':
        request_obj.status = 'overdue'
        Notification.objects.create(
            user=request_obj.user,
            message=f"Your borrow request for {request_obj.property.property_name} is overdue.",
            remarks=remarks
        )
    request_obj.save()
    return redirect('user_borrow_requests')


def request_detail(request, pk):
    request_obj = get_object_or_404(SupplyRequest, pk=pk)
    if request.method == 'POST':
        action = request.POST.get('action')
        remarks = request.POST.get('remarks')
        request_obj.remarks = remarks
        if action == 'approve':
            request_obj.status = 'approved'
            Notification.objects.create(
                user=request_obj.user,
                message=f"Your supply request for {request_obj.supply.supply_name} has been approved.",
                remarks=remarks
            )
            
        elif action == 'rejected':
            request_obj.status = 'rejected'
            Notification.objects.create(
                user=request_obj.user,
                message=f"Your supply request for {request_obj.supply.supply_name} has been rejected.",
                remarks=remarks
            )
        request_obj.approved_date = timezone.now()
        request_obj.save()
        return redirect('user_supply_requests')
    return render(request, 'app/request_details.html', {'request_obj': request_obj})


class UserProfileListView(AdminPermissionMixin, PermissionRequiredMixin, ListView):
    model = UserProfile
    template_name = 'app/manage_users.html'
    permission_required = 'app.view_admin_module'
    required_permission = 'manage_users'
    context_object_name = 'users'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().select_related('user', 'department')
        
        # Get filter parameters from the URL
        search = self.request.GET.get('search', '')
        role = self.request.GET.get('role', '')
        department = self.request.GET.get('department', '')
        status = self.request.GET.get('status', '')
        
        if search:
            queryset = queryset.filter(
                Q(user__username__icontains=search) |
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__email__icontains=search) |
                Q(department__name__icontains=search)
            )
        
        if role:
            queryset = queryset.filter(role=role)
            
        if department:
            queryset = queryset.filter(department__name=department)
        
        if status:
            if status == 'active':
                queryset = queryset.filter(user__is_active=True)
            elif status == 'inactive':
                queryset = queryset.filter(user__is_active=False)
        
        # Order by account creation date (newest first) to fix pagination warning
        return queryset.order_by('-user__date_joined')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Check if there's form data in session (from validation errors)
        form_data = self.request.session.pop('form_data', None)
        form_errors = self.request.session.pop('form_errors', None)
        show_modal = self.request.session.pop('show_modal', False)
        
        if form_data:
            # Recreate form with previous data
            form = UserRegistrationForm(data=form_data)
            if form_errors:
                # Replace the form errors with stored errors
                from django.forms.utils import ErrorDict
                form._errors = ErrorDict()
                for field, errors in form_errors.items():
                    form._errors[field] = errors
        else:
            form = UserRegistrationForm()
        
        context['form'] = form
        context['departments'] = Department.objects.all()
        context['show_modal'] = show_modal
        context['now'] = timezone.now()
        # Add current filter values to context
        context['search'] = self.request.GET.get('search', '')
        context['selected_role'] = self.request.GET.get('role', '')
        context['selected_department'] = self.request.GET.get('department', '')
        context['selected_status'] = self.request.GET.get('status', '')
        return context


@login_required
@permission_required('app.view_admin_module')
def manage_admin_permissions(request, user_id):
    """
    View to manage admin permissions for a specific user
    Allows assigning/removing granular permissions for admin users
    """
    # Check if user has permission to manage permissions
    # Superusers and full-access admins can manage permissions
    can_manage = False
    if request.user.is_superuser:
        can_manage = True
    else:
        try:
            user_profile_check = request.user.userprofile
            if user_profile_check.role == 'ADMIN':
                # Full access admins (not limited) can manage permissions
                if not user_profile_check.has_limited_access:
                    can_manage = True
                # Limited access admins need the manage_users permission
                elif user_profile_check.admin_permissions.filter(codename='manage_users').exists():
                    can_manage = True
        except:
            pass
    
    if not can_manage:
        messages.error(request, "You don't have permission to manage user permissions.")
        return redirect('user_profile_list')
    
    user_profile = get_object_or_404(UserProfile, id=user_id)
    
    # Ensure target user is an admin
    if user_profile.role != 'ADMIN':
        messages.error(request, "Permissions can only be managed for admin users.")
        return redirect('user_profile_list')
    
    # Prevent modifying superuser permissions
    if user_profile.user.is_superuser:
        messages.error(request, "Cannot modify permissions for superuser accounts.")
        return redirect('user_profile_list')
    
    if request.method == 'POST':
        # Get the has_limited_access checkbox
        has_limited_access = request.POST.get('has_limited_access') == 'on'
        user_profile.has_limited_access = has_limited_access
        
        if has_limited_access:
            # Get selected permissions
            selected_permissions = request.POST.getlist('permissions')
            # Clear existing permissions
            user_profile.admin_permissions.clear()
            # Add selected permissions
            for perm_id in selected_permissions:
                try:
                    permission = AdminPermission.objects.get(id=perm_id)
                    user_profile.admin_permissions.add(permission)
                except AdminPermission.DoesNotExist:
                    pass
        else:
            # Full access - clear all specific permissions
            user_profile.admin_permissions.clear()
        
        user_profile.save()
        
        # Log the activity
        ActivityLog.log_activity(
            user=request.user,
            action='update',
            model_name='UserProfile',
            object_repr=user_profile.user.username,
            description=f"Updated admin permissions for {user_profile.user.username}"
        )
        
        messages.success(request, f"Permissions updated successfully for {user_profile.user.get_full_name() or user_profile.user.username}.")
        return redirect('user_profile_list')
    
    # GET request - display form
    all_permissions = AdminPermission.objects.all().order_by('name')
    user_permission_ids = user_profile.admin_permissions.values_list('id', flat=True)
    
    context = {
        'user_profile': user_profile,
        'all_permissions': all_permissions,
        'user_permission_ids': list(user_permission_ids),
    }
    
    return render(request, 'app/manage_admin_permissions.html', context)


@login_required
@permission_required('app.view_admin_module')
def initialize_admin_permissions(request):
    """
    Initialize default admin permissions
    This should be run once after adding the permission system
    """
    if not request.user.is_superuser:
        messages.error(request, "Only superusers can initialize permissions.")
        return redirect('dashboard')
    
    AdminPermission.initialize_permissions()
    
    messages.success(request, "Admin permissions have been initialized successfully.")
    return redirect('user_profile_list')


@login_required
@permission_required('app.view_admin_module')
def get_user_permissions(request, user_id):
    """
    AJAX endpoint to get user permissions data for the modal
    """
    # Check if user has permission to manage permissions
    can_manage = False
    if request.user.is_superuser:
        can_manage = True
    else:
        try:
            user_profile_check = request.user.userprofile
            if user_profile_check.role == 'ADMIN':
                if not user_profile_check.has_limited_access:
                    can_manage = True
                elif user_profile_check.admin_permissions.filter(codename='manage_users').exists():
                    can_manage = True
        except:
            pass
    
    if not can_manage:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        user_profile = UserProfile.objects.get(id=user_id)
        
        # Ensure target user is an admin
        if user_profile.role != 'ADMIN':
            return JsonResponse({'success': False, 'error': 'Permissions can only be managed for admin users'})
        
        # Prevent modifying superuser permissions
        if user_profile.user.is_superuser:
            return JsonResponse({'success': False, 'error': 'Cannot modify permissions for superuser accounts'})
        
        all_permissions = AdminPermission.objects.all().order_by('name')
        user_permission_ids = list(user_profile.admin_permissions.values_list('id', flat=True))
        
        data = {
            'success': True,
            'user_profile': {
                'username': user_profile.user.username,
                'full_name': user_profile.user.get_full_name(),
                'role': user_profile.get_role_display(),
                'department': user_profile.department.name if user_profile.department else None,
                'has_limited_access': user_profile.has_limited_access,
            },
            'all_permissions': [
                {
                    'id': perm.id,
                    'name': perm.name,
                    'codename': perm.codename,
                    'description': perm.description
                }
                for perm in all_permissions
            ],
            'user_permission_ids': user_permission_ids,
        }
        
        return JsonResponse(data)
    except UserProfile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@permission_required('app.view_admin_module')
@require_POST
def save_user_permissions(request, user_id):
    """
    AJAX endpoint to save user permissions from the modal
    """
    # Check if user has permission to manage permissions
    can_manage = False
    if request.user.is_superuser:
        can_manage = True
    else:
        try:
            user_profile_check = request.user.userprofile
            if user_profile_check.role == 'ADMIN':
                if not user_profile_check.has_limited_access:
                    can_manage = True
                elif user_profile_check.admin_permissions.filter(codename='manage_users').exists():
                    can_manage = True
        except:
            pass
    
    if not can_manage:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        user_profile = UserProfile.objects.get(id=user_id)
        
        # Ensure target user is an admin
        if user_profile.role != 'ADMIN':
            return JsonResponse({'success': False, 'error': 'Permissions can only be managed for admin users'})
        
        # Prevent modifying superuser permissions
        if user_profile.user.is_superuser:
            return JsonResponse({'success': False, 'error': 'Cannot modify permissions for superuser accounts'})
        
        # Get the has_limited_access value
        has_limited_access = request.POST.get('has_limited_access') == 'on'
        user_profile.has_limited_access = has_limited_access
        
        if has_limited_access:
            # Get selected permissions
            selected_permissions = request.POST.getlist('permissions')
            # Clear existing permissions
            user_profile.admin_permissions.clear()
            # Add selected permissions
            for perm_id in selected_permissions:
                try:
                    permission = AdminPermission.objects.get(id=perm_id)
                    user_profile.admin_permissions.add(permission)
                except AdminPermission.DoesNotExist:
                    pass
        else:
            # Full access - clear all specific permissions
            user_profile.admin_permissions.clear()
        
        user_profile.save()
        
        # Log the activity
        ActivityLog.log_activity(
            user=request.user,
            action='update',
            model_name='UserProfile',
            object_repr=user_profile.user.username,
            description=f"Updated admin permissions for {user_profile.user.username}"
        )
        
        messages.success(request, f"Permissions updated successfully for {user_profile.user.get_full_name() or user_profile.user.username}.")
        
        return JsonResponse({'success': True})
    except UserProfile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@permission_required('app.view_admin_module')
@login_required
def get_top_requested_supplies(request):
    """
    API endpoint to get top 10 most requested supplies with optional date and department filtering
    """
    try:
        # Get filter parameters from request
        days = request.GET.get('days')  # 30, 90, etc.
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        department_id = request.GET.get('department')  # department filter
        
        # Build filter
        filter_kwargs = {}
        if days:
            try:
                days = int(days)
                cutoff_date = timezone.now() - timedelta(days=days)
                filter_kwargs['request_date__gte'] = cutoff_date
            except ValueError:
                pass
        elif date_from and date_to:
            try:
                from datetime import datetime
                start_dt = datetime.strptime(date_from, '%Y-%m-%d')
                end_dt = datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                filter_kwargs['request_date__range'] = [start_dt, end_dt]
            except ValueError:
                pass
        
        # Add department filter if specified
        if department_id:
            filter_kwargs['user__userprofile__department_id'] = department_id
        
        # Get supply requests from both legacy and batch systems
        legacy_requests = SupplyRequest.objects.filter(**filter_kwargs).values('supply_id', 'supply__supply_name').annotate(
            total_quantity=models.Sum('quantity'),
            request_count=models.Count('id')
        ).order_by('-total_quantity')[:10]
        
        # Build batch items filter
        batch_filter = {
            'batch_request__status__in': ['approved', 'completed', 'for_claiming'],
            'batch_request__request_date__gte': filter_kwargs.get('request_date__gte', timezone.now() - timedelta(days=999))
        }
        
        # Add date range if specified
        if 'request_date__range' in filter_kwargs:
            batch_filter['batch_request__request_date__range'] = filter_kwargs['request_date__range']
        
        # Add department filter if specified
        if department_id:
            batch_filter['batch_request__user__userprofile__department_id'] = department_id
        
        batch_items = SupplyRequestItem.objects.filter(**batch_filter).values('supply_id', 'supply__supply_name').annotate(
            total_quantity=models.Sum('quantity'),
            request_count=models.Count('id')
        ).order_by('-total_quantity')[:10]
        
        # Combine and deduplicate
        combined_data = {}
        for item in legacy_requests:
            combined_data[item['supply_id']] = {
                'supply_id': item['supply_id'],
                'supply_name': item['supply__supply_name'],
                'total_quantity': item['total_quantity'],
                'request_count': item['request_count']
            }
        
        for item in batch_items:
            if item['supply_id'] in combined_data:
                combined_data[item['supply_id']]['total_quantity'] += item['total_quantity']
                combined_data[item['supply_id']]['request_count'] += item['request_count']
            else:
                combined_data[item['supply_id']] = {
                    'supply_id': item['supply_id'],
                    'supply_name': item['supply__supply_name'],
                    'total_quantity': item['total_quantity'],
                    'request_count': item['request_count']
                }
        
        # Sort and get top 10
        sorted_data = sorted(combined_data.values(), key=lambda x: x['total_quantity'], reverse=True)[:10]
        
        return JsonResponse({
            'success': True,
            'data': sorted_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@permission_required('app.view_admin_module')
@login_required
def get_department_requests_filtered(request):
    """
    API endpoint to get department requests with optional date filtering
    """
    try:
        # Get filter parameters from request
        days = request.GET.get('days')  # 30, 90, etc.
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        # Build filter
        filter_kwargs = {}
        if days:
            try:
                days = int(days)
                cutoff_date = timezone.now() - timedelta(days=days)
                filter_kwargs['request_date__gte'] = cutoff_date
            except ValueError:
                pass
        elif date_from and date_to:
            try:
                from datetime import datetime
                start_dt = datetime.strptime(date_from, '%Y-%m-%d')
                end_dt = datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                filter_kwargs['request_date__range'] = [start_dt, end_dt]
            except ValueError:
                pass
        
        # Get requests by department
        departments = Department.objects.all()
        department_request_data = []
        
        for department in departments:
            # Count supply requests
            legacy_supply_count = SupplyRequest.objects.filter(
                user__userprofile__department=department,
                **filter_kwargs
            ).count()
            batch_supply_count = SupplyRequestBatch.objects.filter(
                user__userprofile__department=department,
                **filter_kwargs
            ).count()
            total_supply_requests = legacy_supply_count + batch_supply_count
            
            # Count borrow requests
            borrow_count = BorrowRequest.objects.filter(
                user__userprofile__department=department,
                **({} if not filter_kwargs else {'borrow_date__gte': filter_kwargs.get('request_date__gte', timezone.now() - timedelta(days=999))})
            ).count()
            
            # Count reservations
            reservation_count = Reservation.objects.filter(
                user__userprofile__department=department,
                **({} if not filter_kwargs else {'reservation_date__gte': filter_kwargs.get('request_date__gte', timezone.now() - timedelta(days=999))})
            ).count()
            
            total_requests = total_supply_requests + borrow_count + reservation_count
            
            if total_requests > 0:
                department_request_data.append({
                    'department': department.name,
                    'total_requests': total_requests,
                    'supply_requests': total_supply_requests,
                    'borrow_requests': borrow_count,
                    'reservations': reservation_count
                })
        
        # Sort by total requests
        department_request_data.sort(key=lambda x: x['total_requests'], reverse=True)
        
        return JsonResponse({
            'success': True,
            'data': department_request_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


class DashboardPageView(PermissionRequiredMixin,TemplateView):
    template_name = 'app/dashboard.html'
    permission_required = 'app.view_admin_module'  

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        

        # Notifications
        user_notifications = Notification.objects.filter(user=self.request.user).order_by('-timestamp')
        unread_notifications = user_notifications.filter(is_read=False)

        # Calculate date ranges for percentage changes
        today = timezone.now().date()
        first_day_this_month = today.replace(day=1)
        last_day_last_month = first_day_this_month - timedelta(days=1)
        first_day_last_month = last_day_last_month.replace(day=1)

        #expiry count
        seven_days_later = today + timedelta(days=30) #30 days before the expiry date para ma trigger
        near_expiry_count = Supply.objects.filter(
            expiration_date__range=(today, seven_days_later),
            quantity_info__current_quantity__gt=0
        ).count()
        context['near_expiry_count'] = near_expiry_count
        
        # Calculate near expiry count for last month
        near_expiry_count_last_month = Supply.objects.filter(
            expiration_date__range=(first_day_last_month, last_day_last_month),
            quantity_info__current_quantity__gt=0
        ).count()

        # Supply Status Counts
        supply_status_counts = [
            Supply.objects.filter(
                quantity_info__current_quantity__gt=models.F('quantity_info__minimum_threshold')
            ).count(),  # available
            Supply.objects.filter(
                quantity_info__current_quantity__gt=0,
                quantity_info__current_quantity__lte=models.F('quantity_info__minimum_threshold')
            ).count(),  # low_stock
            Supply.objects.filter(
                quantity_info__current_quantity=0
            ).count(),  # out_of_stock
        ]

        # Property Condition Counts
        property_condition_choices = [
            'In good condition',
            'Needing repair',
            'Unserviceable',
            'Obsolete',
            'No longer needed',
            'Not used since purchased',
        ]
        property_condition_counts = [
            Property.objects.filter(condition=condition).count()
            for condition in property_condition_choices
        ]

        # Request Status Counts (SupplyRequest + SupplyRequestBatch)
        request_status_choices = ['pending', 'approved', 'rejected', 'partially_approved', 'for_claiming']
        request_status_counts = []
        for status in request_status_choices:
            legacy_count = SupplyRequest.objects.filter(status__iexact=status).count()
            batch_count = SupplyRequestBatch.objects.filter(status__iexact=status).count()
            request_status_counts.append(legacy_count + batch_count)

        # Damage Report Status Counts
        damage_status_choices = ['pending', 'reviewed', 'resolved']
        damage_status_counts = [
            DamageReport.objects.filter(status__iexact=status).count()
            for status in damage_status_choices
        ]

        # Borrow Request Trends (last 6 months)
        now_date = now().date()
        borrow_trends_data = []

        for i in reversed(range(6)):
            start_date = now_date.replace(day=1) - timedelta(days=30*i)
            if i == 0:
                end_date = now_date
            else:
                end_date = now_date.replace(day=1) - timedelta(days=30*(i-1))

            # Count old borrow requests
            old_count = BorrowRequest.objects.filter(
                borrow_date__date__gte=start_date,
                borrow_date__date__lt=end_date
            ).count()
            
            # Count new batch borrow requests
            batch_count = BorrowRequestBatch.objects.filter(
                request_date__date__gte=start_date,
                request_date__date__lt=end_date
            ).count()
            
            # Total count combines both systems
            count = old_count + batch_count

            month_str = start_date.strftime('%b %Y')

            borrow_trends_data.append({
                'month': month_str,
                'count': count
            })

        # Property Categories Counts
        try:
            # Get all categories from PropertyCategory model
            categories = PropertyCategory.objects.all()
            property_categories_data = []
            for category in categories:
                count = Property.objects.filter(category=category).count()
                if count > 0:
                    property_categories_data.append({
                        'category': category.name,
                        'count': count
                    })
        except Exception as e:
            logger.error(f"Error getting property categories data: {str(e)}")
            property_categories_data = []

        # User Activity by Role
        try:
            user_roles = [choice[0] for choice in UserProfile.ROLE_CHOICES]
            user_activity_data = []
            for role in user_roles:
                try:
                    count = ActivityLog.objects.filter(
                        user__userprofile__role=role
                    ).count()
                except:
                    count = UserProfile.objects.filter(role=role).count()

                if count > 0:
                    user_activity_data.append({
                        'role': role.replace('_', ' ').title(),
                        'count': count
                    })
        except (AttributeError, NameError):
            user_activity_data = []

        # Department Request Analysis
        try:
            # Get requests by department for both SupplyRequest and BorrowRequest
            departments = Department.objects.all()
            department_request_data = []
            
            for department in departments:
                # Count supply requests from users in this department (legacy system)
                legacy_supply_request_count = SupplyRequest.objects.filter(
                    user__userprofile__department=department
                ).count()
                
                # Count batch supply requests from users in this department (new system)
                batch_supply_request_count = SupplyRequestBatch.objects.filter(
                    user__userprofile__department=department
                ).count()
                
                # Total supply requests (legacy + batch)
                total_supply_requests = legacy_supply_request_count + batch_supply_request_count
                
                # Count borrow requests from users in this department
                borrow_request_count = BorrowRequest.objects.filter(
                    user__userprofile__department=department
                ).count()
                
                # Count reservations from users in this department
                reservation_count = Reservation.objects.filter(
                    user__userprofile__department=department
                ).count()
                
                # Total requests from this department
                total_requests = total_supply_requests + borrow_request_count + reservation_count
                
                if total_requests > 0:
                    department_request_data.append({
                        'department': department.name,
                        'total_requests': total_requests,
                        'supply_requests': total_supply_requests,
                        'borrow_requests': borrow_request_count,
                        'reservations': reservation_count
                    })
            
            # Sort by total requests descending
            department_request_data.sort(key=lambda x: x['total_requests'], reverse=True)
            
        except Exception as e:
            logger.error(f"Error getting department request data: {str(e)}")
            department_request_data = []

        # Recent Requests for Preview Table
        recent_supply_requests = SupplyRequest.objects.select_related(
            'user', 'supply'
        ).order_by('-request_date')[:5]

        recent_batch_requests = SupplyRequestBatch.objects.select_related(
            'user'
        ).order_by('-request_date')[:5]

        recent_borrow_requests = BorrowRequest.objects.select_related(
            'user', 'property'
        ).order_by('-borrow_date')[:5]

        recent_reservations = Reservation.objects.select_related(
            'user', 'item'
        ).order_by('-reservation_date')[:5]

        all_recent_requests = []
        
        # Add legacy supply requests
        for req in recent_supply_requests:
            all_recent_requests.append({
                'type': 'Supply Request',
                'user': req.user.username,
                'item': req.supply.supply_name,
                'quantity': req.quantity,
                'status': req.get_status_display(),
                'date': req.request_date,
                'purpose': req.purpose[:50] + '...' if len(req.purpose) > 50 else req.purpose
            })
        
        # Add batch supply requests
        for req in recent_batch_requests:
            items_text = f"{req.total_items} items"
            if req.total_items <= 3:
                try:
                    items_list = ", ".join([f"{item.supply.supply_name} (x{item.quantity})" for item in req.items.all()])
                    items_text = items_list
                except:
                    items_text = f"{req.total_items} items"
            
            all_recent_requests.append({
                'type': 'Batch Supply Request',
                'user': req.user.username,
                'item': items_text,
                'quantity': req.total_quantity,
                'status': req.get_status_display(),
                'date': req.request_date,
                'purpose': req.purpose[:50] + '...' if len(req.purpose) > 50 else req.purpose
            })
        
        for req in recent_borrow_requests:
            all_recent_requests.append({
                'type': 'Borrow Request',
                'user': req.user.username,
                'item': req.property.property_name,
                'quantity': req.quantity,
                'status': req.get_status_display(),
                'date': req.borrow_date,
                'purpose': req.purpose[:50] + '...' if len(req.purpose) > 50 else req.purpose
            })
        
        for req in recent_reservations:
            all_recent_requests.append({
                'type': 'Reservation',
                'user': req.user.username,
                'item': req.item.property_name,
                'quantity': req.quantity,
                'status': req.get_status_display(),
                'date': req.reservation_date,
                'purpose': req.purpose[:50] + '...' if len(req.purpose) > 50 else req.purpose
            })
            
        all_recent_requests.sort(key=lambda x: x['date'], reverse=True)
        recent_requests_preview = all_recent_requests[:5]

        context.update({
            # Notifications
            'notifications': user_notifications,
            'unread_count': unread_notifications.count(),

            # supply status summary
            'supply_available': supply_status_counts[0],
            'supply_low_stock': supply_status_counts[1],
            'supply_out_of_stock': supply_status_counts[2],

            # property condition summary
            'property_in_good_condition': property_condition_counts[0],
            'property_needing_repair': property_condition_counts[1],
            'property_unserviceable': property_condition_counts[2],
            'property_obsolete': property_condition_counts[3],
            'property_no_longer_needed': property_condition_counts[4],
            'property_not_used_since_purchased': property_condition_counts[5],

            # request status summary
            'request_status_pending': request_status_counts[0],
            'request_status_approved': request_status_counts[1],
            'request_status_rejected': request_status_counts[2],

            # damage status summary
            'damage_status_pending': damage_status_counts[0],
            'damage_status_reviewed': damage_status_counts[1],
            'damage_status_resolved': damage_status_counts[2],

            # JSON serialized data for charts
            'borrow_trends_data': json.dumps(borrow_trends_data),
            'user_activity_by_role': json.dumps(user_activity_data),
            'department_requests_data': json.dumps(department_request_data),

            # total counts for cards
            'supply_count': Supply.objects.count(),
            'property_count': Property.objects.count(),
            'pending_supply_requests': SupplyRequest.objects.filter(status__iexact='pending').count() + SupplyRequestBatch.objects.filter(status__iexact='pending').count(),
            'pending_borrow_requests': BorrowRequest.objects.filter(status__iexact='pending').count() + BorrowRequestBatch.objects.filter(status__iexact='pending').count(),
            'pending_requests': SupplyRequest.objects.filter(status__iexact='pending').count() + SupplyRequestBatch.objects.filter(status__iexact='pending').count() + BorrowRequest.objects.filter(status__iexact='pending').count() + BorrowRequestBatch.objects.filter(status__iexact='pending').count(),
            'damage_reports': DamageReport.objects.filter(status__iexact='pending').count(),
            
            # Recent requests for preview table
            'recent_requests_preview': recent_requests_preview,
            
            # Departments for filters
            'departments': Department.objects.all().order_by('name'),
        })
        
        # Calculate percentage changes for dashboard cards
        # Helper function to calculate percentage change
        def calculate_percentage_change(current, previous):
            if previous == 0:
                return {'percentage': 0, 'direction': 'neutral'} if current == 0 else {'percentage': 100, 'direction': 'positive'}
            change = ((current - previous) / previous) * 100
            return {
                'percentage': abs(round(change, 1)),
                'direction': 'positive' if change > 0 else 'negative' if change < 0 else 'neutral'
            }
        
        # Count items created/added last month for comparison
        supply_count_current = context['supply_count']
        supply_count_last_month = Supply.objects.filter(
            created_at__lt=first_day_this_month
        ).count() if hasattr(Supply, 'created_at') else supply_count_current
        
        property_count_current = context['property_count']
        property_count_last_month = Property.objects.filter(
            created_at__lt=first_day_this_month
        ).count() if hasattr(Property, 'created_at') else property_count_current
        
        pending_requests_current = context['pending_requests']
        pending_requests_last_month = (
            SupplyRequest.objects.filter(
                status__iexact='pending',
                request_date__gte=first_day_last_month,
                request_date__lt=first_day_this_month
            ).count() +
            SupplyRequestBatch.objects.filter(
                status__iexact='pending',
                request_date__gte=first_day_last_month,
                request_date__lt=first_day_this_month
            ).count() +
            BorrowRequest.objects.filter(
                status__iexact='pending',
                borrow_date__gte=first_day_last_month,
                borrow_date__lt=first_day_this_month
            ).count() +
            BorrowRequestBatch.objects.filter(
                status__iexact='pending',
                request_date__gte=first_day_last_month,
                request_date__lt=first_day_this_month
            ).count()
        )
        
        damage_reports_current = context['damage_reports']
        damage_reports_last_month = DamageReport.objects.filter(
            status__iexact='pending',
            report_date__gte=first_day_last_month,
            report_date__lt=first_day_this_month
        ).count()
        
        # Add percentage changes to context
        context.update({
            'supply_change': calculate_percentage_change(supply_count_current, supply_count_last_month),
            'property_change': calculate_percentage_change(property_count_current, property_count_last_month),
            'pending_requests_change': calculate_percentage_change(pending_requests_current, pending_requests_last_month),
            'damage_reports_change': calculate_percentage_change(damage_reports_current, damage_reports_last_month),
            'near_expiry_change': calculate_percentage_change(near_expiry_count, near_expiry_count_last_month),
        })

        return context


@login_required
def get_near_expiry_count(request):
    """API endpoint to get near expiry count based on days filter"""
    from django.http import JsonResponse
    
    try:
        days = int(request.GET.get('days', 30))
        today = timezone.now().date()
        end_date = today + timedelta(days=days)
        
        count = Supply.objects.filter(
            expiration_date__range=(today, end_date),
            quantity_info__current_quantity__gt=0
        ).count()
        
        return JsonResponse({'count': count})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def get_pending_requests_count(request):
    """API endpoint to get pending requests count by type (all, supply, borrow)"""
    from django.http import JsonResponse
    
    try:
        request_type = request.GET.get('type', 'all')
        
        if request_type == 'supply':
            count = (
                SupplyRequest.objects.filter(status__iexact='pending').count() +
                SupplyRequestBatch.objects.filter(status__iexact='pending').count()
            )
        elif request_type == 'borrow':
            count = (
                BorrowRequest.objects.filter(status__iexact='pending').count() +
                BorrowRequestBatch.objects.filter(status__iexact='pending').count()
            )
        else:  # all
            count = (
                SupplyRequest.objects.filter(status__iexact='pending').count() +
                SupplyRequestBatch.objects.filter(status__iexact='pending').count() +
                BorrowRequest.objects.filter(status__iexact='pending').count() +
                BorrowRequestBatch.objects.filter(status__iexact='pending').count()
            )
        
        return JsonResponse({'count': count})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


class ActivityPageView(PermissionRequiredMixin, ListView):
    model = ActivityLog
    template_name = 'app/activity.html'
    permission_required = 'app.view_admin_module'
    permission_denied_message = "You do not have permission to view the activity log."
    context_object_name = 'activitylog_list'
    paginate_by = 20  # Increased from 10 to 20 for better performance (fewer page loads)

    def get_queryset(self):
        # Use select_related to prevent N+1 queries on user lookups
        queryset = ActivityLog.objects.select_related('user').order_by('-timestamp')
        
        # Apply filters
        user_filter = self.request.GET.get('user')
        model_filter = self.request.GET.get('model')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        if user_filter:
            queryset = queryset.filter(user_id=user_filter)
        if model_filter:
            queryset = queryset.filter(model_name=model_filter)
        if start_date:
            try:
                from datetime import datetime
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                start_date_aware = timezone.make_aware(start_date_obj.replace(hour=0, minute=0, second=0))
                queryset = queryset.filter(timestamp__gte=start_date_aware)
            except (ValueError, TypeError):
                pass
        if end_date:
            try:
                from datetime import datetime
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                end_date_aware = timezone.make_aware(end_date_obj.replace(hour=23, minute=59, second=59))
                queryset = queryset.filter(timestamp__lte=end_date_aware)
            except (ValueError, TypeError):
                pass

        # Filter by category if specified (duplicate of model_filter, kept for compatibility)
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(model_name=category)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Load all users who have activity logs (including inactive users)
        # This is important for audit trails - we need to show past activities of deactivated users
        context['users'] = User.objects.filter(
            activitylog__isnull=False
        ).distinct().order_by('username')
        
        # Get all unique models efficiently using database-level distinct
        models = ActivityLog.objects.values_list('model_name', flat=True).distinct().order_by('model_name')
        context['models'] = list(models)
        
        # Get current category for highlighting in template
        context['current_category'] = self.request.GET.get('category', '')
        
        return context



class UserSupplyRequestListView(PermissionRequiredMixin, ListView):
    model = SupplyRequestBatch
    template_name = 'app/requests.html'
    permission_required = 'app.view_admin_module'
    context_object_name = 'batch_requests'
    paginate_by = 8  # Optimized for table layout - shows 8 requests per page for better UX
    
    def get_queryset(self):
        # Show batch requests ordered by oldest first
        return SupplyRequestBatch.objects.select_related('user', 'user__userprofile', 'user__userprofile__department').prefetch_related('items__supply').order_by('request_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get the current tab from request
        current_tab = self.request.GET.get('tab', 'pending')
        
        # Get search and filter parameters
        search_query = self.request.GET.get('search', '').strip()
        department_filter = self.request.GET.get('department', '')
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')
        
        # Base queryset with related data
        base_queryset = SupplyRequestBatch.objects.select_related('user', 'user__userprofile', 'user__userprofile__department').prefetch_related('items__supply').order_by('request_date')
        
        # Apply search filter
        if search_query:
            # Search only by supply request ID/number
            try:
                # Try to extract just the number from the search query
                search_number = ''.join(filter(str.isdigit, search_query))
                if search_number:
                    base_queryset = base_queryset.filter(id=int(search_number))
                else:
                    # If no number found, return empty queryset
                    base_queryset = base_queryset.none()
            except (ValueError, TypeError):
                # If conversion fails, return empty queryset
                base_queryset = base_queryset.none()
        
        # Apply department filter
        if department_filter:
            base_queryset = base_queryset.filter(user__userprofile__department__id=department_filter)
        
        # Apply date range filters
        if date_from:
            try:
                from django.utils import timezone
                from datetime import datetime
                # Convert string to datetime and make it timezone aware
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                date_from_aware = timezone.make_aware(date_from_obj)
                base_queryset = base_queryset.filter(request_date__date__gte=date_from_aware.date())
            except ValueError:
                # Invalid date format, ignore filter
                pass
        if date_to:
            try:
                from django.utils import timezone
                from datetime import datetime
                # Convert string to datetime and make it timezone aware
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                date_to_aware = timezone.make_aware(date_to_obj)
                base_queryset = base_queryset.filter(request_date__date__lte=date_to_aware.date())
            except ValueError:
                # Invalid date format, ignore filter
                pass
        
        # Apply pagination to each status filter
        from django.core.paginator import Paginator
        from django.db.models import Q
        
        # Merge pending and partially_approved into pending
        pending_requests = base_queryset.filter(Q(status='pending') | Q(status='partially_approved'))
        approved_requests = base_queryset.filter(status='approved')
        rejected_requests = base_queryset.filter(status='rejected')
        for_claiming_requests = base_queryset.filter(status='for_claiming')
        # Completed requests ordered by completion date (most recent first)
        completed_requests = base_queryset.filter(status='completed').order_by('-completed_date')
        voided_requests = base_queryset.filter(status='voided').order_by('-request_date')
        
        # Paginate each category
        pending_paginator = Paginator(pending_requests, self.paginate_by)
        approved_paginator = Paginator(approved_requests, self.paginate_by)
        rejected_paginator = Paginator(rejected_requests, self.paginate_by)
        for_claiming_paginator = Paginator(for_claiming_requests, self.paginate_by)
        completed_paginator = Paginator(completed_requests, self.paginate_by)
        voided_paginator = Paginator(voided_requests, self.paginate_by)
        
        # Get current page number for each tab
        pending_page = self.request.GET.get('pending_page', 1)
        approved_page = self.request.GET.get('approved_page', 1)
        rejected_page = self.request.GET.get('rejected_page', 1)
        for_claiming_page = self.request.GET.get('for_claiming_page', 1)
        completed_page = self.request.GET.get('completed_page', 1)
        voided_page = self.request.GET.get('voided_page', 1)
        
        # Get the page objects
        context['pending_batch_requests'] = pending_paginator.get_page(pending_page)
        context['approved_batch_requests'] = approved_paginator.get_page(approved_page)
        context['rejected_batch_requests'] = rejected_paginator.get_page(rejected_page)
        context['for_claiming_batch_requests'] = for_claiming_paginator.get_page(for_claiming_page)
        context['completed_batch_requests'] = completed_paginator.get_page(completed_page)
        context['voided_batch_requests'] = voided_paginator.get_page(voided_page)
        
        # Add current tab to context
        context['current_tab'] = current_tab
        
        # Add search and filter parameters to context
        context['search_query'] = search_query
        context['department_filter'] = department_filter
        context['date_from'] = date_from
        context['date_to'] = date_to
        
        # Build URL parameters string for pagination links
        url_params = []
        if search_query:
            url_params.append(f'search={search_query}')
        if department_filter:
            url_params.append(f'department={department_filter}')
        if date_from:
            url_params.append(f'date_from={date_from}')
        if date_to:
            url_params.append(f'date_to={date_to}')
        
        base_url_params = '&'.join(url_params)
        context['url_params'] = '&' + base_url_params if base_url_params else ''
        
        # Get all departments for the filter dropdown
        from .models import Department
        context['departments'] = Department.objects.all().order_by('name')
        
        # Get all users who have made supply requests for the filter dropdown
        from django.contrib.auth.models import User
        try:
            users_with_requests = User.objects.filter(
                supplyrequestbatch__isnull=False
            ).distinct().order_by('first_name', 'last_name')
            context['users_with_requests'] = users_with_requests
        except Exception as e:
            # Fallback to all active users if there's an issue
            context['users_with_requests'] = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
        
        # Create a mapping of users to their departments for frontend filtering
        users_departments = {}
        for user in context['users_with_requests']:
            try:
                dept = user.userprofile.department
                users_departments[str(user.id)] = str(dept.id) if dept else ''
            except:
                users_departments[str(user.id)] = ''
        context['users_departments'] = users_departments
        
        return context



class UserDamageReportListView(PermissionRequiredMixin, ListView):
    model = DamageReport
    template_name = 'app/reports.html'
    permission_required = 'app.view_admin_module'
    context_object_name = 'damage_reports'
    paginate_by = 15  # Changed from 10 to 15
    
    def get_queryset(self):
        queryset = DamageReport.objects.select_related('user', 'user__userprofile', 'item').order_by('-report_date')
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__username__icontains=search) |
                Q(item__property_name__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Department filter
        department = self.request.GET.get('department')
        if department:
            queryset = queryset.filter(user__userprofile__department_id=department)
        
        # Date range filters
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            try:
                from datetime import datetime
                # Parse the date string and make it timezone-aware
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                date_from_aware = timezone.make_aware(date_from_obj.replace(hour=0, minute=0, second=0))
                queryset = queryset.filter(report_date__gte=date_from_aware)
            except (ValueError, TypeError):
                pass  # Invalid date format, skip filter
        if date_to:
            try:
                from datetime import datetime
                # Parse the date string and make it timezone-aware (end of day)
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                date_to_aware = timezone.make_aware(date_to_obj.replace(hour=23, minute=59, second=59))
                queryset = queryset.filter(report_date__lte=date_to_aware)
            except (ValueError, TypeError):
                pass  # Invalid date format, skip filter
        
        return queryset
    
    def get_context_data(self, **kwargs):
        from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
        context = super().get_context_data(**kwargs)
        all_reports = self.get_queryset()

        # Get URL parameters for building pagination links
        url_params = ""
        search = self.request.GET.get('search', '')
        department = self.request.GET.get('department', '')
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')
        
        params = []
        if search:
            params.append(f'search={search}')
        if department:
            params.append(f'department={department}')
        if date_from:
            params.append(f'date_from={date_from}')
        if date_to:
            params.append(f'date_to={date_to}')
        
        if params:
            url_params = "&" + "&".join(params)

        # Paginate each tab separately with 15 items per page
        paginate_by = 15
        
        # Get page numbers for each tab
        all_page = self.request.GET.get('all_page', 1)
        pending_page = self.request.GET.get('pending_page', 1)
        resolved_page = self.request.GET.get('resolved_page', 1)
        
        # Create paginators for each tab
        # Pending reports: oldest first (ascending order)
        pending_reports = all_reports.filter(status='pending').order_by('report_date')
        # Resolved reports: newest first (descending order)
        resolved_reports = all_reports.filter(status='resolved')
        
        # Paginate pending reports
        pending_paginator = Paginator(pending_reports, paginate_by)
        try:
            pending_page_obj = pending_paginator.page(pending_page)
        except PageNotAnInteger:
            pending_page_obj = pending_paginator.page(1)
        except EmptyPage:
            pending_page_obj = pending_paginator.page(pending_paginator.num_pages)
        
        # Paginate resolved reports
        resolved_paginator = Paginator(resolved_reports, paginate_by)
        try:
            resolved_page_obj = resolved_paginator.page(resolved_page)
        except PageNotAnInteger:
            resolved_page_obj = resolved_paginator.page(1)
        except EmptyPage:
            resolved_page_obj = resolved_paginator.page(resolved_paginator.num_pages)
        
        # Create combined all reports (mix of all statuses) and paginate
        all_reports_combined = all_reports
        all_paginator = Paginator(all_reports_combined, paginate_by)
        try:
            all_page_obj = all_paginator.page(all_page)
        except PageNotAnInteger:
            all_page_obj = all_paginator.page(1)
        except EmptyPage:
            all_page_obj = all_paginator.page(all_paginator.num_pages)

        # Add paginated objects to context
        context['pending_reports'] = pending_page_obj
        context['resolved_reports'] = resolved_page_obj
        context['all_reports'] = all_page_obj
        
        # Add total counts for badges
        context['pending_count'] = pending_reports.count()
        context['resolved_count'] = resolved_reports.count()
        context['all_count'] = all_reports_combined.count()
        
        # Add departments for filter dropdown
        from .models import Department
        context['departments'] = Department.objects.all().order_by('name')
        
        # Add search parameters for form persistence and URL building
        context['search'] = search
        context['department_filter'] = department
        context['date_from'] = date_from
        context['date_to'] = date_to
        context['url_params'] = url_params

        return context


class UserReservationListView(PermissionRequiredMixin, ListView):
    model = ReservationBatch
    template_name = 'app/reservation.html'
    permission_required = 'app.view_admin_module'
    context_object_name = 'reservation_batches'
    paginate_by = 10

    def get_queryset(self):
        # Get all reservation batches
        queryset = ReservationBatch.objects.select_related(
            'user'
        ).prefetch_related(
            'user__userprofile',
            'items__property'
        ).order_by('-request_date')
        
        # Apply search filter
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(user__username__icontains=search_query) |
                Q(items__property__property_name__icontains=search_query) |
                Q(purpose__icontains=search_query) |
                Q(id__icontains=search_query)
            ).distinct()
        
        # Apply department filter
        department = self.request.GET.get('department')
        if department:
            queryset = queryset.filter(user__userprofile__department_id=department)
        
        # Apply date filters
        date_from = self.request.GET.get('date_from')
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(request_date__date__gte=date_from_obj)
            except ValueError:
                pass
                
        date_to = self.request.GET.get('date_to')
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(request_date__date__lte=date_to_obj)
            except ValueError:
                pass
        
        return queryset

    def get_context_data(self, **kwargs):
        from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
        
        # Update reservation batch statuses before displaying
        ReservationBatch.check_and_update_batches()
        
        context = super().get_context_data(**kwargs)
        all_batches = self.get_queryset()
        
        # Group batches by status
        # Pending batches should show oldest first
        pending_batches = all_batches.filter(status='pending').order_by('request_date')
        approved_batches = all_batches.filter(status='approved')
        active_batches = all_batches.filter(status='active')
        completed_batches = all_batches.filter(status='completed')
        rejected_batches = all_batches.filter(status='rejected')
        expired_batches = all_batches.filter(status='expired')
        voided_batches = all_batches.filter(status='voided').order_by('-request_date')
        
        # Paginate each status group with 10 items per page
        paginate_by = 10
        
        # Get page numbers for each status tab
        all_page = self.request.GET.get('all_page', 1)
        pending_page = self.request.GET.get('pending_page', 1)
        approved_page = self.request.GET.get('approved_page', 1)
        active_page = self.request.GET.get('active_page', 1)
        completed_page = self.request.GET.get('completed_page', 1)
        rejected_page = self.request.GET.get('rejected_page', 1)
        expired_page = self.request.GET.get('expired_page', 1)
        voided_page = self.request.GET.get('voided_page', 1)
        
        # Paginate all batches
        all_paginator = Paginator(all_batches, paginate_by)
        try:
            context['all_reservations'] = all_paginator.page(all_page)
        except PageNotAnInteger:
            context['all_reservations'] = all_paginator.page(1)
        except EmptyPage:
            context['all_reservations'] = all_paginator.page(all_paginator.num_pages)
        
        # Paginate pending batches
        pending_paginator = Paginator(pending_batches, paginate_by)
        try:
            context['pending_reservations'] = pending_paginator.page(pending_page)
        except PageNotAnInteger:
            context['pending_reservations'] = pending_paginator.page(1)
        except EmptyPage:
            context['pending_reservations'] = pending_paginator.page(pending_paginator.num_pages)
        
        # Paginate approved batches
        approved_paginator = Paginator(approved_batches, paginate_by)
        try:
            context['approved_reservations'] = approved_paginator.page(approved_page)
        except PageNotAnInteger:
            context['approved_reservations'] = approved_paginator.page(1)
        except EmptyPage:
            context['approved_reservations'] = approved_paginator.page(approved_paginator.num_pages)
        
        # Paginate active batches
        active_paginator = Paginator(active_batches, paginate_by)
        try:
            context['active_reservations'] = active_paginator.page(active_page)
        except PageNotAnInteger:
            context['active_reservations'] = active_paginator.page(1)
        except EmptyPage:
            context['active_reservations'] = active_paginator.page(active_paginator.num_pages)
        
        # Paginate completed batches
        completed_paginator = Paginator(completed_batches, paginate_by)
        try:
            context['completed_reservations'] = completed_paginator.page(completed_page)
        except PageNotAnInteger:
            context['completed_reservations'] = completed_paginator.page(1)
        except EmptyPage:
            context['completed_reservations'] = completed_paginator.page(completed_paginator.num_pages)
        
        # Paginate rejected batches
        rejected_paginator = Paginator(rejected_batches, paginate_by)
        try:
            context['rejected_reservations'] = rejected_paginator.page(rejected_page)
        except PageNotAnInteger:
            context['rejected_reservations'] = rejected_paginator.page(1)
        except EmptyPage:
            context['rejected_reservations'] = rejected_paginator.page(rejected_paginator.num_pages)
        
        # Paginate expired batches
        expired_paginator = Paginator(expired_batches, paginate_by)
        try:
            context['expired_reservations'] = expired_paginator.page(expired_page)
        except PageNotAnInteger:
            context['expired_reservations'] = expired_paginator.page(1)
        except EmptyPage:
            context['expired_reservations'] = expired_paginator.page(expired_paginator.num_pages)
        
        # Paginate voided batches
        voided_paginator = Paginator(voided_batches, paginate_by)
        try:
            context['voided_reservations'] = voided_paginator.page(voided_page)
        except PageNotAnInteger:
            context['voided_reservations'] = voided_paginator.page(1)
        except EmptyPage:
            context['voided_reservations'] = voided_paginator.page(voided_paginator.num_pages)
        
        # Add total counts for badges
        context['all_count'] = all_paginator.count
        context['pending_count'] = pending_paginator.count
        context['approved_count'] = approved_paginator.count
        context['active_count'] = active_paginator.count
        context['completed_count'] = completed_paginator.count
        context['rejected_count'] = rejected_paginator.count
        context['expired_count'] = expired_paginator.count
        context['voided_count'] = voided_paginator.count
        
        # Add departments for the filter dropdown
        context['departments'] = Department.objects.all()
        
        # Prepare reservation data for JavaScript calendar (if needed)
        reservation_data = []
        for batch in all_batches:
            # Get user display name
            if batch.user.first_name or batch.user.last_name:
                user_name = f"{batch.user.first_name} {batch.user.last_name}".strip()
            else:
                user_name = batch.user.username
            
            # Get first item name for display
            first_item = batch.items.first()
            item_name = first_item.property.property_name if first_item else "No items"
            if batch.total_items > 1:
                item_name += f" (+{batch.total_items - 1} more)"
            
            # Prepare batch object
            batch_obj = {
                'id': batch.id,
                'status': batch.status,
                'statusDisplay': batch.get_status_display(),
                'itemName': item_name,
                'userName': user_name,
                'department': str(batch.user.userprofile.department) if batch.user.userprofile.department else '',
                'quantity': batch.total_quantity,
                'purpose': batch.purpose[:50] + ('...' if len(batch.purpose) > 50 else ''),
                'neededDate': batch.earliest_needed_date.strftime('%Y-%m-%d') if batch.earliest_needed_date else '',
                'returnDate': batch.latest_return_date.strftime('%Y-%m-%d') if batch.latest_return_date else '',
                'requestDate': batch.request_date.strftime('%Y-%m-%d %H:%M'),
                'approvedDate': batch.approved_date.strftime('%Y-%m-%d %H:%M') if batch.approved_date else None,
                'remarks': batch.remarks if batch.remarks else '',
            }
                
            reservation_data.append(batch_obj)
        
        context['reservation_data'] = json.dumps(reservation_data)
        
        return context



class SupplyListView(PermissionRequiredMixin, ListView):
    model = Supply
    template_name = 'app/supply.html'
    permission_required = 'app.view_admin_module'
    context_object_name = 'supplies'
    paginate_by = 15  # Enable pagination with 15 items per page

    def get_queryset(self):
        queryset = Supply.objects.filter(is_archived=False).select_related('quantity_info', 'category', 'subcategory').order_by('supply_name')
        
        # Apply search filter
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(supply_name__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(category__name__icontains=search_query) |
                Q(subcategory__name__icontains=search_query) |
                Q(barcode__icontains=search_query)
            )
        
        # Apply category filter
        category_filters = self.request.GET.getlist('category')
        if category_filters:
            queryset = queryset.filter(category__name__in=category_filters)
        
        # Apply status filter
        status_filters = self.request.GET.getlist('status')
        if status_filters:
            status_conditions = []
            for status in status_filters:
                if status == 'available':
                    # Available: current_quantity > minimum_threshold
                    status_conditions.append(Q(quantity_info__current_quantity__gt=F('quantity_info__minimum_threshold')))
                elif status == 'low_stock':
                    # Low stock: current_quantity > 0 AND current_quantity <= minimum_threshold
                    status_conditions.append(
                        Q(quantity_info__current_quantity__gt=0) & 
                        Q(quantity_info__current_quantity__lte=F('quantity_info__minimum_threshold'))
                    )
                elif status == 'out_of_stock':
                    # Out of stock: current_quantity = 0
                    status_conditions.append(Q(quantity_info__current_quantity=0))
            
            if status_conditions:
                status_q = status_conditions[0]
                for condition in status_conditions[1:]:
                    status_q |= condition
                queryset = queryset.filter(status_q)
        
        # Apply availability filter
        availability_filters = self.request.GET.getlist('availability')
        if availability_filters:
            availability_conditions = []
            for availability in availability_filters:
                if availability == 'Yes':
                    availability_conditions.append(Q(available_for_request=True))
                elif availability == 'No':
                    availability_conditions.append(Q(available_for_request=False))
            
            if availability_conditions:
                availability_q = availability_conditions[0]
                for condition in availability_conditions[1:]:
                    availability_q |= condition
                queryset = queryset.filter(availability_q)
        
        # Apply expiry filter
        expiry_filters = self.request.GET.getlist('expiry')
        if expiry_filters:
            today = timezone.now().date()
            
            expiry_conditions = []
            for expiry in expiry_filters:
                if expiry == 'expired':
                    expiry_conditions.append(Q(expiration_date__lt=today))
                elif expiry == 'expiring_soon':
                    # Expires within 30 days
                    expiry_conditions.append(Q(expiration_date__gte=today, expiration_date__lte=today + timezone.timedelta(days=30)))
                elif expiry == 'not_expired':
                    expiry_conditions.append(Q(expiration_date__gt=today + timezone.timedelta(days=30)) | Q(expiration_date__isnull=True))
                elif expiry == 'no_expiry':
                    expiry_conditions.append(Q(expiration_date__isnull=True))
            
            if expiry_conditions:
                expiry_q = expiry_conditions[0]
                for condition in expiry_conditions[1:]:
                    expiry_q |= condition
                queryset = queryset.filter(expiry_q)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = SupplyForm()
        context['form'] = form
        context['today'] = date.today()
        
        # Get the paginated supplies from the current page
        supplies = context['supplies']
        
        # Ensure barcodes are generated for supplies that don't have them
        from .utils import generate_barcode_image
        for supply in supplies:
            if not supply.barcode or not supply.barcode_image:
                barcode_text = f"SUP-{supply.id}"
                supply.barcode = barcode_text
                filename, content = generate_barcode_image(barcode_text)
                supply.barcode_image.save(filename, content, save=False)
                supply.save(update_fields=['barcode', 'barcode_image'])
            
            # Calculate days until expiration
            if supply.expiration_date:
                days_until = (supply.expiration_date - date.today()).days
                supply.days_until_expiration = days_until
        
        # Get all categories and subcategories for the filter dropdowns
        context['categories'] = SupplyCategory.objects.prefetch_related('supply_set').all()
        context['subcategories'] = SupplySubcategory.objects.prefetch_related('supply_set').all()
        
        # Get all non-archived supplies for the "Report Bad Stock" modal dropdown
        # This includes all supplies (not paginated) so users can select any supply regardless of page
        context['all_supplies'] = Supply.objects.filter(
            is_archived=False
        ).select_related('quantity_info', 'category', 'subcategory').order_by('supply_name')
        
        # Only include paginated supplies for grouped view (not all supplies)
        # This improves page load performance by not loading entire database
        grouped = defaultdict(list)
        for supply in supplies:
            category_name = supply.category.name if supply.category else 'Uncategorized'
            grouped[category_name].append(supply)
        context['grouped_supplies'] = dict(sorted(grouped.items()))
        
        # Add user permissions to context
        context['can_report_bad_stock'] = has_admin_permission(self.request.user, 'report_bad_stock')
        context['can_edit_supply'] = has_admin_permission(self.request.user, 'edit_supply')
        
        return context

@permission_required('app.view_admin_module')
def add_supply(request):
    if request.method == 'POST':
        form = SupplyForm(request.POST)
        if form.is_valid():
            try:
                # First save the supply
                supply = form.save(commit=False)
                supply._logged_in_user = request.user
                supply.save(user=request.user)

                # Create and save quantity info
                quantity_info = SupplyQuantity.objects.create(
                    supply=supply,
                    current_quantity=form.cleaned_data['current_quantity'],
                    minimum_threshold=form.cleaned_data['minimum_threshold']
                )

                # available_for_request is already set from the form, no need to override

                # Create activity log
                ActivityLog.log_activity(
                    user=request.user,
                    action='create',
                    model_name='Supply',
                    object_repr=str(supply),
                    description=f"Added new supply '{supply.supply_name}' with initial quantity {form.cleaned_data['current_quantity']}"
                )
                messages.success(request, 'Supply added successfully.')
                # For AJAX requests, return JSON
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': 'Supply added successfully.'})
            except Exception as e:
                if 'supply' in locals() and supply.pk:
                    supply.delete()
                error_msg = f'Error adding supply: {str(e)}'
                messages.error(request, error_msg)
                # For AJAX requests, return JSON
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
        else:
            # Display form validation errors
            error_messages = {}
            for field, errors in form.errors.items():
                error_messages[field] = [str(e) for e in errors]
            # For AJAX requests, return JSON with field errors
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': error_messages})
            # For regular requests, show error message
            error_text = "; ".join([f"{field}: {', '.join(errors)}" for field, errors in error_messages.items()])
            messages.error(request, f'Please fix the following errors: {error_text}')
    return redirect('supply_list')


@permission_required('app.view_admin_module')
@login_required
@admin_permission_required('edit_supply')
def edit_supply(request):
    if request.method == 'POST':
        supply_id = request.POST.get('id')
        supply = get_object_or_404(Supply, id=supply_id)
        
        try:
            # Store old values for activity log
            old_values = {
                'supply_name': supply.supply_name,
                'description': supply.description,
                'unit': supply.unit,
                'available_for_request': supply.available_for_request,
                'category': supply.category.id if supply.category else None,
                'subcategory': supply.subcategory.id if supply.subcategory else None,
                'date_received': supply.date_received,
                'expiration_date': supply.expiration_date,
                'current_quantity': supply.quantity_info.current_quantity if supply.quantity_info else 0,
                'minimum_threshold': supply.quantity_info.minimum_threshold if supply.quantity_info else 0
            }

            # Update supply fields
            supply.supply_name = request.POST.get('supply_name')
            supply.description = request.POST.get('description')
            supply.unit = request.POST.get('unit') or None
            supply.available_for_request = request.POST.get('available_for_request') == 'on'
            
            # Handle category and subcategory
            category_id = request.POST.get('category')
            subcategory_id = request.POST.get('subcategory')
            
            if category_id:
                supply.category = get_object_or_404(SupplyCategory, id=category_id)
            else:
                supply.category = None
                
            if subcategory_id:
                supply.subcategory = get_object_or_404(SupplySubcategory, id=subcategory_id)
            else:
                supply.subcategory = None
                
            supply.date_received = request.POST.get('date_received')
            supply.expiration_date = request.POST.get('expiration_date') or None
            
            # Only update minimum_threshold, NOT current_quantity
            minimum_threshold = int(request.POST.get('minimum_threshold', 0))
            
            if not supply.quantity_info:
                # Create new quantity info, set current_quantity to 0 by default
                quantity_info = SupplyQuantity(
                    supply=supply,
                    current_quantity=0,
                    minimum_threshold=minimum_threshold
                )
                quantity_info.save(user=request.user)
            else:
                # Only update minimum_threshold
                supply.quantity_info.minimum_threshold = minimum_threshold
                supply.quantity_info.save(user=request.user)
            
            # Save supply with user information
            supply.save(user=request.user)

            # Create activity log for the update
            changes = []
            for field, old_value in old_values.items():
                if field == 'minimum_threshold':
                    new_value = minimum_threshold
                elif field == 'current_quantity':
                    new_value = old_value  # current_quantity is not changed here
                else:
                    new_value = getattr(supply, field)
                if str(old_value) != str(new_value):
                    changes.append(f"{field}: {old_value} → {new_value}")

            if changes:
                ActivityLog.log_activity(
                    user=request.user,
                    action='update',
                    model_name='Supply',
                    object_repr=str(supply),
                    description=f"Updated supply '{supply.supply_name}'. Changes: {', '.join(changes)}"
                )

            messages.success(request, 'Supply updated successfully!')
            return redirect('supply_list')
        except Exception as e:
            messages.error(request, f'Error updating supply: {str(e)}')
            return redirect('supply_list')
    return JsonResponse({'error': 'Invalid request method'}, status=400)

@permission_required('app.view_admin_module')
def delete_supply(request, pk):
    supply = get_object_or_404(Supply, pk=pk)
    if request.method == 'POST':
        supply_name = supply.supply_name
        repr_str = str(supply)
        supply._logged_in_user = request.user
        supply.delete()

        ActivityLog.objects.create(
            user=request.user,
            action='delete',
            model_name='Supply',
            object_repr=repr_str,
            description=f"Supply '{supply_name}' was deleted."
        )
        messages.success(request, 'Supply deleted successfully.')
    return redirect('supply_list')


@permission_required('app.view_admin_module')
@require_POST
@admin_permission_required('report_bad_stock')
def report_bad_stock(request):
    """Handle bad stock report submission via AJAX"""
    try:
        form = BadStockReportForm(request.POST)
        if form.is_valid():
            bad_stock = form.save(commit=False)
            bad_stock.reported_by = request.user
            bad_stock.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Successfully removed {bad_stock.quantity_removed} units of {bad_stock.supply.supply_name} from inventory.'
            })
        else:
            errors = {}
            for field, error_list in form.errors.items():
                errors[field] = error_list[0]
            return JsonResponse({
                'success': False,
                'errors': errors
            })
    except ValueError as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'An unexpected error occurred. Please try again.'
        })


@permission_required('app.view_admin_module')
def bad_stock_list(request):
    """Return list of bad stock reports as JSON for datatable"""
    reports = BadStockReport.objects.select_related(
        'supply', 'reported_by'
    ).order_by('-reported_at')
    
    # Apply filters if provided
    supply_id = request.GET.get('supply_id')
    if supply_id:
        reports = reports.filter(supply_id=supply_id)
    
    date_from = request.GET.get('date_from')
    if date_from:
        reports = reports.filter(reported_at__gte=date_from)
    
    date_to = request.GET.get('date_to')
    if date_to:
        reports = reports.filter(reported_at__lte=date_to)
    
    data = []
    for report in reports:
        data.append({
            'id': report.id,
            'supply_name': report.supply.supply_name,
            'quantity_removed': report.quantity_removed,
            'remarks': report.remarks,
            'reported_by': report.reported_by.get_full_name() or report.reported_by.username if report.reported_by else 'Unknown',
            'reported_at': report.reported_at.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return JsonResponse({'data': data})


class PropertyListView(PermissionRequiredMixin, ListView):
    model = Property
    template_name = 'app/property.html'
    context_object_name = 'properties'
    permission_required = 'app.view_admin_module'
    paginate_by = 15

    def get_queryset(self):
        queryset = Property.objects.filter(is_archived=False).select_related('category').order_by('property_name')
        
        # Apply search filter
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(property_name__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(category__name__icontains=search_query) |
                Q(property_number__icontains=search_query)
            )
        
        # Apply category filter (support multiple categories)
        category_filters = self.request.GET.getlist('category')
        if category_filters:
            queryset = queryset.filter(category__name__in=category_filters)
        
        # Apply availability filter
        availability_filter = self.request.GET.get('availability')
        if availability_filter == 'available':
            queryset = queryset.filter(availability='available')
        elif availability_filter == 'not_available':
            queryset = queryset.exclude(availability='available')
        
        # Apply condition filter (support multiple conditions)
        condition_filters = self.request.GET.getlist('condition')
        if condition_filters:
            queryset = queryset.filter(condition__in=condition_filters)
        
        # Apply price range filter
        price_range_filter = self.request.GET.get('priceRange')
        if price_range_filter == 'below50000':
            queryset = queryset.filter(unit_value__lt=50000)
        elif price_range_filter == 'above50000':
            queryset = queryset.filter(unit_value__gte=50000)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get the paginated properties from the parent context
        properties = context['properties']
        
        # Ensure barcodes are generated for properties that don't have them
        from .utils import generate_barcode_image
        for prop in properties:
            if not prop.barcode or not prop.barcode_image:
                barcode_text = prop.property_number if prop.property_number else f"PROP-{prop.id}"
                prop.barcode = barcode_text
                filename, content = generate_barcode_image(barcode_text)
                prop.barcode_image.save(filename, content, save=False)
                prop.save(update_fields=['barcode', 'barcode_image'])
        
        # Get all categories for the filter dropdown
        context['categories'] = PropertyCategory.objects.all()
        context['form'] = PropertyForm()
        
        # Build properties_by_category from ALL non-archived properties (not just paginated ones)
        # This ensures the Property Inventory modal dropdown shows all available properties
        all_properties = Property.objects.filter(is_archived=False).select_related('category').order_by('property_name')
        properties_by_category = defaultdict(list)
        for prop in all_properties:
            properties_by_category[prop.category].append(prop)
        context['properties_by_category'] = dict(properties_by_category)
        
        # Add user permissions to context
        context['can_report_lost_items'] = has_admin_permission(self.request.user, 'report_lost_items')
        context['can_manage_lost_items'] = has_admin_permission(self.request.user, 'manage_lost_items')
        context['can_edit_property'] = has_admin_permission(self.request.user, 'edit_property')
        
        return context
    
@require_POST
def modify_supply_quantity_generic(request):
    try:
        supply_id = request.POST.get("supply_id")
        if not supply_id:
            return JsonResponse({
                'success': False,
                'error': 'Supply ID is required'
            })

        amount = int(request.POST.get("amount", 0))
        supply = get_object_or_404(Supply, pk=supply_id)

        quantity_info = supply.quantity_info
        old_quantity = quantity_info.current_quantity

        # Always add the quantity since we removed the action type selection
        quantity_info.current_quantity += amount

        # Update supply's available_for_request status
        supply.available_for_request = (quantity_info.current_quantity > 0)
        supply.save(user=request.user)

        SupplyHistory.objects.create(
            supply=supply,
            user=request.user,
            action='quantity_update',
            field_name='quantity',
            old_value=str(old_quantity),
            new_value=str(quantity_info.current_quantity),
            remarks=f"add {amount}"
        )
        
        # Save with skip_history to avoid duplicate entry
        quantity_info.save(skip_history=True)

        # Add activity log
        ActivityLog.log_activity(
            user=request.user,
            action='quantity_update',
            model_name='Supply',
            object_repr=str(supply),
            description=f"Added {amount} units to supply '{supply.supply_name}'. Changed quantity from {old_quantity} to {quantity_info.current_quantity}"
        )

        messages.success(request, f"Quantity successfully added.")
        
        # Check if request expects JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Quantity updated successfully',
                'new_quantity': quantity_info.current_quantity
            })
        return redirect('supply_list')

    except SupplyQuantity.DoesNotExist:
        error_msg = "Supply quantity information not found."
        messages.error(request, error_msg)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': error_msg
            })
        return redirect('supply_list')
    except ValueError as e:
        error_msg = str(e)
        messages.error(request, error_msg)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': error_msg
            })
        return redirect('supply_list')

@permission_required('app.view_admin_module')
def add_property(request):
    if request.method == 'POST':
        form = PropertyForm(request.POST)
        if form.is_valid():
            try:
                # First save the property
                prop = form.save(commit=False)
                
                # Set initial quantity equal to overall_quantity for new properties
                if not prop.pk:
                    prop.quantity = form.cleaned_data['overall_quantity']
                
                prop._logged_in_user = request.user
                prop.save(user=request.user)  # Pass the user to save method

                # Ensure barcode and barcode image are generated
                if not prop.barcode or not prop.barcode_image:
                    from .utils import generate_barcode_image
                    barcode_text = prop.property_number if prop.property_number else f"PROP-{prop.id}"
                    prop.barcode = barcode_text
                    filename, content = generate_barcode_image(barcode_text)
                    prop.barcode_image.save(filename, content, save=False)
                    prop.save(update_fields=['barcode', 'barcode_image'], user=request.user)
                
                # Create activity log using log_activity method
                ActivityLog.log_activity(
                    user=request.user,
                    action='create',
                    model_name='Property',
                    object_repr=str(prop),
                    description=f"Added new property '{prop.property_name}' with property number {prop.property_number}, overall quantity {prop.overall_quantity}, and initial quantity {prop.quantity}"
                )
                messages.success(request, 'Property added successfully.')
            except Exception as e:
                # If anything goes wrong, delete the property to maintain data consistency
                if 'prop' in locals() and prop.pk:
                    prop.delete()
                messages.error(request, f'Error adding property: {str(e)}')
                return redirect('property_list')
        else:
            # Add form errors to messages
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            messages.error(request, f'Please correct the following errors: {", ".join(error_messages)}')
    return redirect('property_list')

@permission_required('app.view_admin_module')
@login_required
@admin_permission_required('edit_property')
def edit_property(request):
    if request.method == 'POST':
        property_id = request.POST.get('id')
        property_obj = get_object_or_404(Property, id=property_id)

        try:
            # Store old values for activity log
            old_values = {
                'property_number': property_obj.property_number,
                'property_name': property_obj.property_name,
                'description': property_obj.description,
                'barcode': property_obj.barcode,
                'unit_of_measure': property_obj.unit_of_measure,
                'unit_value': property_obj.unit_value,
                'overall_quantity': property_obj.overall_quantity,
                'quantity': property_obj.quantity,
                'location': property_obj.location,
                'accountable_person': property_obj.accountable_person,
                'year_acquired': property_obj.year_acquired,
                'condition': property_obj.condition,
                'category': property_obj.category,
                'availability': property_obj.availability
            }

            # Update property fields from POST data
            property_number = request.POST.get('property_number')
            property_obj.property_number = property_number.upper() if property_number else property_number
            property_obj.property_name = request.POST.get('property_name')
            property_obj.description = request.POST.get('description')
            property_obj.barcode = request.POST.get('barcode')
            property_obj.unit_of_measure = request.POST.get('unit_of_measure')

            # Convert unit_value to float safely - only if provided
            unit_value_str = request.POST.get('unit_value')
            if unit_value_str is not None and unit_value_str != '':
                try:
                    new_unit_value = float(unit_value_str)
                    property_obj.unit_value = new_unit_value
                except (TypeError, ValueError):
                    # Keep the existing value if conversion fails
                    pass

            # Convert overall_quantity to int safely
            try:
                property_obj.overall_quantity = int(request.POST.get('overall_quantity', 0))
            except (TypeError, ValueError):
                property_obj.overall_quantity = 0

            property_obj.location = request.POST.get('location')
            property_obj.accountable_person = request.POST.get('accountable_person')
            
            # Handle year_acquired date field
            year_acquired_str = request.POST.get('year_acquired')
            if year_acquired_str:
                try:
                    from datetime import datetime
                    property_obj.year_acquired = datetime.strptime(year_acquired_str, '%Y-%m-%d').date()
                except ValueError as e:
                    property_obj.year_acquired = None
            else:
                property_obj.year_acquired = None

            # Handle ForeignKey category assignment
            category_id = request.POST.get('category')
            if category_id:
                property_obj.category = get_object_or_404(PropertyCategory, id=category_id)
            else:
                property_obj.category = None  # or keep existing?

            # Handle condition and availability
            new_condition = request.POST.get('condition')
            property_obj.condition = new_condition

            unavailable_conditions = ['Needing repair', 'Unserviceable', 'No longer needed', 'Obsolete']
            if new_condition in unavailable_conditions:
                property_obj.availability = 'not_available'
            else:
                # Only use submitted availability if condition allows it
                property_obj.availability = request.POST.get('availability')

            # Save the updated property, assuming your model's save method accepts user param
            property_obj.save(user=request.user)
            
            # Debug: Print the saved values
            print(f"DEBUG: After save - accountable_person: '{property_obj.accountable_person}'")
            print(f"DEBUG: After save - year_acquired: '{property_obj.year_acquired}'")

            # Create activity log for changes
            changes = []
            for field, old_value in old_values.items():
                new_value = getattr(property_obj, field)
                # For category FK, compare by id or name for clarity
                if field == 'category':
                    old_val_repr = old_value.name if old_value else 'None'
                    new_val_repr = new_value.name if new_value else 'None'
                    if old_val_repr != new_val_repr:
                        changes.append(f"category: {old_val_repr} → {new_val_repr}")
                else:
                    if str(old_value) != str(new_value):
                        if field == 'overall_quantity':
                            changes.append(f"overall quantity: {old_value} → {new_value} (current quantity updated accordingly)")
                        elif field == 'quantity_per_physical_count':
                            changes.append(f"quantity per physical count: {old_value} → {new_value}")
                        elif field == 'condition' and new_value in unavailable_conditions:
                            changes.append(f"condition: {old_value} → {new_value} (automatically set availability to Not Available)")
                        else:
                            changes.append(f"{field.replace('_', ' ')}: {old_value} → {new_value}")

            if changes:
                ActivityLog.log_activity(
                    user=request.user,
                    action='update',
                    model_name='Property',
                    object_repr=str(property_obj),
                    description=f"Updated property '{property_obj.property_name}'. Changes: {', '.join(changes)}"
                )

            messages.success(request, 'Property updated successfully!')

        except Exception as e:
            messages.error(request, f'Error updating property: {str(e)}')

        return redirect('property_list')

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@permission_required('app.view_admin_module')
def change_property_number(request, pk):
    """View to handle changing property numbers"""
    property_obj = get_object_or_404(Property, pk=pk)
    
    if request.method == 'POST':
        form = PropertyNumberChangeForm(request.POST, instance=property_obj)
        if form.is_valid():
            old_number = property_obj.property_number
            new_number = form.cleaned_data['new_property_number']
            
            # Save with user context for proper history tracking
            form.save(user=request.user)
            
            # Create specific activity log for property number change
            ActivityLog.log_activity(
                user=request.user,
                action='update',
                model_name='Property',
                object_repr=str(property_obj),
                description=f"Changed property number from '{old_number}' to '{new_number}' for property '{property_obj.property_name}'"
            )
            
            return JsonResponse({
                'success': True, 
                'message': f'Property number changed from {old_number} to {new_number}',
                'new_number': new_number,
                'old_number': old_number
            })
        else:
            return JsonResponse({
                'success': False, 
                'errors': form.errors
            })
    
    else:
        form = PropertyNumberChangeForm(instance=property_obj)
    
    # For GET requests or form errors, return the form data
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method' if request.method != 'POST' else 'Form validation failed'
    })

@permission_required('app.view_admin_module')
def delete_property(request, pk):
    prop = get_object_or_404(Property, pk=pk)
    if request.method == 'POST':
        prop_name = prop.property_name
        repr_str = str(prop)
        prop._logged_in_user = request.user
        prop.delete()

        ActivityLog.objects.create(
            user=request.user,
            action='delete',
            model_name='Property',
            object_repr=repr_str,
            description=f"Property '{prop_name}' was deleted."
        )

        messages.success(request, 'Property deleted successfully.')
    return redirect('property_list')

@permission_required('app.view_admin_module')
@require_POST
def admin_mark_property_damaged(request, property_id):
    """
    View for admins to mark a property as damaged directly.
    Creates a damage report and immediately marks the property as damaged.
    """
    from .forms import AdminDamageReportForm
    
    property_obj = get_object_or_404(Property, id=property_id)
    
    form = AdminDamageReportForm(request.POST, request.FILES)
    
    if form.is_valid():
        damage_report = form.save(commit=False)
        damage_report.user = request.user  # Admin is the reporter
        damage_report.item = property_obj  # Set the property
        damage_report.status = 'approved'  # Automatically approved since admin created it
        damage_report.save()
        
        # Immediately mark the property as damaged
        property_obj.condition = 'Needing repair'
        property_obj.availability = 'not_available'
        property_obj.save(user=request.user)
        
        # Log the activity
        ActivityLog.log_activity(
            user=request.user,
            action='report',
            model_name='DamageReport',
            object_repr=str(property_obj.property_name),
            description=f"Admin marked property '{property_obj.property_name}' as damaged and moved to Damage Items. Reason: {damage_report.description[:100]}..."
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Property marked as damaged and moved to Damage Items successfully.',
            'damage_report_id': damage_report.id
        })
    else:
        return JsonResponse({
            'success': False,
            'errors': form.errors
        }, status=400)

@require_POST
@permission_required('app.view_admin_module')
@admin_permission_required('report_lost_items')
def report_lost_item(request, property_id):
    """
    View for admins to report a property as lost.
    When admin directly reports via this endpoint, it immediately marks as Lost without verification.
    Creates a lost item report and marks the property condition as Lost.
    """
    from .forms import LostItemForm
    
    property_obj = get_object_or_404(Property, id=property_id)
    
    form = LostItemForm(request.POST, request.FILES)
    
    if form.is_valid():
        lost_item = form.save(commit=False)
        lost_item.user = request.user  # Admin is the reporter
        lost_item.item = property_obj  # Set the property
        
        # Skip automatic status changes since admin is directly marking it
        lost_item.save(skip_auto_status=True)
        
        # Admin directly reporting - immediately mark as Lost without verification needed
        property_obj.condition = 'Lost'
        property_obj.availability = 'not_available'
        property_obj.save(update_fields=['condition', 'availability'])
        
        # Log the activity
        ActivityLog.log_activity(
            user=request.user,
            action='report',
            model_name='LostItem',
            object_repr=str(property_obj.property_name),
            description=f"Admin directly reported and marked property '{property_obj.property_name}' as Lost (no verification needed). Description: {lost_item.description[:100]}..."
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Item successfully reported and marked as Lost by admin.',
            'lost_item_id': lost_item.id
        })
    else:
        return JsonResponse({
            'success': False,
            'errors': form.errors
        }, status=400)

@login_required
@permission_required('app.view_admin_module')
def get_all_properties(request):
    """
    API endpoint to get all properties for the lost item dropdown.
    Returns all non-archived, available properties with distinct IDs.
    """
    properties = Property.objects.filter(
        is_archived=False
    ).values('id', 'property_name', 'property_number').distinct().order_by('property_name')
    
    properties_data = [{
        'id': prop['id'],
        'property_name': prop['property_name'],
        'property_number': prop['property_number'] or 'N/A'
    } for prop in properties]
    
    return JsonResponse({'properties': properties_data})

@permission_required('app.view_admin_module')
def add_property_category(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        uacs = request.POST.get('uacs')  # Get UACS field
        
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if name:
            # Check if category already exists
            if PropertyCategory.objects.filter(name=name).exists():
                if is_ajax:
                    return JsonResponse({
                        'success': False, 
                        'error': f'Category "{name}" already exists.'
                    })
                else:
                    messages.error(request, f'Category "{name}" already exists.')
            else:
                try:
                    # Convert UACS to integer if provided, otherwise set to None
                    uacs_value = None
                    if uacs and uacs.strip():
                        try:
                            uacs_value = int(uacs)
                        except ValueError:
                            if is_ajax:
                                return JsonResponse({
                                    'success': False, 
                                    'error': 'UACS must be a valid number.'
                                })
                            else:
                                messages.error(request, 'UACS must be a valid number.')
                                return redirect('add_property_category')
                    
                    category = PropertyCategory.objects.create(name=name, uacs=uacs_value)
                    
                    # Log the activity
                    ActivityLog.log_activity(
                        user=request.user,
                        action='create',
                        model_name='PropertyCategory',
                        object_repr=str(category),
                        description=f"Added new property category '{name}'" + (f" with UACS {uacs_value}" if uacs_value else "")
                    )
                    
                    if is_ajax:
                        return JsonResponse({
                            'success': True, 
                            'message': 'Category added successfully.',
                            'category': {
                                'id': category.id,
                                'name': category.name,
                                'uacs': category.uacs
                            }
                        })
                    else:
                        messages.success(request, 'Category added successfully.')
                except Exception as e:
                    if is_ajax:
                        return JsonResponse({
                            'success': False, 
                            'error': f'Error adding category: {str(e)}'
                        })
                    else:
                        messages.error(request, f'Error adding category: {str(e)}')
        else:
            if is_ajax:
                return JsonResponse({
                    'success': False, 
                    'error': 'Category name is required.'
                })
            else:
                messages.error(request, 'Category name is required.')
    
    return redirect('property_list')

@login_required
def get_property_categories(request):
    """Return all property categories as JSON for dropdown updates"""
    categories = PropertyCategory.objects.all().order_by('name')
    categories_data = [
        {'id': category.id, 'name': category.name}
        for category in categories
    ]
    return JsonResponse({'categories': categories_data})

class CheckOutPageView(PermissionRequiredMixin, TemplateView):
    template_name = 'app/checkout.html'
    permission_required = 'app.view_admin_module'  # Adjust permission as needed

class LandingPageView(TemplateView):
    """View for the landing page that provides login options for both admin and regular users."""
    template_name = "app/landing_page.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

#meow   
class CustomLoginView(LoginView):
    template_name = 'registration/login.html'
    
    def get_form_class(self):
        """Use custom authentication form"""
        from .forms import CustomAuthenticationForm
        return CustomAuthenticationForm
    
    def form_invalid(self, form):
        """Override to replace default inactive message with custom one"""
        # Check if the error is about inactive account and replace the message
        if form.errors.get('__all__'):
            for i, error in enumerate(form.errors['__all__']):
                # Replace Django's default inactive account messages
                if 'inactive' in str(error).lower() or 'disabled' in str(error).lower():
                    form.errors['__all__'][i] = 'This account is inactive. Please contact the administrator for assistance.'
        return super().form_invalid(form)
    
    def form_valid(self, form):
        """
        Handle successful login - enforce single session per user.
        If user has an existing active session, invalidate it.
        """
        user = form.get_user()
        
        # Invalidate previous session if it exists
        previous_session = UserSession.objects.filter(user=user).first()
        if previous_session:
            # Delete the old session from the sessions table to force logout
            try:
                from django.contrib.sessions.models import Session
                Session.objects.filter(session_key=previous_session.session_key).delete()
            except Exception as e:
                pass
            previous_session.delete()
        
        # Call parent's form_valid to create new session
        response = super().form_valid(form)
        
        # Create new UserSession record after session is created
        session_key = self.request.session.session_key
        if session_key:
            try:
                UserSession.objects.update_or_create(
                    user=user,
                    defaults={
                        'session_key': session_key,
                        'ip_address': self.get_client_ip(self.request),
                        'user_agent': self.request.META.get('HTTP_USER_AGENT', '')[:500]
                    }
                )
                
                # Log login activity
                ActivityLog.log_activity(
                    user=user,
                    action='login',
                    model_name='User',
                    object_repr=user.username,
                    description=f"User {user.username} logged in"
                )
            except Exception as e:
                # Log any errors but don't fail the login
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error creating UserSession: {str(e)}")
        
        return response
    
    @staticmethod
    def get_client_ip(request):
        """Get the client's IP address"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip
    
    def get_success_url(self):
        user = self.request.user

        if user.groups.filter(name='ADMIN').exists():
            return reverse('dashboard')
        elif user.groups.filter(name='USER').exists():
            return reverse('user_dashboard')
        else:
            logout(self.request)
            messages.error(self.request, "You do not have permission to access the system. Contact the Admin.")
            return reverse('login')


    def dashboard(request):
        return render(request, 'app/dashboard.html')

    def user_dashboard(request):
        return render(request, 'userpanel/user_dashboard.html')



def logout_view(request):
    if request.user.is_authenticated:
        username = request.user.username
        
        # Delete UserSession record on logout
        try:
            UserSession.objects.filter(user=request.user).delete()
        except:
            pass
        
        ActivityLog.log_activity(
            user=request.user,
            action='logout',
            model_name='User',
            object_repr=username,
            description=f"User {username} logged out"
        )
    return LogoutView.as_view()(request)

@login_required
def create_supply_request(request):
    """
    New cart-based supply request system.
    Users can add multiple items to a cart before submitting.
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_to_cart':
            # Handle adding item to list
            return add_to_list(request)
        elif action == 'remove_from_cart':
            # Handle removing item from list
            return remove_from_list(request)
        elif action == 'update_cart':
            # Handle updating item quantity in list
            return update_list_item(request)
        elif action == 'submit_request':
            # Handle final request submission
            return submit_list_request(request)
    
    # GET request - show the cart page
    cart_items = request.session.get('supply_cart', [])
    
    # Get recent batch requests (new system)
    recent_batch_requests = SupplyRequestBatch.objects.filter(user=request.user).order_by('-request_date')[:5]
    
    # Get recent single requests (legacy system) 
    recent_single_requests = SupplyRequest.objects.filter(user=request.user).order_by('-request_date')[:5]
    
    # Combine and format recent requests
    recent_requests_data = []
    
    # Add batch requests
    for req in recent_batch_requests:
        items_text = f"{req.total_items} items"
        if req.total_items <= 3:
            items_list = ", ".join([f"{item.supply.supply_name} (x{item.quantity})" for item in req.items.all()])
            items_text = items_list
        
        recent_requests_data.append({
            'id': req.id,
            'item': items_text,
            'quantity': req.total_quantity,
            'status': req.status,
            'date': req.request_date,
            'purpose': req.purpose,
            'type': 'batch'
        })
    
    # Add single requests (for backward compatibility)
    for req in recent_single_requests:
        recent_requests_data.append({
            'id': req.id,
            'item': req.supply.supply_name,
            'quantity': req.quantity,
            'status': req.status,
            'date': req.request_date,
            'purpose': req.purpose,
            'type': 'single'
        })
    
    # Sort by date
    recent_requests_data.sort(key=lambda x: x['date'], reverse=True)
    recent_requests_data = recent_requests_data[:5]  # Keep only 5 most recent
    
    context = {
        'cart_items': cart_items,
        'item_form': SupplyRequestItemForm(),
        'batch_form': SupplyRequestBatchForm(),
        'available_supplies': Supply.objects.filter(
            available_for_request=True,
            quantity_info__current_quantity__gt=0
        ).select_related('quantity_info'),
        'recent_requests': recent_requests_data,
    }
    
    return render(request, 'userpanel/user_request.html', context)

@login_required
def add_to_list(request):
    """Add an item to the supply request list"""
    if request.method == 'POST':
        supply_id = request.POST.get('supply_id')
        quantity = int(request.POST.get('quantity', 0))
        
        try:
            supply = Supply.objects.get(id=supply_id)
            
            # Validate quantity
            available_quantity = supply.quantity_info.current_quantity
            if quantity > available_quantity:
                return JsonResponse({
                    'success': False,
                    'message': f'Only {available_quantity} units of {supply.supply_name} are available.'
                })
            
            # Get or create cart in session
            cart = request.session.get('supply_cart', [])
            
            # Check if item already exists in cart
            item_exists = False
            for item in cart:
                if item['supply_id'] == supply_id:
                    # Update quantity (prevent duplicates)
                    new_quantity = item['quantity'] + quantity
                    if new_quantity > available_quantity:
                        return JsonResponse({
                            'success': False,
                            'message': f'Cannot add {quantity} more. Total would exceed available quantity ({available_quantity}).'
                        })
                    item['quantity'] = new_quantity
                    item_exists = True
                    break
            
            if not item_exists:
                cart.append({
                    'supply_id': supply_id,
                    'supply_name': supply.supply_name,
                    'quantity': quantity,
                    'available_quantity': available_quantity
                })
            
            request.session['supply_cart'] = cart
            
            return JsonResponse({
                'success': True,
                'message': f'Added {supply.supply_name} to list.',
                'list_count': len(cart)
            })
            
        except Supply.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Supply item not found.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error adding item to cart: {str(e)}'
            })

@login_required
def remove_from_list(request):
    """Remove an item from the supply request list"""
    if request.method == 'POST':
        supply_id = request.POST.get('supply_id')
        
        cart = request.session.get('supply_cart', [])
        cart = [item for item in cart if item['supply_id'] != supply_id]
        
        request.session['supply_cart'] = cart
        
        return JsonResponse({
            'success': True,
            'message': 'Item removed from cart.',
            'cart_count': len(cart)
        })

@login_required
def clear_supply_list(request):
    """Clear all items from the supply request list"""
    if request.method == 'POST':
        request.session['supply_cart'] = []
        request.session.modified = True
        
        return JsonResponse({
            'success': True,
            'message': 'List cleared successfully.',
            'cart_count': 0
        })
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method.'
    })

@login_required
def update_list_item(request):
    """Update quantity of an item in the list"""
    if request.method == 'POST':
        supply_id = request.POST.get('supply_id')
        new_quantity = int(request.POST.get('quantity', 0))
        
        try:
            supply = Supply.objects.get(id=supply_id)
            available_quantity = supply.quantity_info.current_quantity
            
            if new_quantity > available_quantity:
                return JsonResponse({
                    'success': False,
                    'message': f'Only {available_quantity} units available.'
                })
            
            cart = request.session.get('supply_cart', [])
            
            for item in cart:
                if item['supply_id'] == supply_id:
                    item['quantity'] = new_quantity
                    break
            
            request.session['supply_cart'] = cart
            
            return JsonResponse({
                'success': True,
                'message': 'Cart updated.',
                'cart_count': len(cart)
            })
            
        except Supply.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Supply item not found.'
            })

@login_required
def submit_list_request(request):
    """Submit the list as a batch supply request"""
    if request.method == 'POST':
        purpose = request.POST.get('purpose', '').strip()
        cart = request.session.get('supply_cart', [])
        
        if not cart:
            messages.error(request, 'Your request list is empty. Please add items before submitting.')
            return redirect('create_supply_request')
        
        if not purpose:
            messages.error(request, 'Please provide a purpose for your request.')
            return redirect('create_supply_request')
        
        try:
            # Create the batch request
            batch_request = SupplyRequestBatch.objects.create(
                user=request.user,
                purpose=purpose,
                status='pending'
            )
            
            # Create individual items
            for cart_item in cart:
                supply = Supply.objects.get(id=cart_item['supply_id'])
                SupplyRequestItem.objects.create(
                    batch_request=batch_request,
                    supply=supply,
                    quantity=cart_item['quantity']
                )
            
            # Log activity - build item list from supply objects, not cart data
            supplies_for_logging = []
            for cart_item in cart[:3]:
                try:
                    supply = Supply.objects.get(id=cart_item['supply_id'])
                    supplies_for_logging.append(f"{supply.supply_name} (x{cart_item['quantity']})")
                except Supply.DoesNotExist:
                    supplies_for_logging.append(f"Unknown Item (x{cart_item['quantity']})")
            
            item_list = ", ".join(supplies_for_logging)
            if len(cart) > 3:
                item_list += f" and {len(cart) - 3} more items"
            
            ActivityLog.log_activity(
                user=request.user,
                action='request',
                model_name='SupplyRequestBatch',
                object_repr=f"Batch #{batch_request.id}",
                description=f"Submitted batch supply request with {len(cart)} items: {item_list}"
            )
            
            # Clear the cart
            request.session['supply_cart'] = []
            request.session.modified = True
            
            messages.success(request, f'Supply request submitted successfully! Your request ID is #{batch_request.id}.')
            return redirect('user_unified_request')
            
        except Exception as e:
            messages.error(request, f'Error submitting request: {str(e)}')
            return redirect('user_unified_request')


@login_required
def create_borrow_request(request):
    if request.method == 'POST':
        property_id = request.POST.get('property_id')
        quantity = request.POST.get('quantity')
        borrow_date = request.POST.get('borrow_date')
        return_date = request.POST.get('return_date')
        purpose = request.POST.get('purpose')
        
        property_item = get_object_or_404(Property, id=property_id)
        borrow_request = BorrowRequest.objects.create(
            user=request.user,
            property=property_item,
            quantity=quantity,
            borrow_date=borrow_date,
            return_date=return_date,
            purpose=purpose,
            status='pending'
        )

        ActivityLog.log_activity(
            user=request.user,
            action='request',
            model_name='Property',
            object_repr=property_item.property_name,
            description=f"Requested to borrow {quantity} units of {property_item.property_name} from {borrow_date} to {return_date} for {purpose}"
        )

        messages.success(request, 'Borrow request created successfully.')
        return redirect('create_borrow_request')


@permission_required('app.view_admin_module')
@login_required
def approve_borrow_request(request, request_id):
    borrow_request = get_object_or_404(BorrowRequest, id=request_id)
    borrow_request.status = 'approved'
    borrow_request.save()

    ActivityLog.log_activity(
        user=request.user,
        action='approve',
        model_name='BorrowRequest',
        object_repr=f"Request #{request_id}",
        description=f"Approved borrow request for {borrow_request.quantity} units of {borrow_request.property.property_name}"
    )

    messages.success(request, 'Borrow request approved successfully.')
    return redirect('borrow_requests')


@permission_required('app.view_admin_module')
@login_required
def reject_borrow_request(request, request_id):
    borrow_request = get_object_or_404(BorrowRequest, id=request_id)
    borrow_request.status = 'rejected'
    borrow_request.save()

    ActivityLog.log_activity(
        user=request.user,
        action='reject',
        model_name='BorrowRequest',
        object_repr=f"Request #{request_id}",
        description=f"Rejected borrow request for {borrow_request.quantity} units of {borrow_request.property.property_name}"
    )

    messages.success(request, 'Borrow request rejected successfully.')
    return redirect('borrow_requests')

@login_required
def get_supply_history(request, supply_id):
    try:
        supply = get_object_or_404(Supply, id=supply_id)
        history = supply.history.all().order_by('-timestamp')
        
        def get_field_display_name(field_name):
            """Convert field names to user-friendly display names"""
            field_mapping = {
                'supply_name': 'Supply Name',
                'category': 'Category',
                'subcategory': 'Subcategory',
                'description': 'Description',
                'minimum_threshold': 'Minimum Threshold',
                'current_quantity': 'Current Quantity',
                'available_for_request': 'Available for Request',
                'date_received': 'Date Received',
                'expiration_date': 'Expiration Date',
                'initial_creation': 'Supply Creation'
            }
            return field_mapping.get(field_name, field_name.replace('_', ' ').title())
        
        def get_action_display(action):
            """Convert action codes to user-friendly display names"""
            action_mapping = {
                'create': 'Created',
                'update': 'Updated',
                'quantity_update': 'Quantity Changed'
            }
            return action_mapping.get(action, action.replace('_', ' ').title())
        
        def format_change_description(entry):
            """Create a descriptive change summary"""
            try:
                field_display = get_field_display_name(entry.field_name)
                
                if entry.action == 'create':
                    return f"Supply was created in the system"
                elif entry.action == 'quantity_update':
                    if entry.remarks:
                        return f"{field_display}: {entry.remarks}"
                    else:
                        return f"{field_display} was modified"
                elif entry.action == 'update':
                    if entry.field_name == 'available_for_request':
                        old_display = 'Available' if entry.old_value == 'True' else 'Not Available'
                        new_display = 'Available' if entry.new_value == 'True' else 'Not Available'
                        return f"Availability changed from '{old_display}' to '{new_display}'"
                    elif entry.field_name == 'category':
                        return f"Category changed from '{entry.old_value or 'None'}' to '{entry.new_value or 'None'}'"
                    elif entry.field_name == 'subcategory':
                        return f"Subcategory changed from '{entry.old_value or 'None'}' to '{entry.new_value or 'None'}'"
                    elif entry.field_name == 'current_quantity':
                        return f"Current Quantity changed from {entry.old_value or '0'} to {entry.new_value or '0'}"
                    elif entry.field_name == 'minimum_threshold':
                        return f"Minimum Threshold changed from {entry.old_value or '0'} to {entry.new_value or '0'}"
                    elif 'quantity' in entry.field_name:
                        return f"{field_display} changed from {entry.old_value or '0'} to {entry.new_value or '0'}"
                    else:
                        return f"{field_display} was updated"
                else:
                    return f"{field_display} was modified"
            except Exception as e:
                return f"Supply change recorded"
        
        def format_value_change(entry):
            """Format the value change display"""
            try:
                if entry.action == 'create':
                    return entry.new_value if entry.new_value else 'Supply created'
                elif entry.old_value and entry.new_value and entry.old_value != entry.new_value:
                    # Handle special formatting for different field types
                    if 'quantity' in entry.field_name:
                        return f"{entry.old_value} → {entry.new_value}"
                    elif entry.field_name == 'available_for_request':
                        old_display = 'Available' if entry.old_value == 'True' else 'Not Available'
                        new_display = 'Available' if entry.new_value == 'True' else 'Not Available'
                        return f"{old_display} → {new_display}"
                    else:
                        # Escape HTML and limit length for long values
                        old_val = str(entry.old_value)[:100] + ('...' if len(str(entry.old_value)) > 100 else '')
                        new_val = str(entry.new_value)[:100] + ('...' if len(str(entry.new_value)) > 100 else '')
                        return f"{old_val} → {new_val}"
                elif entry.new_value:
                    new_val = str(entry.new_value)[:100] + ('...' if len(str(entry.new_value)) > 100 else '')
                    return f"Set to: {new_val}"
                elif entry.old_value:
                    # Field was cleared/removed
                    if 'quantity' in entry.field_name.lower():
                        return f"{entry.old_value} → 0"
                    else:
                        old_val = str(entry.old_value)[:100] + ('...' if len(str(entry.old_value)) > 100 else '')
                        return f"Removed: {old_val}"
                else:
                    return "No value change"
            except Exception as e:
                return "Value changed"
        
        history_data = []
        for entry in history:
            try:
                # Format timestamp for better readability
                formatted_date = entry.timestamp.strftime('%m/%d/%Y')
                formatted_time = entry.timestamp.strftime('%I:%M %p')
                formatted_datetime = f"{formatted_date}<br><small style='color: #6c757d;'>{formatted_time}</small>"
                
                history_data.append({
                    'id': entry.id,
                    'datetime': formatted_datetime,
                    'action': entry.action,
                    'action_display': get_action_display(entry.action),
                    'field_name': entry.field_name,
                    'field_display': get_field_display_name(entry.field_name),
                    'old_value': entry.old_value if entry.old_value is not None else '',
                    'new_value': entry.new_value if entry.new_value is not None else '',
                    'value_change': format_value_change(entry),
                    'change_description': format_change_description(entry),
                    'user': entry.user.get_full_name() if entry.user and entry.user.get_full_name() else (entry.user.username if entry.user else 'System'),
                    'remarks': entry.remarks if entry.remarks else '',
                    'timestamp_iso': entry.timestamp.isoformat(),
                    'raw_timestamp': entry.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                    # Debug information
                    'debug_field': entry.field_name,
                    'debug_old': entry.old_value,
                    'debug_new': entry.new_value
                })
            except Exception as e:
                # Log error but continue processing other entries
                print(f"Error processing supply history entry {entry.id}: {str(e)}")
                continue
        
        return JsonResponse({
            'success': True,
            'history': history_data,
            'supply_name': supply.supply_name,
            'supply_id': supply.id
        })
        
    except Supply.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Supply not found'
        }, status=404)
    except Exception as e:
        print(f"Error in get_supply_history: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while fetching supply history'
        }, status=500)

@login_required
def get_supply_quantity_activity(request, supply_id):
    """
    Fetch quantity change activity for a specific supply.
    Returns JSON with quantity-related history entries only.
    Supports filtering by user, date range, month, and activity type.
    """
    try:
        supply = get_object_or_404(Supply, id=supply_id)
        
        # Get filter parameters
        user_filter = request.GET.get('user_filter', 'all')
        activity_type_filter = request.GET.get('activity_type_filter', 'all')
        date_filter_type = request.GET.get('date_filter_type', 'all')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        month = request.GET.get('month')
        year = request.GET.get('year')
        
        # Filter history for quantity-related changes only
        quantity_history = supply.history.filter(
            field_name__in=['quantity', 'current_quantity', 'initial_quantity']
        ).order_by('-timestamp')
        
        # Apply user filter (filter by requestor name in remarks)
        if user_filter != 'all' and user_filter:
            # Filter by requestor name mentioned in remarks
            quantity_history = quantity_history.filter(remarks__icontains=f'Supply request by {user_filter}')
        
        # Apply activity type filter
        if activity_type_filter != 'all':
            if activity_type_filter == 'supply_request':
                quantity_history = quantity_history.filter(remarks__icontains='Supply request by')
            elif activity_type_filter == 'manual':
                quantity_history = quantity_history.exclude(remarks__icontains='Supply request by').exclude(action='create')
            elif activity_type_filter == 'addition':
                # For additions, we need to check if new_value > old_value
                addition_entries = []
                for entry in quantity_history:
                    old_qty = int(entry.old_value) if entry.old_value and entry.old_value.isdigit() else 0
                    new_qty = int(entry.new_value) if entry.new_value and entry.new_value.isdigit() else 0
                    if entry.action == 'create' or new_qty > old_qty:
                        addition_entries.append(entry.id)
                quantity_history = quantity_history.filter(id__in=addition_entries)
            elif activity_type_filter == 'deduction':
                # For deductions, we need to check if new_value < old_value
                deduction_entries = []
                for entry in quantity_history:
                    old_qty = int(entry.old_value) if entry.old_value and entry.old_value.isdigit() else 0
                    new_qty = int(entry.new_value) if entry.new_value and entry.new_value.isdigit() else 0
                    if entry.action != 'create' and new_qty < old_qty:
                        deduction_entries.append(entry.id)
                quantity_history = quantity_history.filter(id__in=deduction_entries)
        
        # Apply date filters
        if date_filter_type == 'date_range' and start_date and end_date:
            try:
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                quantity_history = quantity_history.filter(
                    timestamp__range=[start_dt, end_dt]
                )
            except ValueError:
                pass  # Invalid date format, ignore filter
        elif date_filter_type == 'month' and month and year:
            try:
                month_int = int(month)
                year_int = int(year)
                quantity_history = quantity_history.filter(
                    timestamp__year=year_int,
                    timestamp__month=month_int
                )
            except ValueError:
                pass  # Invalid month or year, ignore filter
        
        # Extract unique requestor names from remarks (e.g., "Supply request by john vales (Batch #4)")
        import re
        
        request_entries = supply.history.filter(
            field_name__in=['quantity', 'current_quantity', 'initial_quantity'],
            remarks__icontains='Supply request by'
        )
        
        requestor_names = set()
        for entry in request_entries:
            if entry.remarks and 'Supply request by' in entry.remarks:
                # Extract requestor name using regex pattern
                # Pattern matches "Supply request by [name] (" or "Supply request by [name]" at end
                match = re.search(r'Supply request by ([^(]+?)(?:\s*\(|$)', entry.remarks)
                if match:
                    requestor_name = match.group(1).strip()
                    if requestor_name:
                        requestor_names.add(requestor_name)
        
        # Create users_data with requestor names
        users_data = []
        for name in sorted(requestor_names):
            users_data.append({
                'id': name,  # Use name as ID since we're filtering by name
                'name': name,
                'username': name
            })
        
        activity_data = []
        for entry in quantity_history:
            try:
                # Calculate quantity change
                old_qty = int(entry.old_value) if entry.old_value and entry.old_value.isdigit() else 0
                new_qty = int(entry.new_value) if entry.new_value and entry.new_value.isdigit() else 0
                quantity_change = new_qty - old_qty
                
                # Determine action type based on change and remarks
                if entry.action == 'create':
                    action_type = 'Initial Creation'
                    quantity_change = new_qty  # For creation, the change is the initial amount
                elif 'Supply request by' in (entry.remarks or ''):
                    action_type = 'Supply Request'
                    # Keep the original quantity_change calculation
                elif quantity_change > 0:
                    action_type = 'Addition'
                elif quantity_change < 0:
                    action_type = 'Deduction'
                else:
                    action_type = 'Adjustment'
                
                activity_data.append({
                    'id': entry.id,
                    'date': entry.timestamp.isoformat(),
                    'action': action_type,
                    'quantity_change': quantity_change,
                    'previous_quantity': old_qty,
                    'new_quantity': new_qty,
                    'user': entry.user.get_full_name() if entry.user and entry.user.get_full_name() else (entry.user.username if entry.user else 'System'),
                    'user_id': entry.user.id if entry.user else None,
                    'remarks': entry.remarks if entry.remarks else '',
                    'timestamp': entry.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                })
            except Exception as e:
                print(f"Error processing quantity activity entry {entry.id}: {str(e)}")
                continue
        
        return JsonResponse({
            'success': True,
            'activity': activity_data,
            'users': users_data,
            'supply_name': supply.supply_name,
            'supply_id': supply.id,
            'total_entries': len(activity_data),
            'filters_applied': {
                'user_filter': user_filter,
                'activity_type_filter': activity_type_filter,
                'date_filter_type': date_filter_type
            }
        })
        
    except Supply.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Supply not found'
        }, status=404)
    except Exception as e:
        print(f"Error in get_supply_quantity_activity: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while fetching quantity activity'
        }, status=500)

@login_required
def get_property_history(request, property_id):
    try:
        property_obj = get_object_or_404(Property, id=property_id)
        history = property_obj.history.all().order_by('-timestamp')
        
        def get_field_display_name(field_name):
            """Convert field names to user-friendly display names"""
            field_mapping = {
                'property_number': 'Property Number',
                'property_name': 'Property Name',
                'category': 'Category',
                'description': 'Description',
                'barcode': 'Barcode',
                'unit_of_measure': 'Unit of Measure',
                'unit_value': 'Unit Value',
                'overall_quantity': 'Overall Quantity',
                'quantity': 'Current Quantity',
                'quantity_per_physical_count': 'Physical Count Quantity',
                'location': 'Location',
                'accountable_person': 'Accountable Person',
                'year_acquired': 'Year Acquired',
                'condition': 'Condition',
                'availability': 'Availability',
                'initial_creation': 'Property Creation'
            }
            return field_mapping.get(field_name, field_name.replace('_', ' ').title())
        
        def get_action_display(action):
            """Convert action codes to user-friendly display names"""
            action_mapping = {
                'create': 'Created',
                'update': 'Updated',
                'quantity_update': 'Quantity Changed'
            }
            return action_mapping.get(action, action.replace('_', ' ').title())
        
        def format_change_description(entry):
            """Create a descriptive change summary"""
            try:
                field_display = get_field_display_name(entry.field_name)
                
                if entry.action == 'create':
                    return f"Property was created in the system"
                elif entry.action == 'quantity_update':
                    if entry.remarks:
                        return f"{field_display}: {entry.remarks}"
                    else:
                        return f"{field_display} was modified"
                elif entry.action == 'update':
                    if entry.field_name == 'condition':
                        return f"Condition changed from '{entry.old_value}' to '{entry.new_value}'"
                    elif entry.field_name == 'availability':
                        old_display = 'Available' if entry.old_value == 'available' else 'Not Available'
                        new_display = 'Available' if entry.new_value == 'available' else 'Not Available'
                        return f"Availability changed from '{old_display}' to '{new_display}'"
                    elif entry.field_name == 'category':
                        return f"Category changed from '{entry.old_value or 'None'}' to '{entry.new_value or 'None'}'"
                    elif entry.field_name == 'quantity_per_physical_count':
                        return f"Physical Count Quantity changed from {entry.old_value or '0'} to {entry.new_value or '0'}"
                    elif entry.field_name == 'overall_quantity':
                        return f"Overall Quantity changed from {entry.old_value or '0'} to {entry.new_value or '0'}"
                    elif entry.field_name == 'quantity':
                        return f"Current Quantity changed from {entry.old_value or '0'} to {entry.new_value or '0'}"
                    elif 'quantity' in entry.field_name:
                        return f"{field_display} changed from {entry.old_value or '0'} to {entry.new_value or '0'}"
                    else:
                        return f"{field_display} was updated"
                else:
                    return f"{field_display} was modified"
            except Exception as e:
                return f"Property change recorded"
        
        def format_value_change(entry):
            """Format the value change display"""
            try:
                if entry.action == 'create':
                    return entry.new_value if entry.new_value else 'Property created'
                elif entry.old_value and entry.new_value and entry.old_value != entry.new_value:
                    # Handle special formatting for different field types
                    if entry.field_name == 'unit_value':
                        try:
                            old_val = f"₱{float(entry.old_value):,.2f}"
                            new_val = f"₱{float(entry.new_value):,.2f}"
                            return f"{old_val} → {new_val}"
                        except (ValueError, TypeError):
                            return f"{entry.old_value} → {entry.new_value}"
                    elif 'quantity' in entry.field_name:
                        return f"{entry.old_value} → {entry.new_value}"
                    elif entry.field_name == 'availability':
                        old_display = 'Available' if entry.old_value == 'available' else 'Not Available'
                        new_display = 'Available' if entry.new_value == 'available' else 'Not Available'
                        return f"{old_display} → {new_display}"
                    else:
                        # Escape HTML and limit length for long values
                        old_val = str(entry.old_value)[:100] + ('...' if len(str(entry.old_value)) > 100 else '')
                        new_val = str(entry.new_value)[:100] + ('...' if len(str(entry.new_value)) > 100 else '')
                        return f"{old_val} → {new_val}"
                elif entry.new_value:
                    new_val = str(entry.new_value)[:100] + ('...' if len(str(entry.new_value)) > 100 else '')
                    return f"Set to: {new_val}"
                elif entry.old_value:
                    # Field was cleared/removed
                    # Only show "Removed" for fields where it makes sense (not quantities)
                    if 'quantity' in entry.field_name.lower():
                        return f"{entry.old_value} → 0"
                    else:
                        old_val = str(entry.old_value)[:100] + ('...' if len(str(entry.old_value)) > 100 else '')
                        return f"Removed: {old_val}"
                else:
                    return "No value change"
            except Exception as e:
                return "Value changed"
        
        history_data = []
        for entry in history:
            try:
                # Debug logging for quantity-related entries
                if 'quantity' in entry.field_name.lower():
                    print(f"DEBUG: Processing quantity entry - Field: {entry.field_name}, Old: {entry.old_value}, New: {entry.new_value}, Action: {entry.action}")
                
                # Format timestamp for better readability
                formatted_date = entry.timestamp.strftime('%m/%d/%Y')
                formatted_time = entry.timestamp.strftime('%I:%M %p')
                formatted_datetime = f"{formatted_date}<br><small style='color: #6c757d;'>{formatted_time}</small>"
                
                history_data.append({
                    'id': entry.id,
                    'datetime': formatted_datetime,
                    'action': entry.action,
                    'action_display': get_action_display(entry.action),
                    'field_name': entry.field_name,
                    'field_display': get_field_display_name(entry.field_name),
                    'old_value': entry.old_value if entry.old_value is not None else '',
                    'new_value': entry.new_value if entry.new_value is not None else '',
                    'value_change': format_value_change(entry),
                    'change_description': format_change_description(entry),
                    'user': entry.user.get_full_name() if entry.user and entry.user.get_full_name() else (entry.user.username if entry.user else 'System'),
                    'remarks': entry.remarks if entry.remarks else '',
                    'timestamp_iso': entry.timestamp.isoformat(),
                    'raw_timestamp': entry.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                    # Debug information
                    'debug_field': entry.field_name,
                    'debug_old': entry.old_value,
                    'debug_new': entry.new_value
                })
            except Exception as e:
                # Log error but continue processing other entries
                print(f"Error processing history entry {entry.id}: {str(e)}")
                continue
        
        return JsonResponse({
            'success': True,
            'history': history_data,
            'property_name': property_obj.property_name,
            'property_number': property_obj.property_number
        })
        
    except Property.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Property not found'
        }, status=404)
    except Exception as e:
        print(f"Error in get_property_history: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while fetching property history'
        }, status=500)

@require_POST
def modify_supply_quantity_generic(request):
    try:
        supply_id = request.POST.get("supply_id")
        if not supply_id:
            return JsonResponse({
                'success': False,
                'error': 'Supply ID is required'
            })

        amount = int(request.POST.get("amount", 0))
        supply = get_object_or_404(Supply, pk=supply_id)

        quantity_info = supply.quantity_info
        old_quantity = quantity_info.current_quantity

        # Always add the quantity since we removed the action type selection
        quantity_info.current_quantity += amount

        # Update supply's available_for_request status
        supply.available_for_request = (quantity_info.current_quantity > 0)
        supply.save(user=request.user)

        SupplyHistory.objects.create(
            supply=supply,
            user=request.user,
            action='quantity_update',
            field_name='quantity',
            old_value=str(old_quantity),
            new_value=str(quantity_info.current_quantity),
            remarks=f"add {amount}"
        )
        
        # Save with skip_history to avoid duplicate entry
        quantity_info.save(skip_history=True)

        # Add activity log
        ActivityLog.log_activity(
            user=request.user,
            action='quantity_update',
            model_name='Supply',
            object_repr=str(supply),
            description=f"Added {amount} units to supply '{supply.supply_name}'. Changed quantity from {old_quantity} to {quantity_info.current_quantity}"
        )

        messages.success(request, f"Quantity successfully added.")
        
        # Check if request expects JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Quantity updated successfully',
                'new_quantity': quantity_info.current_quantity
            })
        return redirect('supply_list')

    except SupplyQuantity.DoesNotExist:
        error_msg = "Supply quantity information not found."
        messages.error(request, error_msg)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': error_msg
            })
        return redirect('supply_list')
    except ValueError as e:
        error_msg = str(e)
        messages.error(request, error_msg)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': error_msg
            })
        return redirect('supply_list')

class AdminPasswordChangeView(LoginRequiredMixin, PasswordChangeView):
    template_name = 'app/password_change.html'
    success_url = reverse_lazy('password_change_done')

    def get_template_names(self):
        return [self.template_name]

    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Clear the must_change_password flag after successful password change
        try:
            profile = UserProfile.objects.get(user=self.request.user)
            if profile.must_change_password:
                profile.must_change_password = False
                profile.save()
        except UserProfile.DoesNotExist:
            pass
        
        return response

class AdminPasswordChangeDoneView(LoginRequiredMixin, PasswordChangeDoneView):
    template_name = 'app/password_change_done.html'

    def get_template_names(self):
        return [self.template_name]
    
    def get(self, request, *args, **kwargs):
        # Add success message on the done page
        messages.success(request, 'Your password was successfully updated!')
        # Display success message on the done page, then redirect
        return render(request, self.get_template_names()[0])

@permission_required('app.view_admin_module')
@login_required
def add_category(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        
        if not name:
            return JsonResponse({'success': False, 'error': 'Category name is required'})
        
        try:
            category = SupplyCategory.objects.create(name=name)
            return JsonResponse({
                'success': True,
                'category': {
                    'id': category.id,
                    'name': category.name
                }
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@permission_required('app.view_admin_module')
@login_required
def add_subcategory(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        
        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'})
        
        try:
            subcategory = SupplySubcategory.objects.create(name=name)
            return JsonResponse({
                'success': True,
                'subcategory': {
                    'id': subcategory.id,
                    'name': subcategory.name
                }
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def get_subcategories(request):
    category_id = request.GET.get('category_id')
    if category_id:
        subcategories = SupplySubcategory.objects.filter(category_id=category_id)
        return JsonResponse({
            'subcategories': [
                {'id': sub.id, 'name': sub.name}
                for sub in subcategories
            ]
        })
    return JsonResponse({'subcategories': []})
   #new
@permission_required('app.view_admin_module')
@login_required
def update_property_category(request):
    if request.method == 'POST':
        category_id = request.POST.get('id')
        new_name = request.POST.get('name')
        new_uacs = request.POST.get('uacs')
       
        try:
            category = PropertyCategory.objects.get(id=category_id)
            old_name = category.name
            old_uacs = category.uacs
            
            # Update name
            category.name = new_name
            
            # Update UACS - convert to int if provided, otherwise set to None
            if new_uacs and new_uacs.strip():
                try:
                    category.uacs = int(new_uacs)
                except ValueError:
                    return JsonResponse({'success': False, 'error': 'UACS must be a valid number'})
            else:
                category.uacs = None
            
            category.save()
           
            # Log the activity with details
            changes = []
            if old_name != new_name:
                changes.append(f"name from '{old_name}' to '{new_name}'")
            if old_uacs != category.uacs:
                changes.append(f"UACS from '{old_uacs}' to '{category.uacs}'")
            
            description = f"Updated category {' and '.join(changes)}" if changes else f"Updated category '{category.name}'"
            
            ActivityLog.log_activity(
                user=request.user,
                action='update',
                model_name='PropertyCategory',
                object_repr=str(category),
                description=description
            )
           
            return JsonResponse({'success': True})
        except PropertyCategory.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Category not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
           
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@permission_required('app.view_admin_module')
@login_required
def delete_property_category(request, category_id):
    if request.method == 'POST':
        try:
            category = PropertyCategory.objects.get(id=category_id)
            category_name = category.name
            category.delete()
           
            # Log the activity
            ActivityLog.log_activity(
                user=request.user,
                action='delete',
                model_name='PropertyCategory',
                object_repr=category_name,
                description=f"Deleted property category '{category_name}'"
            )
           
            messages.success(request, 'Category deleted successfully.')
        except PropertyCategory.DoesNotExist:
            messages.error(request, 'Category not found.')
        except Exception as e:
            messages.error(request, f'Error deleting category: {str(e)}')
           
    return redirect('property_list')

@permission_required('app.view_admin_module')
def export_supply_to_excel(request):
    """Export supply data to Excel with filters"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Supply Inventory"

    # Title and Header Styling
    title_font = ws['A1'].font.copy(bold=True, size=16)
    header_font = ws['A1'].font.copy(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="152D64", end_color="152D64", fill_type="solid")
    thin_border = Side(border_style="thin", color="000000")
    category_font = ws['A1'].font.copy(bold=True, size=12)

    # Get selected fields from request
    selected_fields = request.POST.getlist('fields', [
        'barcode', 'supply_name', 'description', 'unit', 'category', 'subcategory',
        'current_quantity', 'status', 'date_received', 'expiration_date', 'available_for_request'
    ])

    # Define all possible fields and their display names
    field_mapping = {
        'barcode': 'Supply ID',
        'supply_name': 'Supply Name',
        'description': 'Description',
        'unit': 'Unit of Measure',
        'category': 'Category',
        'subcategory': 'Sub Category',
        'current_quantity': 'Current Quantity',
        'status': 'Stock Status',
        'date_received': 'Date Received',
        'expiration_date': 'Expiration Date',
        'available_for_request': 'Available for Request'
    }

    # Create header row with title and metadata
    num_columns = len(selected_fields)
    merge_range = f'A1:{chr(64 + num_columns)}1'
    ws.merge_cells(merge_range)
    ws['A1'] = 'INVENTORY REPORT FOR SUPPLY'
    ws['A1'].font = title_font
    ws['A1'].alignment = ws['A1'].alignment.copy(horizontal='center')

    # Add metadata
    ws['A3'] = 'Department:'
    ws['B3'] = request.user.userprofile.department.name if hasattr(request.user, 'userprofile') and request.user.userprofile.department else '_____________________'
    ws['G3'] = 'Date:'
    ws['H3'] = datetime.now().strftime("%B %d, %Y")
    
    ws['A4'] = 'Prepared by:'
    ws['B4'] = f'{request.user.first_name} {request.user.last_name}'
    ws['G4'] = 'Page:'
    ws['H4'] = '1 of 1'

    # Headers start at row 6
    ws['A6'] = 'SUPPLY INVENTORY'
    ws['A6'].font = ws['A6'].font.copy(bold=True)

    # Get all supplies with related data and group by category
    supplies = Supply.objects.select_related(
        'category', 'subcategory', 'quantity_info'
    ).order_by('category__name', 'supply_name').all()

    # Group supplies by category
    supplies_by_category = {}
    for supply in supplies:
        category_name = supply.category.name if supply.category else 'Uncategorized'
        if category_name not in supplies_by_category:
            supplies_by_category[category_name] = []
        supplies_by_category[category_name].append(supply)

    current_row = 8  # Start after the title and metadata

    # Process each category
    for category_name, category_supplies in supplies_by_category.items():
        # Add category header
        ws.cell(row=current_row, column=1, value=category_name)
        ws.cell(row=current_row, column=1).font = category_font
        current_row += 1

        # Add headers for this category
        headers = [field_mapping[field] for field in selected_fields]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        current_row += 1

        # Add data for this category
        for supply in category_supplies:
            row_data = []
            for field in selected_fields:
                if field == 'barcode':
                    row_data.append(supply.barcode or 'N/A')
                elif field == 'supply_name':
                    row_data.append(supply.supply_name)
                elif field == 'description':
                    row_data.append(supply.description or 'N/A')
                elif field == 'unit':
                    row_data.append(supply.unit or 'N/A')
                elif field == 'category':
                    row_data.append(supply.category.name if supply.category else 'N/A')
                elif field == 'subcategory':
                    row_data.append(supply.subcategory.name if supply.subcategory else 'N/A')
                elif field == 'current_quantity':
                    row_data.append(supply.quantity_info.current_quantity if supply.quantity_info else 0)
                elif field == 'status':
                    row_data.append(supply.get_status_display)
                elif field == 'date_received':
                    row_data.append(supply.date_received.strftime('%Y-%m-%d') if supply.date_received else 'N/A')
                elif field == 'expiration_date':
                    row_data.append(supply.expiration_date.strftime('%Y-%m-%d') if supply.expiration_date else 'N/A')
                elif field == 'available_for_request':
                    row_data.append('Yes' if supply.available_for_request else 'No')

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
            current_row += 1

        current_row += 1  # Add space between categories

    # Add signature section
    signature_row = current_row + 2
    ws.cell(row=signature_row, column=1, value='Prepared by:')
    ws.cell(row=signature_row, column=4, value='Reviewed by:')
    ws.cell(row=signature_row, column=7, value='Approved by:')

    ws.cell(row=signature_row + 2, column=1, value='_____________________')
    ws.cell(row=signature_row + 2, column=4, value='_____________________')
    ws.cell(row=signature_row + 2, column=7, value='_____________________')

    ws.cell(row=signature_row + 3, column=1, value='Inventory Officer')
    ws.cell(row=signature_row + 3, column=4, value='Department Head')
    ws.cell(row=signature_row + 3, column=7, value='Property Custodian')

    # Add notes section
    notes_row = signature_row + 5
    ws.cell(row=notes_row, column=1, value='Notes:')
    ws.cell(row=notes_row + 1, column=1, value='1. This report shows the current inventory status of supplies.')
    ws.cell(row=notes_row + 2, column=1, value='2. Please verify physical count against this report.')
    ws.cell(row=notes_row + 3, column=1, value='3. Report any discrepancies to the inventory officer.')

    # Auto-adjust column widths
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for cell in column[1:]:
            if isinstance(cell, MergedCell):
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=supply_inventory_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response

@login_required
def generate_quantity_activity_report(request):
    """
    Generate a PDF or Excel report for supply quantity activity.
    Supports filtering by date range, month, or all time.
    """
    try:
        # Get parameters
        supply_id = request.GET.get('supply_id')
        filter_type = request.GET.get('filter_type', 'all')
        user_filter = request.GET.get('user_filter', 'all')
        activity_type_filter = request.GET.get('activity_type_filter', 'all')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        month = request.GET.get('month')
        year = request.GET.get('year')
        format_type = request.GET.get('format', 'pdf')  # 'pdf' or 'excel'
        
        if not supply_id:
            return JsonResponse({'error': 'Supply ID is required'}, status=400)
        
        # Get supply
        supply = get_object_or_404(Supply, id=supply_id)
        
        # Filter quantity activity
        quantity_history = supply.history.filter(
            field_name__in=['quantity', 'current_quantity', 'initial_quantity']
        ).order_by('-timestamp')
        
        # Apply user filter (filter by requestor name in remarks)
        if user_filter != 'all' and user_filter:
            quantity_history = quantity_history.filter(remarks__icontains=f'Supply request by {user_filter}')
        
        # Apply activity type filter
        if activity_type_filter != 'all':
            if activity_type_filter == 'supply_request':
                quantity_history = quantity_history.filter(remarks__icontains='Supply request by')
            elif activity_type_filter == 'manual':
                quantity_history = quantity_history.exclude(remarks__icontains='Supply request by').exclude(action='create')
            elif activity_type_filter == 'addition':
                # For additions, we need to check if new_value > old_value
                addition_entries = []
                for entry in quantity_history:
                    old_qty = int(entry.old_value) if entry.old_value and entry.old_value.isdigit() else 0
                    new_qty = int(entry.new_value) if entry.new_value and entry.new_value.isdigit() else 0
                    if entry.action == 'create' or new_qty > old_qty:
                        addition_entries.append(entry.id)
                quantity_history = quantity_history.filter(id__in=addition_entries)
            elif activity_type_filter == 'deduction':
                # For deductions, we need to check if new_value < old_value
                deduction_entries = []
                for entry in quantity_history:
                    old_qty = int(entry.old_value) if entry.old_value and entry.old_value.isdigit() else 0
                    new_qty = int(entry.new_value) if entry.new_value and entry.new_value.isdigit() else 0
                    if entry.action != 'create' and new_qty < old_qty:
                        deduction_entries.append(entry.id)
                quantity_history = quantity_history.filter(id__in=deduction_entries)
        
        # Apply date filters
        if filter_type == 'date_range' and start_date and end_date:
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                quantity_history = quantity_history.filter(
                    timestamp__range=[start_dt, end_dt]
                )
            except ValueError:
                return JsonResponse({'error': 'Invalid date format'}, status=400)
        elif filter_type == 'month' and month and year:
            try:
                month_int = int(month)
                year_int = int(year)
                quantity_history = quantity_history.filter(
                    timestamp__year=year_int,
                    timestamp__month=month_int
                )
            except ValueError:
                return JsonResponse({'error': 'Invalid month or year'}, status=400)
        
        # Process activity data
        activity_data = []
        total_additions = 0
        total_deductions = 0
        
        for entry in quantity_history:
            try:
                old_qty = int(entry.old_value) if entry.old_value and entry.old_value.isdigit() else 0
                new_qty = int(entry.new_value) if entry.new_value and entry.new_value.isdigit() else 0
                quantity_change = new_qty - old_qty
                
                if entry.action == 'create':
                    action_type = 'Initial Creation'
                    quantity_change = new_qty
                elif quantity_change > 0:
                    action_type = 'Addition'
                    total_additions += quantity_change
                elif quantity_change < 0:
                    action_type = 'Deduction'
                    total_deductions += abs(quantity_change)
                else:
                    action_type = 'Adjustment'
                
                activity_data.append({
                    'date': entry.timestamp.strftime('%Y-%m-%d'),
                    'time': entry.timestamp.strftime('%H:%M:%S'),
                    'action': action_type,
                    'quantity_change': quantity_change,
                    'previous_quantity': old_qty,
                    'new_quantity': new_qty,
                    'user': entry.user.get_full_name() if entry.user and entry.user.get_full_name() else (entry.user.username if entry.user else 'System'),
                    'remarks': entry.remarks if entry.remarks else 'N/A'
                })
            except Exception as e:
                print(f"Error processing entry {entry.id}: {str(e)}")
                continue
        
        net_change = total_additions - total_deductions
        
        # Prepare filter info for the report
        filter_info = {
            'filter_type': filter_type,
            'user_filter': user_filter,
            'activity_type_filter': activity_type_filter
        }
        
        if format_type == 'excel':
            return _generate_excel_quantity_report(supply, activity_data, total_additions, total_deductions, net_change, filter_info)
        else:
            return _generate_pdf_quantity_report(supply, activity_data, total_additions, total_deductions, net_change, filter_info)
        
    except Supply.DoesNotExist:
        return JsonResponse({'error': 'Supply not found'}, status=404)
    except Exception as e:
        print(f"Error generating quantity activity report: {str(e)}")
        return JsonResponse({'error': 'An error occurred while generating the report'}, status=500)

def _generate_excel_quantity_report(supply, activity_data, total_additions, total_deductions, net_change, filter_info):
    """Generate Excel report for quantity activity"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Quantity Activity - {supply.supply_name}"
    
    # Set column widths
    ws.column_dimensions['A'].width = 15  # Date
    ws.column_dimensions['B'].width = 12  # Time
    ws.column_dimensions['C'].width = 18  # Action
    ws.column_dimensions['D'].width = 15  # Quantity Change
    ws.column_dimensions['E'].width = 15  # Previous Qty
    ws.column_dimensions['F'].width = 12  # New Qty
    ws.column_dimensions['G'].width = 20  # User
    ws.column_dimensions['H'].width = 30  # Remarks
    
    # Title and summary
    ws['A1'] = f"Quantity Activity Report - {supply.supply_name}"
    ws['A1'].font = Font(size=16, bold=True)
    
    # Add filter information
    row_offset = 3
    if filter_info['user_filter'] != 'all':
        ws[f'A{row_offset}'] = f"Filtered by Requestor: {filter_info['user_filter']}"
        row_offset += 1
    if filter_info['activity_type_filter'] != 'all':
        ws[f'A{row_offset}'] = f"Activity Type Filter: {filter_info['activity_type_filter']}"
        row_offset += 1
    
    ws[f'A{row_offset}'] = f"Total Additions: {total_additions}"
    ws[f'A{row_offset + 1}'] = f"Total Deductions: {total_deductions}"
    ws[f'A{row_offset + 2}'] = f"Net Change: {net_change}"
    ws[f'A{row_offset + 3}'] = f"Total Transactions: {len(activity_data)}"
    
    # Headers
    headers = ['Date', 'Time', 'Action', 'Quantity Change', 'Previous Qty', 'New Qty', 'User', 'Remarks']
    header_row = row_offset + 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data rows
    for row, activity in enumerate(activity_data, header_row + 1):
        ws.cell(row=row, column=1, value=activity['date'])
        ws.cell(row=row, column=2, value=activity['time'])
        ws.cell(row=row, column=3, value=activity['action'])
        ws.cell(row=row, column=4, value=activity['quantity_change'])
        ws.cell(row=row, column=5, value=activity['previous_quantity'])
        ws.cell(row=row, column=6, value=activity['new_quantity'])
        ws.cell(row=row, column=7, value=activity['user'])
        ws.cell(row=row, column=8, value=activity['remarks'])
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="quantity_activity_{supply.supply_name}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response

def _generate_pdf_quantity_report(supply, activity_data, total_additions, total_deductions, net_change, filter_info):
    """Generate PDF report for quantity activity"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="quantity_activity_{supply.supply_name}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    story.append(Paragraph(f"Quantity Activity Report - {supply.supply_name}", title_style))
    
    # Add filter information
    filter_text = []
    if filter_info['user_filter'] != 'all':
        filter_text.append(f"<b>Filtered by Requestor:</b> {filter_info['user_filter']}")
    if filter_info['activity_type_filter'] != 'all':
        filter_text.append(f"<b>Activity Type Filter:</b> {filter_info['activity_type_filter']}")
    
    if filter_text:
        filter_para = Paragraph("<br/>".join(filter_text), styles['Normal'])
        story.append(filter_para)
        story.append(Spacer(1, 12))
    
    # Summary section
    summary_data = [
        ['Summary', '', '', ''],
        ['Total Additions:', total_additions, 'Total Deductions:', total_deductions],
        ['Net Change:', net_change, 'Total Transactions:', len(activity_data)]
    ]
    
    summary_table = Table(summary_data, colWidths=[120, 80, 120, 80])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Activity table
    if activity_data:
        table_data = [['Date', 'Time', 'Action', 'Change', 'Previous', 'New', 'User', 'Remarks']]
        
        for activity in activity_data:
            table_data.append([
                activity['date'],
                activity['time'],
                activity['action'],
                str(activity['quantity_change']),
                str(activity['previous_quantity']),
                str(activity['new_quantity']),
                activity['user'][:15] + '...' if len(activity['user']) > 15 else activity['user'],
                activity['remarks'][:20] + '...' if len(activity['remarks']) > 20 else activity['remarks']
            ])
        
        table = Table(table_data, colWidths=[80, 60, 80, 50, 60, 50, 100, 120])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        story.append(table)
    else:
        story.append(Paragraph("No quantity activity found for the selected period.", styles['Normal']))
    
    doc.build(story)
    return response

@login_required
def export_property_to_pdf_ics(request):
    """Export properties with unit value below 50,000 as Excel Inventory Custodian Slip (ICS)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.utils import timezone
    
    # Get form data
    college_campus = request.POST.get('college_campus', 'BACOOR')
    accountable_person_filter = request.POST.get('accountable_person_ics', '(All)')
    
    # Get all properties
    properties = Property.objects.select_related('category').all()
    
    # Filter properties with unit value (unit cost) below 50,000
    filtered_properties = []
    for prop in properties:
        unit_cost = float(prop.unit_value or 0)
        qty_per_card = int(prop.overall_quantity or 0)  # Use overall_quantity instead of quantity
        total_value = unit_cost * qty_per_card
        
        # Filter based on unit cost being below 50,000
        if unit_cost < 50000:
            # Apply accountable person filter if specified
            if accountable_person_filter == '(All)' or prop.accountable_person == accountable_person_filter:
                filtered_properties.append({
                    'article': prop.property_name or 'N/A',  # Property name as Article
                    'description': prop.description or '',
                    'property_number': prop.property_number or 'N/A',
                    'unit_of_measure': prop.unit_of_measure or 'unit',
                    'unit_cost': unit_cost,
                    'total_value': total_value,
                    'qty_per_card': qty_per_card,
                    'qty_per_physical_count': prop.quantity_per_physical_count or qty_per_card,
                    'remarks': 'test',
                    'accountable_person': prop.accountable_person or 'N/A',
                    'year_acquired': prop.year_acquired.strftime('%m/%d/%Y') if prop.year_acquired else 'N/A'
                })
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "ICS Below 50000"
    
    # Header styling
    header_fill = PatternFill(start_color="152d64", end_color="152d64", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:K1')
    ws['A1'] = 'List of Inventories / Inventory Custodian Slip (ICS)'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add info rows
    ws.merge_cells('A2:K2')
    ws['A2'] = f'College / Campus: {college_campus}'
    ws['A2'].alignment = Alignment(horizontal='left')
    
    ws.merge_cells('A3:K3')
    ws['A3'] = f'Accountable Person: {accountable_person_filter}'
    ws['A3'].alignment = Alignment(horizontal='left')
    
    ws.merge_cells('A4:K4')
    ws['A4'] = f'Generated on: {timezone.now().strftime("%B %d, %Y %I:%M %p")}'
    ws['A4'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Article', 'Description', 'Property Number', 'Unit Of Measure', 
               'Unit Cost', 'Total Value', 'QTY Per card', 'Qty Per Physical Count', 
               'Remarks', 'Accountable Person', 'Year Acquired']
    ws.append([])  # Empty row
    ws.append(headers)
    
    # Style headers (row 6)
    for cell in ws[6]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Add data
    if filtered_properties:
        for prop in filtered_properties:
            ws.append([
                prop['article'],
                prop['description'],
                prop['property_number'],
                prop['unit_of_measure'],
                prop['unit_cost'],
                prop['total_value'],
                prop['qty_per_card'],
                prop['qty_per_physical_count'],
                prop['remarks'],
                prop['accountable_person'],
                prop['year_acquired']
            ])
    else:
        ws.append(['No properties found with unit cost below ₱50,000.00', '', '', '', '', '', '', '', '', '', ''])
    
    # Style data rows
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Format number columns
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=5, max_col=6):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Center align specific columns
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
        # Property Number
        row[2].alignment = Alignment(horizontal='center', vertical='center')
        # Unit of Measure
        row[3].alignment = Alignment(horizontal='center', vertical='center')
        # QTY columns
        row[6].alignment = Alignment(horizontal='center', vertical='center')
        row[7].alignment = Alignment(horizontal='center', vertical='center')
    
    # Adjust column widths
    column_widths = [20, 30, 18, 15, 12, 12, 12, 18, 15, 20, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ICS_Below_50000_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response


@login_required
def export_inventory_count_form_cvsu(request):
    """
    Generate Excel Inventory Count Form matching CvSU format.
    Filters items with unit value >= 50,000 and groups them by PPE Account Group (category).
    All categories are in one sheet with footer sections between them.
    """
    from openpyxl.styles import Alignment, Border, Side, Font as OpenpyxlFont
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from PIL import Image as PILImage
    from collections import OrderedDict
    
    # Filter properties with unit_value at or above 50,000
    properties = Property.objects.filter(
        unit_value__gte=50000
    ).select_related('category').order_by('category__name', 'property_name')
    
    # Check if there are any properties
    if not properties.exists():
        messages.warning(request, 'No properties with unit value at or above ₱50,000.00 were found. Please ensure you have properties meeting this criteria in your inventory.')
        return redirect('property')
    
    # Group properties by category
    properties_by_category = OrderedDict()
    for prop in properties:
        if prop.category:
            category_key = prop.category.id
            if category_key not in properties_by_category:
                properties_by_category[category_key] = {
                    'category': prop.category,
                    'properties': []
                }
            properties_by_category[category_key]['properties'].append(prop)
        else:
            # Handle properties without category
            if 'uncategorized' not in properties_by_category:
                properties_by_category['uncategorized'] = {
                    'category': None,
                    'properties': []
                }
            properties_by_category['uncategorized']['properties'].append(prop)
    
    # Get optional column selections from request
    include_accountable_person = request.POST.get('include_accountable_person') == 'yes'
    include_year_acquired = request.POST.get('include_year_acquired') == 'yes'
    
    # Calculate total columns and last column letter
    base_columns = 11  # A-K (Article to Remarks)
    total_columns = base_columns
    if include_accountable_person:
        total_columns += 1
    if include_year_acquired:
        total_columns += 1
    last_col_letter = chr(64 + total_columns)  # Convert to letter (A=65, so A=1+64)
    
    # Create workbook with single sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Count Form"
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_font = OpenpyxlFont(name='Calibri', size=10, bold=True)
    normal_font = OpenpyxlFont(name='Calibri', size=9)
    title_font = OpenpyxlFont(name='Calibri', size=11, bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 18  # Article/Item
    ws.column_dimensions['B'].width = 35  # Description
    ws.column_dimensions['C'].width = 15  # Old Property No.
    ws.column_dimensions['D'].width = 15  # New Property No.
    ws.column_dimensions['E'].width = 10  # Unit of Measure
    ws.column_dimensions['F'].width = 12  # Unit Value
    ws.column_dimensions['G'].width = 10  # Qty per Property Card
    ws.column_dimensions['H'].width = 10  # Qty per Physical Count
    ws.column_dimensions['I'].width = 20  # Location
    ws.column_dimensions['J'].width = 15  # Condition
    ws.column_dimensions['K'].width = 15  # Remarks
    
    # Set widths for optional columns
    current_col_index = 12  # Start after K (column 11)
    if include_accountable_person:
        ws.column_dimensions[chr(64 + current_col_index)].width = 18  # Accountable Person
        current_col_index += 1
    if include_year_acquired:
        ws.column_dimensions[chr(64 + current_col_index)].width = 12  # Year Acquired
    
    current_row = 1
    
    # Process each category
    for idx, (cat_key, cat_data) in enumerate(properties_by_category.items()):
        category = cat_data['category']
        category_properties = cat_data['properties']
        
        # Add header for each category section
        # Add logo for EVERY section - position it on the left side overlapping rows 1-5
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'cvsu logo.png')
        if os.path.exists(logo_path):
            try:
                # Load and resize logo
                img = OpenpyxlImage(logo_path)
                # Resize logo to approximately 80x80 pixels to span multiple rows
                img.width = 80
                img.height = 80
                # Position logo at C row (left area, will span across rows 1-5)
                ws.add_image(img, f'C{current_row}')
            except Exception as e:
                print(f"Error adding logo: {e}")
        
        # Row 1: Republic of the Philippines
        ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = 'Republic of the Philippines'
        cell.font = OpenpyxlFont(name='Calibri', size=10, bold=False)
        cell.alignment = center_alignment
        current_row += 1
        
        # Row 2: CAVITE STATE UNIVERSITY
        ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = 'CAVITE STATE UNIVERSITY'
        cell.font = OpenpyxlFont(name='Calibri', size=12, bold=True)
        cell.alignment = center_alignment
        current_row += 1
        
        # Row 3: Don Severino de las Alas Campus
        ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = 'Don Severino de las Alas Campus'
        cell.font = OpenpyxlFont(name='Calibri', size=10, bold=True)
        cell.alignment = center_alignment
        current_row += 1
        
        # Row 4: Indang, Cavite
        ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = 'Indang, Cavite'
        cell.font = OpenpyxlFont(name='Calibri', size=10, bold=False)
        cell.alignment = center_alignment
        current_row += 1
        
        # Row 5: www.cvsu.edu.ph
        ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = 'www.cvsu.edu.ph'
        cell.font = OpenpyxlFont(name='Calibri', size=9, bold=False, color='0000FF', underline='single')
        cell.alignment = center_alignment
        current_row += 1
        
        # Blank row
        current_row += 1
        
        # Inventory Count Form title
        ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = 'Inventory Count Form'
        cell.font = OpenpyxlFont(name='Calibri', size=12, bold=True)
        cell.alignment = center_alignment
        current_row += 1
        
        # Blank row
        current_row += 1
        
        # PPE Account Group - Label in column A
        cell = ws.cell(row=current_row, column=1)
        cell.value = 'PPE Account Group:'
        cell.font = OpenpyxlFont(name='Calibri', size=10, bold=True)
        cell.alignment = left_alignment
        
        # UACS + Category in columns B-C (merged with bottom border)
        ws.merge_cells(f'B{current_row}:C{current_row}')
        cell_b = ws.cell(row=current_row, column=2)
        cell_c = ws.cell(row=current_row, column=3)
        
        if category and category.uacs:
            cell_b.value = f'{category.uacs} - {category.name}'
        elif category:
            cell_b.value = f'{category.name}'
        else:
            cell_b.value = 'Uncategorized'
        
        cell_b.font = OpenpyxlFont(name='Calibri', size=10, bold=False)
        cell_b.alignment = left_alignment
        # Add bottom border to both B and C cells
        cell_b.border = Border(bottom=Side(style='thin', color='000000'))
        cell_c.border = Border(bottom=Side(style='thin', color='000000'))
        current_row += 1
        
        # Blank row
        current_row += 1
        
        # Table Headers (base columns)
        headers = [
            'Article/Item',
            'Description',
            'Old Property No.\nassigned',
            'New Property No.\nassigned\n(to be filled up during\nvalidation)',
            'Unit of\nMeasure',
            'Unit Value',
            'Quantity per\nProperty Card',
            'Quantity per\nPhysical Count',
            'Location/Whereabouts\n(Building, Floor and\nRoom No.)',
            'Condition\n(in good condition,\nneeding repair,\nunserviceable,\nobsolete, etc.)',
            'Remarks\n(Non-existing\nor Missing)'
        ]
        
        # Add optional headers
        if include_accountable_person:
            headers.append('Accountable Person')
        if include_year_acquired:
            headers.append('Year Acquired')
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Set row height for header
        ws.row_dimensions[current_row].height = 60
        current_row += 1
        
        # Data rows
        for prop in category_properties:
            # Article/Item
            cell = ws.cell(row=current_row, column=1)
            cell.value = prop.property_name or ''
            cell.font = normal_font
            cell.alignment = left_alignment
            cell.border = thin_border
            
            # Description
            cell = ws.cell(row=current_row, column=2)
            cell.value = prop.description or ''
            cell.font = normal_font
            cell.alignment = left_alignment
            cell.border = thin_border
            
            # Old Property No.
            cell = ws.cell(row=current_row, column=3)
            cell.value = prop.old_property_number or ''
            cell.font = normal_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # New Property No.
            cell = ws.cell(row=current_row, column=4)
            cell.value = prop.property_number or ''
            cell.font = normal_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # Unit of Measure
            cell = ws.cell(row=current_row, column=5)
            cell.value = prop.unit_of_measure or ''
            cell.font = normal_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # Unit Value
            cell = ws.cell(row=current_row, column=6)
            cell.value = float(prop.unit_value) if prop.unit_value else 0
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = thin_border
            cell.number_format = '#,##0.00'
            
            # Quantity per Property Card
            cell = ws.cell(row=current_row, column=7)
            cell.value = prop.quantity or 0
            cell.font = normal_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # Quantity per Physical Count
            cell = ws.cell(row=current_row, column=8)
            cell.value = prop.quantity_per_physical_count or 0
            cell.font = normal_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # Location/Whereabouts
            cell = ws.cell(row=current_row, column=9)
            cell.value = prop.location or ''
            cell.font = normal_font
            cell.alignment = left_alignment
            cell.border = thin_border
            
            # Condition
            cell = ws.cell(row=current_row, column=10)
            cell.value = prop.condition or ''
            cell.font = normal_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # Remarks
            cell = ws.cell(row=current_row, column=11)
            cell.value = ''  # Empty for manual entry
            cell.font = normal_font
            cell.alignment = left_alignment
            cell.border = thin_border
            
            # Optional columns - dynamically add after Remarks
            opt_col = 12  # Start after column 11 (Remarks)
            if include_accountable_person:
                cell = ws.cell(row=current_row, column=opt_col)
                cell.value = prop.accountable_person or ''
                cell.font = normal_font
                cell.alignment = left_alignment
                cell.border = thin_border
                opt_col += 1
            
            if include_year_acquired:
                cell = ws.cell(row=current_row, column=opt_col)
                cell.value = prop.year_acquired.strftime('%Y') if prop.year_acquired else ''
                cell.font = normal_font
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # Set row height
            ws.row_dimensions[current_row].height = 30
            current_row += 1
        
        # Add footer section after each category
        # Blank row
        current_row += 1
        
        # Note section - "Note:" in column A (bold), rest of text in B-lastcol (not bold)
        cell = ws.cell(row=current_row, column=1)
        cell.value = 'Note:'
        cell.font = OpenpyxlFont(name='Calibri', size=9, bold=True)
        cell.alignment = left_alignment
        
        # Note content text - merge dynamically based on total columns
        ws.merge_cells(f'B{current_row}:{last_col_letter}{current_row}')
        cell = ws[f'B{current_row}']
        cell.value = 'for PPE items without Property No., provide in the "Remarks" column other information such as Serial No./Model No./brief description that can be useful during the reconciliation process.'
        cell.font = OpenpyxlFont(name='Calibri', size=9, bold=False)
        cell.alignment = left_alignment
        current_row += 1
        
        # Blank rows for spacing
        current_row += 2
        
        # Prepared by section (LEFT SIDE - Columns A-C)
        prepared_row = current_row
        cell = ws.cell(row=current_row, column=1)
        cell.value = 'Prepared by:'
        cell.font = OpenpyxlFont(name='Calibri', size=9, bold=True)
        cell.alignment = left_alignment
        
        # Reviewed by section (RIGHT SIDE - Column G)
        cell = ws.cell(row=current_row, column=7)
        cell.value = 'Reviewed by:'
        cell.font = OpenpyxlFont(name='Calibri', size=9, bold=True)
        cell.alignment = left_alignment
        current_row += 1
        
        # Blank row before signatures
        current_row += 1
        
        # Signature line for Prepared by (bottom border only in column B)
        cell = ws.cell(row=current_row, column=2)
        cell.border = Border(bottom=Side(style='thin', color='000000'))
        cell.alignment = center_alignment
        
        # Signature line for Reviewed by (bottom border in columns H-J)
        ws.merge_cells(f'H{current_row}:J{current_row}')
        for col in range(8, 11):  # Columns H, I, J
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(bottom=Side(style='thin', color='000000'))
            cell.alignment = center_alignment
        current_row += 1
        
        # Title for Prepared by (only in column B)
        cell = ws.cell(row=current_row, column=2)
        cell.value = 'Concerned Inventory Committee Member'
        cell.font = OpenpyxlFont(name='Calibri', size=9)
        cell.alignment = center_alignment
        
        # Title for Reviewed by (merged in H-J)
        ws.merge_cells(f'H{current_row}:J{current_row}')
        cell = ws.cell(row=current_row, column=8)
        cell.value = 'Chairman, Inventory Committee'
        cell.font = OpenpyxlFont(name='Calibri', size=9)
        cell.alignment = center_alignment
        current_row += 1
        
        # Blank rows
        current_row += 2
        
        # Date section (LEFT SIDE)
        cell = ws.cell(row=current_row, column=1)
        cell.value = 'Date:'
        cell.font = OpenpyxlFont(name='Calibri', size=9, bold=True)
        cell.alignment = left_alignment
        current_row += 1
        
        # Blank row before date line
        current_row += 1
        
        # Date line for left side (bottom border only in column A)
        cell = ws.cell(row=current_row, column=1)
        cell.border = Border(bottom=Side(style='thin', color='000000'))
        cell.alignment = center_alignment
        current_row += 1
        
        # Add spacing before next category (7 blank rows)
        if idx < len(properties_by_category) - 1:
            current_row += 7
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Inventory_Count_Over50k_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


@login_required
def get_supply_by_barcode(request, barcode):
    import logging
    import re
    logger = logging.getLogger(__name__)

    # Preserve original for logging, then normalize scanned barcode
    original_barcode = barcode
    # Replace non-breaking spaces and collapse whitespace, then strip
    barcode = (barcode or '').replace('\u00A0', ' ')
    barcode = re.sub(r"\s+", ' ', barcode).strip()

    logger.info(f"Barcode lookup - Original: {repr(original_barcode)} | Normalized: {repr(barcode)} | Length: {len(barcode)}")
    logger.info(f"Barcode bytes: {original_barcode.encode('utf-8')}")

    supply = None

    # 1) Try exact (case-sensitive) then case-insensitive exact
    try:
        supply = Supply.objects.get(barcode=barcode)
        logger.info(f"Found supply by exact barcode match: {supply.id}, archived: {supply.is_archived}")
    except Supply.DoesNotExist:
        try:
            supply = Supply.objects.get(barcode__iexact=barcode)
            logger.info(f"Found supply by case-insensitive barcode match: {supply.id}, archived: {supply.is_archived}")
        except Supply.DoesNotExist:
            logger.info(f"No direct barcode match for '{barcode}', attempting ID and fuzzy searches")

    # 2) If still not found, try extracting numeric ID from SUP-{id}
    if not supply and barcode.upper().startswith('SUP-'):
        try:
            supply_id = int(barcode.split('-')[1])
            logger.info(f"Attempting lookup by extracted ID: {supply_id}")
            supply = Supply.objects.get(id=supply_id)
            logger.info(f"Found supply by ID: {supply.id}, archived: {supply.is_archived}")
        except (IndexError, ValueError, Supply.DoesNotExist) as e:
            logger.debug(f"Lookup by extracted ID failed: {e}")

    # 3) Fallback: scan for exact barcode match (case-insensitive, trimmed) or barcode image filename
    # REMOVED partial "contains" matching to avoid matching wrong items on partial scans
    if not supply and len(barcode) >= 5:  # Only run fallback if barcode is reasonable length
        logger.info("Running fallback search for exact/filename matches")
        candidates = Supply.objects.filter(is_archived=False)[:500]  # Only non-archived
        scanned_lower = barcode.lower()
        for cand in candidates:
            try:
                # Exact match (case-insensitive, trimmed)
                cand_barcode = (cand.barcode or '').strip().lower()
                if cand_barcode and cand_barcode == scanned_lower:
                    supply = cand
                    logger.info(f"Matched by exact candidate barcode: {supply.id}")
                    break

                # If barcode image present, compare filename (SUP-26.png etc.)
                if cand.barcode_image:
                    img_name = cand.barcode_image.name.split('/')[-1].split('\\')[-1].split('.')[0]
                    if img_name and img_name.strip().lower() == scanned_lower:
                        supply = cand
                        logger.info(f"Matched by barcode image filename: {supply.id}")
                        break
            except Exception:
                continue

    # If still not found, return not found with log samples
    if not supply:
        logger.error(f"Supply not found for barcode after all attempts: {repr(barcode)}")
        try:
            sample = list(Supply.objects.filter(barcode__icontains='SUP').values_list('id', 'barcode')[:10])
            logger.info(f"Sample supplies: {sample}")
        except Exception:
            logger.exception('Error fetching sample supplies')
        return JsonResponse({
            'success': False,
            'error': f'Supply not found for barcode: {barcode}'
        })

    # Check if supply is archived
    if supply.is_archived:
        logger.info(f"Supply {supply.id} is archived; rejecting barcode use")
        return JsonResponse({
            'success': False,
            'error': f'Supply "{supply.supply_name}" is archived and cannot be used'
        })
    
    return JsonResponse({
        'success': True,
        'supply': {
            'id': supply.id,
            'name': supply.supply_name,
            'current_quantity': supply.quantity_info.current_quantity if hasattr(supply, 'quantity_info') else 0,
            'description': supply.description or ''
        }
    })

@login_required
def get_all_supply_barcodes(request):
    """
    Returns all supply barcodes as JSON for the barcode selection modal
    """
    try:
        # Get all non-archived supplies
        supplies = Supply.objects.filter(is_archived=False).order_by('supply_name')
        
        barcodes_data = []
        for supply in supplies:
            # Get barcode URL - prefer barcode_image if available
            barcode_url = supply.barcode_image.url if supply.barcode_image else supply.barcode
            
            # Get category name
            category_name = supply.category.name if supply.category else 'N/A'
            
            barcodes_data.append({
                'id': supply.id,
                'supply_name': supply.supply_name,
                'category': category_name,
                'barcode': barcode_url
            })
        
        return JsonResponse({
            'success': True,
            'count': len(barcodes_data),
            'barcodes': barcodes_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@permission_required('app.view_admin_module')
@require_POST
def modify_property_quantity_generic(request):
    try:
        property_id = request.POST.get("property_id")
        amount = int(request.POST.get("amount", 0))
        reason = request.POST.get("reason", "").strip()  # Get optional reason for deduction

        prop = get_object_or_404(Property, pk=property_id)
        old_quantity = prop.quantity

        # Always add the quantity since we removed the action type selection
        prop.quantity += amount
        prop.overall_quantity += amount
        prop.save()

        # Build remarks with reason if provided
        if amount < 0:
            # Deduction - include reason if provided
            if reason:
                remarks = f"Deducted {abs(amount)} units. Reason: {reason}. Modified by: {request.user.get_full_name() or request.user.username}"
            else:
                remarks = f"Deducted {abs(amount)} units. Modified by: {request.user.get_full_name() or request.user.username}"
        else:
            # Addition
            if reason:
                remarks = f"Added {amount} units. Reason: {reason}. Modified by: {request.user.get_full_name() or request.user.username}"
            else:
                remarks = f"Added {amount} units. Modified by: {request.user.get_full_name() or request.user.username}"

        PropertyHistory.objects.create(
            property=prop,
            user=request.user,
            action='quantity_update',
            field_name='quantity',
            old_value=str(old_quantity),
            new_value=str(prop.quantity),
            remarks=remarks
        )

        # Add activity log
        ActivityLog.log_activity(
            user=request.user,
            action='quantity_update',
            model_name='Property',
            object_repr=str(prop),
            description=f"{'Added' if amount > 0 else 'Deducted'} {abs(amount)} units to property '{prop.property_name}'. Changed quantity from {old_quantity} to {prop.quantity}" + (f". Reason: {reason}" if reason else "")
        )

        messages.success(request, f"Quantity successfully {'added' if amount > 0 else 'deducted'}.")
        
        # Return JSON response for AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Quantity updated successfully',
                'new_quantity': prop.quantity
            })
        return redirect('property_list')
        
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
        messages.error(request, f"Error modifying quantity: {str(e)}")
        return redirect('property_list')

@login_required
def get_property_by_barcode(request, barcode):
    # Strip whitespace from scanned barcode (handles padded barcodes)
    barcode = barcode.strip()
    
    # Check if request is AJAX/JSON
    if not request.headers.get('Content-Type') == 'application/json' and not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # For non-AJAX requests, handle authentication differently
        if not request.user.is_authenticated:
            return JsonResponse({
                'success': False,
                'error': 'Authentication required'
            }, status=401)
    
    try:
        # Try to find the property by the exact barcode first
        property = Property.objects.get(barcode=barcode)
    except Property.DoesNotExist:
        try:
            # If not found, try to find by property number
            property = Property.objects.get(property_number=barcode)
        except Property.DoesNotExist:
            try:
                # If still not found, try to find by ID (extract ID from PROP-{id} format)
                if barcode.startswith('PROP-'):
                    property_id = barcode.split('-')[1]
                    property = Property.objects.get(id=property_id)
                else:
                    raise Property.DoesNotExist
            except (Property.DoesNotExist, IndexError, ValueError):
                return JsonResponse({
                    'success': False,
                    'error': 'Property not found'
                })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Database error: {str(e)}'
        })
    
    return JsonResponse({
        'success': True,
        'property': {
            'id': property.id,
            'property_name': property.property_name,
            'property_number': property.property_number,
            'quantity': property.quantity
        }
    })

@login_required
def modify_property_quantity_batch(request):
    """
    Handle batch property quantity modifications from barcode scanning
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        items = data.get('items', [])
        
        if not items:
            return JsonResponse({'success': False, 'error': 'No items provided'})
        
        updated_count = 0
        errors = []
        
        for item_data in items:
            try:
                property_id = item_data.get('property_id')
                quantity_to_add = int(item_data.get('quantity_to_add', 0))
                
                if quantity_to_add <= 0:
                    continue
                    
                prop = get_object_or_404(Property, pk=property_id)
                old_quantity = prop.quantity
                
                # Add the quantity
                prop.quantity += quantity_to_add
                prop.overall_quantity += quantity_to_add
                prop.save()
                
                # Create history record
                PropertyHistory.objects.create(
                    property=prop,
                    user=request.user,
                    action='quantity_update',
                    field_name='quantity',
                    old_value=str(old_quantity),
                    new_value=str(prop.quantity),
                    remarks=f"Batch added {quantity_to_add} units. Modified by: {request.user.get_full_name() or request.user.username}"
                )
                
                # Add activity log
                ActivityLog.log_activity(
                    user=request.user,
                    action='quantity_update',
                    model_name='Property',
                    object_repr=str(prop),
                    description=f"Batch added {quantity_to_add} units to property '{prop.property_name}'. Changed quantity from {old_quantity} to {prop.quantity}"
                )
                
                updated_count += 1
                
            except Exception as e:
                errors.append(f"Error updating property {property_id}: {str(e)}")
                continue
        
        if updated_count > 0:
            return JsonResponse({
                'success': True,
                'updated_count': updated_count,
                'message': f'Successfully updated {updated_count} properties',
                'errors': errors if errors else None
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'No properties were updated',
                'errors': errors
            })
            
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def get_all_property_barcodes(request):
    """
    Returns all property barcodes as JSON for the barcode selection modal
    """
    try:
        # Get all non-archived properties
        properties = Property.objects.filter(is_archived=False).order_by('property_name')
        
        barcodes_data = []
        for prop in properties:
            # Get barcode URL - prefer barcode_image if available
            barcode_url = prop.barcode_image.url if prop.barcode_image else prop.barcode
            
            barcodes_data.append({
                'id': prop.id,
                'property_name': prop.property_name,
                'property_number': prop.property_number or 'Not assigned',
                'barcode': barcode_url
            })
        
        return JsonResponse({
            'success': True,
            'count': len(barcodes_data),
            'barcodes': barcodes_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@permission_required('app.view_admin_module')
@login_required
def archive_supply(request, pk):
    supply = get_object_or_404(Supply, pk=pk)
    if request.method == 'POST':
        # Allow archive if expired, otherwise only if quantity is zero
        if not supply.is_expired and hasattr(supply, 'quantity_info') and supply.quantity_info.current_quantity > 0:
            messages.error(request, f"Cannot archive supply '{supply.supply_name}' because it still has {supply.quantity_info.current_quantity} units.")
            return redirect('supply_list')
        
        supply.is_archived = True
        supply.save(user=request.user)
        
        ActivityLog.log_activity(
            user=request.user,
            action='archive',
            model_name='Supply',
            object_repr=str(supply),
            description=f"Archived supply '{supply.supply_name}'"
        )
        messages.success(request, 'Supply archived successfully.')
    return redirect('supply_list')

@permission_required('app.view_admin_module')
@login_required
def unarchive_supply(request, pk):
    supply = get_object_or_404(Supply, pk=pk)
    if request.method == 'POST':
        supply.is_archived = False
        supply.save(user=request.user)
        
        ActivityLog.log_activity(
            user=request.user,
            action='unarchive',
            model_name='Supply',
            object_repr=str(supply),
            description=f"Unarchived supply '{supply.supply_name}'"
        )
        messages.success(request, 'Supply unarchived successfully.')
    return redirect('archived_items')

@permission_required('app.view_admin_module')
@login_required
def archive_property(request, pk):
    property_obj = get_object_or_404(Property, pk=pk)
    if request.method == 'POST':
        # Check if property condition is Obsolete or Unserviceable
        if property_obj.condition not in ['Obsolete', 'Unserviceable']:
            messages.error(request, f"Cannot archive property '{property_obj.property_name}'. Property must be marked as 'Obsolete' or 'Unserviceable' to archive.")
            return redirect('property_list')
        
        property_obj.is_archived = True
        property_obj.save(user=request.user)
        
        ActivityLog.log_activity(
            user=request.user,
            action='archive',
            model_name='Property',
            object_repr=str(property_obj),
            description=f"Archived property '{property_obj.property_name}'"
        )
        messages.success(request, 'Property archived successfully.')
    return redirect('property_list')

@permission_required('app.view_admin_module')
@login_required
def unarchive_property(request, pk):
    property_obj = get_object_or_404(Property, pk=pk)
    if request.method == 'POST':
        property_obj.is_archived = False
        property_obj.save(user=request.user)
        
        ActivityLog.log_activity(
            user=request.user,
            action='unarchive',
            model_name='Property',
            object_repr=str(property_obj),
            description=f"Unarchived property '{property_obj.property_name}'"
        )
        messages.success(request, 'Property unarchived successfully.')
    return redirect('archived_items')

@permission_required('app.view_admin_module')
@login_required
@admin_permission_required('delete_archived_items')
def delete_archived_supply(request, pk):
    supply = get_object_or_404(Supply, pk=pk)
    if request.method == 'POST':
        # Check if the supply is archived before deleting
        if not supply.is_archived:
            messages.error(request, 'Cannot delete a supply that is not archived.')
            return redirect('archived_items')
        
        supply_name = supply.supply_name
        
        try:
            ActivityLog.log_activity(
                user=request.user,
                action='delete',
                model_name='Supply',
                object_repr=str(supply),
                description=f"Permanently deleted archived supply '{supply_name}'"
            )
            
            # Delete the supply and all related objects
            supply.delete()
            messages.success(request, f"Supply '{supply_name}' has been permanently deleted.")
        except Exception as e:
            messages.error(request, f"Error deleting supply: {str(e)}")
    return redirect('archived_items')

@permission_required('app.view_admin_module')
@login_required
@admin_permission_required('delete_archived_items')
def delete_archived_property(request, pk):
    property_obj = get_object_or_404(Property, pk=pk)
    if request.method == 'POST':
        # Check if the property is archived before deleting
        if not property_obj.is_archived:
            messages.error(request, 'Cannot delete a property that is not archived.')
            return redirect('archived_items')
        
        property_name = property_obj.property_name
        
        try:
            ActivityLog.log_activity(
                user=request.user,
                action='delete',
                model_name='Property',
                object_repr=str(property_obj),
                description=f"Permanently deleted archived property '{property_name}'"
            )
            
            # Delete the property and all related objects
            property_obj.delete()
            messages.success(request, f"Property '{property_name}' has been permanently deleted.")
        except Exception as e:
            messages.error(request, f"Error deleting property: {str(e)}")
    return redirect('archived_items')

class ArchivedItemsView(PermissionRequiredMixin, TemplateView):
    template_name = 'app/archived_items.html'
    permission_required = 'app.view_admin_module'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get archived supplies
        supplies = Supply.objects.filter(is_archived=True).select_related(
            'category', 'subcategory', 'quantity_info'
        ).order_by('supply_name')
        
        # Barcodes are already stored in database, no need to generate
        # They will be displayed using barcode_image field
        
        # Group supplies by category
        supplies_by_category = defaultdict(list)
        for supply in supplies:
            category_name = supply.category.name if supply.category else 'Uncategorized'
            supplies_by_category[category_name].append(supply)
        
        # Get archived properties
        properties = Property.objects.filter(is_archived=True).select_related('category').order_by('property_name')
        
        # Barcodes are already stored in database, no need to generate
        # They will be displayed using barcode_image field
        
        # Group properties by category
        properties_by_category = defaultdict(list)
        for prop in properties:
            category_name = prop.category.name if prop.category else 'Uncategorized'
            properties_by_category[category_name].append(prop)
        
        context.update({
            'supplies_by_category': dict(supplies_by_category),
            'properties_by_category': dict(properties_by_category),
        })
        return context

@permission_required('app.view_admin_module')
@login_required
def update_supply_category(request):
    if request.method == 'POST':
        category_id = request.POST.get('id')
        new_name = request.POST.get('name')
       
        try:
            category = SupplyCategory.objects.get(id=category_id)
            old_name = category.name
            category.name = new_name
            category.save()
           
            # Log the activity
            ActivityLog.log_activity(
                user=request.user,
                action='update',
                model_name='SupplyCategory',
                object_repr=str(category),
                description=f"Updated category name from '{old_name}' to '{new_name}'"
            )
           
            # Return more data for UI update
            return JsonResponse({
                'success': True,
                'category': {
                    'id': category.id,
                    'name': category.name,
                    'supply_count': category.supply_set.count()
                }
            })
        except SupplyCategory.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Category not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
           
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@permission_required('app.view_admin_module')
@login_required
def delete_supply_category(request, category_id):
    if request.method == 'POST':
        try:
            category = SupplyCategory.objects.get(id=category_id)
            category_name = category.name
            category.delete()
           
            # Log the activity
            ActivityLog.log_activity(
                user=request.user,
                action='delete',
                model_name='SupplyCategory',
                object_repr=category_name,
                description=f"Deleted supply category '{category_name}'"
            )
           
            messages.success(request, 'Category deleted successfully.')
        except SupplyCategory.DoesNotExist:
            messages.error(request, 'Category not found.')
        except Exception as e:
            messages.error(request, f'Error deleting category: {str(e)}')
           
    return redirect('supply_list')

@permission_required('app.view_admin_module')
@login_required
def update_supply_subcategory(request):
    if request.method == 'POST':
        subcategory_id = request.POST.get('id')
        new_name = request.POST.get('name')
       
        try:
            subcategory = SupplySubcategory.objects.get(id=subcategory_id)
            old_name = subcategory.name
            subcategory.name = new_name
            subcategory.save()
           
            # Log the activity
            ActivityLog.log_activity(
                user=request.user,
                action='update',
                model_name='SupplySubcategory',
                object_repr=str(subcategory),
                description=f"Updated subcategory name from '{old_name}' to '{new_name}'"
            )
           
            # Return more data for UI update
            return JsonResponse({
                'success': True,
                'subcategory': {
                    'id': subcategory.id,
                    'name': subcategory.name,
                    'supply_count': subcategory.supply_set.count()
                }
            })
        except SupplySubcategory.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Subcategory not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
           
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@permission_required('app.view_admin_module')
@login_required
def delete_supply_subcategory(request, subcategory_id):
    if request.method == 'POST':
        try:
            print(f"Attempting to delete subcategory {subcategory_id}")
            subcategory = SupplySubcategory.objects.get(id=subcategory_id)
            subcategory_name = subcategory.name
           
            # Check if there are any supplies using this subcategory
            supply_count = subcategory.supply_set.count()
            print(f"Subcategory {subcategory_name} has {supply_count} supplies")
           
            if subcategory.supply_set.exists():
                print(f"Cannot delete subcategory {subcategory_name} - has supplies assigned")
                messages.error(request, 'Cannot delete subcategory that has supplies assigned to it.')
                return redirect('supply_list')
               
            subcategory.delete()
            print(f"Successfully deleted subcategory {subcategory_name}")
           
            # Log the activity
            ActivityLog.log_activity(
                user=request.user,
                action='delete',
                model_name='SupplySubcategory',
                object_repr=subcategory_name,
                description=f"Deleted supply subcategory '{subcategory_name}'"
            )
           
            messages.success(request, 'Subcategory deleted successfully.')
        except SupplySubcategory.DoesNotExist:
            print(f"Subcategory {subcategory_id} not found")
            messages.error(request, 'Subcategory not found.')
        except Exception as e:
            print(f"Error deleting subcategory: {str(e)}")
            messages.error(request, f'Error deleting subcategory: {str(e)}')
           
    return redirect('supply_list')

def supply_list(request):
    # Redirect to the class-based view
    return redirect('supplies')


# Batch Request Item Management Views
@permission_required('app.view_admin_module')
@login_required
@require_POST
@admin_permission_required('approve_supply_request')
def approve_batch_item(request, batch_id, item_id):
    """Approve an individual item in a batch request"""
    batch_request = get_object_or_404(SupplyRequestBatch, id=batch_id)
    item = get_object_or_404(SupplyRequestItem, id=item_id, batch_request=batch_request)
    scroll_position = request.POST.get('scroll_position', '0')
    
    # Get approved quantity from form
    approved_quantity = request.POST.get('approved_quantity', item.quantity)
    try:
        approved_quantity = int(approved_quantity)
        if approved_quantity <= 0 or approved_quantity > item.quantity:
            messages.error(request, f'Invalid approved quantity. Must be between 1 and {item.quantity}.')
            response = redirect('batch_request_detail', batch_id=batch_id)
            response['Location'] += f'#scroll-{scroll_position}'
            return response
    except (ValueError, TypeError):
        approved_quantity = item.quantity
    
    # CHECK AVAILABLE STOCK BEFORE APPROVING
    quantity_info = item.supply.quantity_info
    available_qty = quantity_info.available_quantity
    
    if approved_quantity > available_qty:
        error_msg = f'Cannot approve {approved_quantity} units of {item.supply.supply_name}. Only {available_qty} units available (Current: {quantity_info.current_quantity}, Reserved: {quantity_info.reserved_quantity}).'
        
        # Check if this is an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': error_msg
            })
        else:
            messages.error(request, error_msg)
            response = redirect('batch_request_detail', batch_id=batch_id)
            response['Location'] += f'#scroll-{scroll_position}'
            return response
    
    # Mark item as approved
    item.approved = True
    item.status = 'approved'  # Also update the status field
    item.approved_quantity = approved_quantity
    remarks = request.POST.get('remarks', '')
    if remarks:
        item.remarks = remarks
    item.save()
    
    # Reserve the approved quantity to prevent overbooking
    quantity_info.reserved_quantity += approved_quantity
    quantity_info.save(user=request.user)
    
    # Check if all items in the batch are processed
    total_items = batch_request.items.count()
    approved_items = batch_request.items.filter(approved=True).count()
    rejected_items = batch_request.items.filter(approved=False, status='rejected').count()
    processed_items = approved_items + rejected_items
    
    # Update batch status
    if processed_items == total_items:
        # All items have been processed (approved or rejected)
        if approved_items > 0:
            batch_request.status = 'for_claiming'
            # Set approved_date if not already set
            if not batch_request.approved_date:
                batch_request.approved_date = timezone.now()
                batch_request.approved_by = request.user
        else:
            batch_request.status = 'rejected'
    elif approved_items > 0:
        batch_request.status = 'partially_approved'
        # Set approved_date when first item is approved
        if not batch_request.approved_date:
            batch_request.approved_date = timezone.now()
            batch_request.approved_by = request.user
    
    batch_request.save()
    
    # Create notification for user (individual item notification)
    Notification.objects.create(
        user=batch_request.user,
        message=f"Item '{item.supply.supply_name}' in your batch request #{batch_request.id} has been approved for {item.approved_quantity} units.",
        remarks=remarks
    )
    
    # Send batch completion email ONLY if all items are now processed
    if processed_items == total_items:
        from .utils import send_batch_request_completion_email
        approved_items = batch_request.items.filter(status='approved')
        rejected_items = batch_request.items.filter(status='rejected')
        send_batch_request_completion_email(batch_request, approved_items, rejected_items)
    
    # Log activity
    ActivityLog.log_activity(
        user=request.user,
        action='approve',
        model_name='SupplyRequestItem',
        object_repr=f"Batch #{batch_id} - {item.supply.supply_name}",
        description=f"Approved {item.approved_quantity} out of {item.quantity} units of {item.supply.supply_name} in batch request #{batch_id}"
    )
    
    # Remove success message since user can see status change immediately on the page
    # messages.success(request, f'Item "{item.supply.supply_name}" approved successfully.')
    response = redirect('batch_request_detail', batch_id=batch_id)
    response['Location'] += f'#scroll-{scroll_position}'
    return response

@permission_required('app.view_admin_module')
@login_required
@require_POST
@admin_permission_required('approve_supply_request')
def reject_batch_item(request, batch_id, item_id):
    """Reject an individual item in a batch request"""
    batch_request = get_object_or_404(SupplyRequestBatch, id=batch_id)
    item = get_object_or_404(SupplyRequestItem, id=item_id, batch_request=batch_request)
    scroll_position = request.POST.get('scroll_position', '0')
    
    # If the item was previously approved, release the reserved quantity
    if item.status == 'approved' and item.approved_quantity:
        quantity_info = item.supply.quantity_info
        quantity_info.reserved_quantity = max(0, quantity_info.reserved_quantity - item.approved_quantity)
        quantity_info.save(user=request.user)
    
    # Mark item as not approved and add remarks
    item.approved = False
    item.status = 'rejected'  # Also update the status field
    remarks = request.POST.get('remarks', '')
    if remarks:
        item.remarks = remarks
    item.save()
    
    # Check if all items in the batch are processed
    total_items = batch_request.items.count()
    approved_items = batch_request.items.filter(approved=True).count()
    rejected_items = batch_request.items.filter(approved=False, status='rejected').count()
    processed_items = approved_items + rejected_items
    
    # Update batch status
    if processed_items == total_items:
        # All items have been processed (approved or rejected)
        if approved_items > 0:
            batch_request.status = 'for_claiming'
            # Set approved_date if not already set
            if not batch_request.approved_date:
                batch_request.approved_date = timezone.now()
        else:
            batch_request.status = 'rejected'
    elif approved_items > 0:
        batch_request.status = 'partially_approved'
        # Set approved_date when first item is approved
        if not batch_request.approved_date:
            batch_request.approved_date = timezone.now()
    
    batch_request.save()
    
    # Create notification for user (individual item notification)
    Notification.objects.create(
        user=batch_request.user,
        message=f"Item '{item.supply.supply_name}' in your batch request #{batch_request.id} has been rejected.",
        remarks=remarks
    )
    
    # Send batch completion email ONLY if all items are now processed
    if processed_items == total_items:
        from .utils import send_batch_request_completion_email
        approved_items = batch_request.items.filter(status='approved')
        rejected_items = batch_request.items.filter(status='rejected')
        send_batch_request_completion_email(batch_request, approved_items, rejected_items)
    
    # Log activity
    ActivityLog.log_activity(
        user=request.user,
        action='reject',
        model_name='SupplyRequestItem',
        object_repr=f"Batch #{batch_id} - {item.supply.supply_name}",
        description=f"Rejected {item.quantity} units of {item.supply.supply_name} in batch request #{batch_id}"
    )
    
    # Remove success message since user can see status change immediately on the page
    # messages.success(request, f'Item "{item.supply.supply_name}" rejected successfully.')
    response = redirect('batch_request_detail', batch_id=batch_id)
    response['Location'] += f'#scroll-{scroll_position}'
    return response

@permission_required('app.view_admin_module')
@login_required
def batch_request_detail(request, batch_id):
    """View detailed information about a batch request with individual item actions"""
    batch_request = get_object_or_404(SupplyRequestBatch, id=batch_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        item_id = request.POST.get('item_id')
        remarks = request.POST.get('remarks', '').strip()
        
        # Handle void action for for_claiming requests
        if action == 'void_request' and batch_request.status == 'for_claiming':
            # Check permission
            from .permissions import has_admin_permission
            if not has_admin_permission(request.user, 'void_request'):
                messages.error(request, 'You do not have permission to void requests.')
                return redirect('batch_request_detail', batch_id=batch_id)
            
            # Validate that remarks is provided
            if not remarks:
                messages.error(request, 'Please provide a reason for voiding this request.')
                return redirect('batch_request_detail', batch_id=batch_id)
            
            # Release all reserved quantities from approved items
            approved_items = batch_request.items.filter(status='approved')
            for item in approved_items:
                if item.approved_quantity:
                    quantity_info = item.supply.quantity_info
                    quantity_info.reserved_quantity = max(0, quantity_info.reserved_quantity - item.approved_quantity)
                    quantity_info.save(user=request.user)
            
            # Update ALL items to voided status (not just approved ones)
            batch_request.items.all().update(status='voided')
            
            # Change batch status to voided
            batch_request.status = 'voided'
            batch_request.remarks = remarks
            batch_request.save()
            
            # Create notification for user
            Notification.objects.create(
                user=batch_request.user,
                message=f"Your batch supply request #{batch_request.id} has been voided by admin.",
                remarks=remarks
            )
            
            # Send email notification to the requester
            if batch_request.user.email:
                from django.core.mail import send_mail
                from django.conf import settings
                
                subject = f'Supply Request #{batch_request.id} - Voided'
                message = f"""Dear {batch_request.user.get_full_name() or batch_request.user.username},

Your supply request (Batch #{batch_request.id}) has been voided by the administrator.

Reason for voiding:
{remarks}

All reserved quantities have been released back to the inventory.

If you have any questions, please contact the administrator.

Best regards,
Resource Hive Management System
"""
                try:
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        [batch_request.user.email],
                        fail_silently=False,
                    )
                except Exception as e:
                    # Log the error but don't fail the request
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f"Failed to send void notification email: {str(e)}")
            
            # Log activity
            ActivityLog.log_activity(
                user=request.user,
                action='update',
                model_name='SupplyRequestBatch',
                object_repr=f"Batch Request #{batch_request.id}",
                description=f"Voided for_claiming request, released reserved quantities. Reason: {remarks[:100]}"
            )
            
            messages.success(request, 'Request voided successfully. Reserved quantities have been released and the requester has been notified.')
            return redirect('batch_request_detail', batch_id=batch_id)
        
        if item_id:
            item = get_object_or_404(SupplyRequestItem, id=item_id, batch_request=batch_request)
            
            if action == 'approve':
                # Get approved quantity from form (defaults to requested quantity)
                approved_quantity = request.POST.get('approved_quantity', item.quantity)
                try:
                    approved_quantity = int(approved_quantity)
                    if approved_quantity <= 0 or approved_quantity > item.quantity:
                        approved_quantity = item.quantity
                except (ValueError, TypeError):
                    approved_quantity = item.quantity
                
                # CHECK AVAILABLE STOCK BEFORE APPROVING
                quantity_info = item.supply.quantity_info
                available_qty = quantity_info.available_quantity
                
                if approved_quantity > available_qty:
                    messages.error(request, f'Cannot approve {approved_quantity} units of {item.supply.supply_name}. Only {available_qty} units available (Current: {quantity_info.current_quantity}, Reserved: {quantity_info.reserved_quantity}).')
                    return redirect('batch_request_detail', batch_id=batch_id)
                
                item.status = 'approved'
                item.approved_quantity = approved_quantity
                if remarks:
                    item.remarks = remarks
                item.save()
                
                # Reserve the approved quantity to prevent overbooking
                quantity_info.reserved_quantity += approved_quantity
                quantity_info.save(user=request.user)
                
                # Create notification
                Notification.objects.create(
                    user=batch_request.user,
                    message=f"Item '{item.supply.supply_name}' in your batch request #{batch_request.id} has been approved for {approved_quantity} units.",
                    remarks=remarks
                )
                
                messages.success(request, f'Item "{item.supply.supply_name}" approved successfully.')
                
            elif action == 'reject':
                # If the item was previously approved, release the reserved quantity
                if item.status == 'approved' and item.approved_quantity:
                    quantity_info = item.supply.quantity_info
                    quantity_info.reserved_quantity = max(0, quantity_info.reserved_quantity - item.approved_quantity)
                    quantity_info.save(user=request.user)
                
                item.status = 'rejected'
                if remarks:
                    item.remarks = remarks
                item.save()
                
                # Create notification
                Notification.objects.create(
                    user=batch_request.user,
                    message=f"Item '{item.supply.supply_name}' in your batch request #{batch_request.id} has been rejected.",
                    remarks=remarks
                )
                
                messages.success(request, f'Item "{item.supply.supply_name}" rejected successfully.')
            
            # Update batch status based on individual item statuses
            total_items = batch_request.items.count()
            approved_items = batch_request.items.filter(status='approved').count()
            rejected_items = batch_request.items.filter(status='rejected').count()
            pending_items = batch_request.items.filter(status='pending').count()
            processed_items = approved_items + rejected_items
            
            if processed_items == total_items:
                # All items have been processed (approved or rejected)
                if approved_items > 0:
                    # At least some items were approved
                    batch_request.status = 'for_claiming'
                    if approved_items == total_items:
                        batch_request.approved_date = timezone.now()
                else:
                    # All items were rejected
                    batch_request.status = 'rejected'
            elif approved_items > 0:
                # Some approved, some still pending
                batch_request.status = 'partially_approved'
            else:
                # No items approved yet (either all pending or all rejected)
                if rejected_items > 0 and pending_items > 0:
                    # Some rejected, some pending
                    batch_request.status = 'pending'
                else:
                    # All pending or all rejected
                    batch_request.status = 'pending' if pending_items > 0 else 'rejected'
            
            batch_request.save()
            
            # Send batch completion email ONLY if all items are now processed
            if processed_items == total_items:
                from .utils import send_batch_request_completion_email
                approved_items_qs = batch_request.items.filter(status='approved')
                rejected_items_qs = batch_request.items.filter(status='rejected')
                send_batch_request_completion_email(batch_request, approved_items_qs, rejected_items_qs)
        
        return redirect('batch_request_detail', batch_id=batch_id)
    
    from .permissions import has_admin_permission
    context = {
        'batch_request': batch_request,
        'items': batch_request.items.all().order_by('supply__supply_name'),
        'can_void_request': has_admin_permission(request.user, 'void_request'),
    }
    
    return render(request, 'app/batch_request_detail.html', context)


# Claiming Workflow Views
@permission_required('app.view_admin_module')
@login_required
@require_POST
def claim_batch_items(request, batch_id):
    """
    Handle claiming of all approved items in a batch request.
    This will deduct stock and mark items as completed.
    """
    batch_request = get_object_or_404(SupplyRequestBatch, id=batch_id)
    
    # Only allow claiming if batch is in for_claiming status
    if batch_request.status != 'for_claiming':
        messages.error(request, 'This request is not available for claiming.')
        return redirect_with_tab(request, 'user_supply_requests')
    
    # Get all approved items that haven't been claimed yet
    approved_items = batch_request.items.filter(status='approved', claimed_date__isnull=True)
    
    if not approved_items.exists():
        messages.error(request, 'No approved items available for claiming.')
        return redirect_with_tab(request, 'user_supply_requests')
    
    # Check stock availability for all items before processing
    insufficient_stock_items = []
    for item in approved_items:
        available_quantity = 0
        if hasattr(item.supply, 'quantity_info') and item.supply.quantity_info:
            available_quantity = item.supply.quantity_info.current_quantity or 0
        
        # Handle case where approved_quantity might be None (fallback to requested quantity)
        approved_qty = item.approved_quantity or item.quantity or 0
        
        if available_quantity < approved_qty:
            insufficient_stock_items.append(f"{item.supply.supply_name} (available: {available_quantity}, needed: {approved_qty})")
    
    if insufficient_stock_items:
        messages.error(request, f"Insufficient stock for items: {', '.join(insufficient_stock_items)}")
        return redirect_with_tab(request, 'user_supply_requests')
    
    # Process each approved item
    claimed_items = []
    for item in approved_items:
        # Handle case where approved_quantity might be None (fallback to requested quantity)
        approved_qty = item.approved_quantity or item.quantity or 0
        
        # Deduct from stock
        if hasattr(item.supply, 'quantity_info') and item.supply.quantity_info:
            old_qty = item.supply.quantity_info.current_quantity or 0
            new_qty = max(0, old_qty - approved_qty)
            item.supply.quantity_info.current_quantity = new_qty
            
            # Also release the reserved quantity
            item.supply.quantity_info.reserved_quantity = max(0, item.supply.quantity_info.reserved_quantity - approved_qty)
            
            # Create SupplyHistory entry for quantity deduction
            requester_name = batch_request.user.get_full_name() if batch_request.user.get_full_name() else batch_request.user.username
            SupplyHistory.objects.create(
                supply=item.supply,
                user=request.user,  # The admin/staff who processed the claim
                action='quantity_update',
                field_name='current_quantity',
                old_value=str(old_qty),
                new_value=str(new_qty),
                remarks=f"Supply request by {requester_name} (Batch #{batch_id})"
            )
            
            # Save with skip_history to avoid duplicate entry
            item.supply.quantity_info.save(skip_history=True)
        
        # Update the approved_quantity if it was None
        if item.approved_quantity is None:
            item.approved_quantity = item.quantity
        
        # Mark item as completed and claimed
        item.status = 'completed'
        item.claimed_date = timezone.now()
        item.save()
        
        claimed_items.append(item)
        
        # Log activity
        ActivityLog.log_activity(
            user=request.user,
            action='claim',
            model_name='SupplyRequestItem',
            object_repr=f"Batch #{batch_id} - {item.supply.supply_name}",
            description=f"Claimed {item.approved_quantity or item.quantity} units of {item.supply.supply_name} from batch request #{batch_id}"
        )
    
    # Update batch status and dates
    batch_request.status = 'completed'
    batch_request.claimed_date = timezone.now()
    batch_request.completed_date = timezone.now()
    batch_request.claimed_by = request.user
    batch_request.save()
    
    # Create notification for the requester
    Notification.objects.create(
        user=batch_request.user,
        message=f"Your supply request #{batch_request.id} has been completed and items have been claimed.",
        remarks=f"Total items claimed: {len(claimed_items)}"
    )
    
    # Log batch completion
    ActivityLog.log_activity(
        user=request.user,
        action='complete',
        model_name='SupplyRequestBatch',
        object_repr=f"Batch #{batch_id}",
        description=f"Completed batch request #{batch_id} with {len(claimed_items)} items claimed"
    )
    
    return redirect_with_tab(request, 'user_supply_requests')


@permission_required('app.view_admin_module') 
@login_required
@require_POST
def claim_individual_item(request, batch_id, item_id):
    """
    Handle claiming of an individual approved item.
    This will deduct stock and mark the specific item as completed.
    """
    batch_request = get_object_or_404(SupplyRequestBatch, id=batch_id)
    item = get_object_or_404(SupplyRequestItem, id=item_id, batch_request=batch_request)
    
    # Only allow claiming if item is approved and not yet claimed
    if item.status != 'approved' or item.claimed_date is not None:
        messages.error(request, 'This item is not available for claiming.')
        return redirect('batch_request_detail', batch_id=batch_id)
    
    # Check stock availability
    available_quantity = 0
    if hasattr(item.supply, 'quantity_info') and item.supply.quantity_info:
        available_quantity = item.supply.quantity_info.current_quantity or 0
    
    # Handle case where approved_quantity might be None (fallback to requested quantity)
    approved_qty = item.approved_quantity or item.quantity or 0
    
    if available_quantity < approved_qty:
        messages.error(request, f"Insufficient stock for {item.supply.supply_name}. Available: {available_quantity}, needed: {approved_qty}")
        return redirect('batch_request_detail', batch_id=batch_id)
    
    # Deduct from stock
    if hasattr(item.supply, 'quantity_info') and item.supply.quantity_info:
        old_qty = item.supply.quantity_info.current_quantity or 0
        new_qty = max(0, old_qty - approved_qty)
        item.supply.quantity_info.current_quantity = new_qty
        
        # Also release the reserved quantity
        item.supply.quantity_info.reserved_quantity = max(0, item.supply.quantity_info.reserved_quantity - approved_qty)
        
        # Create SupplyHistory entry for quantity deduction
        requester_name = batch_request.user.get_full_name() if batch_request.user.get_full_name() else batch_request.user.username
        SupplyHistory.objects.create(
            supply=item.supply,
            user=request.user,  # The admin/staff who processed the claim
            action='quantity_update',
            field_name='current_quantity',
            old_value=str(old_qty),
            new_value=str(new_qty),
            remarks=f"Supply request by {requester_name} (Batch #{batch_id})"
        )
        
        # Save with skip_history to avoid duplicate entry
        item.supply.quantity_info.save(skip_history=True)
    
    # Update the approved_quantity if it was None
    if item.approved_quantity is None:
        item.approved_quantity = item.quantity
    
    # Mark item as completed and claimed
    item.status = 'completed'
    item.claimed_date = timezone.now()
    item.save()
    
    # Check if all approved items in the batch are now completed
    remaining_approved_items = batch_request.items.filter(status='approved', claimed_date__isnull=True)
    if not remaining_approved_items.exists():
        # All approved items have been claimed, mark batch as completed
        batch_request.status = 'completed'
        batch_request.completed_date = timezone.now()
        if not batch_request.claimed_date:
            batch_request.claimed_date = timezone.now()
        if not batch_request.claimed_by:
            batch_request.claimed_by = request.user
        batch_request.save()
        
        # Notify user
        Notification.objects.create(
            user=batch_request.user,
            message=f"Your supply request #{batch_request.id} has been completed.",
            remarks="All approved items have been claimed."
        )
    
    # Log activity
    ActivityLog.log_activity(
        user=request.user,
        action='claim',
        model_name='SupplyRequestItem',
        object_repr=f"Batch #{batch_id} - {item.supply.supply_name}",
        description=f"Claimed {item.approved_quantity or item.quantity} units of {item.supply.supply_name} from batch request #{batch_id}"
    )
    
    return redirect('batch_request_detail', batch_id=batch_id)

# Batch Borrow Request Management Views
@permission_required('app.view_admin_module')
@login_required
def borrow_batch_request_detail(request, batch_id):
    """View detailed information about a batch borrow request with individual item actions"""
    batch_request = get_object_or_404(BorrowRequestBatch, id=batch_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        item_id = request.POST.get('item_id')
        remarks = request.POST.get('remarks', '').strip()
        
        # Handle void action for for_claiming requests
        if action == 'void_request' and batch_request.status == 'for_claiming':
            # Check permission
            from .permissions import has_admin_permission
            if not has_admin_permission(request.user, 'void_request'):
                messages.error(request, 'You do not have permission to void requests.')
                return redirect('borrow_batch_request_detail', batch_id=batch_id)
            
            # Validate that remarks is provided
            if not remarks:
                messages.error(request, 'Please provide a reason for voiding this borrow request.')
                return redirect('borrow_batch_request_detail', batch_id=batch_id)
            
            # Release reserved quantities from approved items
            approved_items = batch_request.items.filter(status='approved')
            for item in approved_items:
                if item.approved_quantity:
                    # Release the reserved quantity
                    item.property.reserved_quantity = max(0, item.property.reserved_quantity - item.approved_quantity)
                    item.property.save()
            
            # Update ALL items to voided status (not just approved ones)
            batch_request.items.all().update(status='voided')
            
            # Change batch status to voided
            batch_request.status = 'voided'
            batch_request.remarks = remarks
            batch_request.save()
            
            # Create notification for user
            Notification.objects.create(
                user=batch_request.user,
                message=f"Your batch borrow request #{batch_request.id} has been voided by admin.",
                remarks=remarks
            )
            
            # Send email notification to the requester
            if batch_request.user.email:
                from django.core.mail import send_mail
                from django.conf import settings
                
                subject = f'Borrow Request #{batch_request.id} - Voided'
                message = f"""Dear {batch_request.user.get_full_name() or batch_request.user.username},

Your borrow request (Batch #{batch_request.id}) has been voided by the administrator.

Reason for voiding:
{remarks}

All reserved quantities have been released back to the inventory.

If you have any questions, please contact the administrator.

Best regards,
Resource Hive Management System
"""
                try:
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        [batch_request.user.email],
                        fail_silently=False,
                    )
                except Exception as e:
                    # Log the error but don't fail the request
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f"Failed to send void notification email: {str(e)}")
            
            # Log activity
            ActivityLog.log_activity(
                user=request.user,
                action='update',
                model_name='BorrowRequestBatch',
                object_repr=f"Batch Borrow Request #{batch_request.id}",
                description=f"Voided for_claiming request, released reserved quantities. Reason: {remarks[:100]}"
            )
            
            messages.success(request, 'Borrow request voided successfully. Reserved quantities have been released and the requester has been notified.')
            return redirect('borrow_batch_request_detail', batch_id=batch_id)
        
        if item_id:
            item = get_object_or_404(BorrowRequestItem, id=item_id, batch_request=batch_request)
            
            if action == 'approve':
                # Check availability before approving
                available_quantity = item.property.quantity or 0
                approved_quantity = int(request.POST.get('approved_quantity', item.quantity))
                
                if approved_quantity > available_quantity:
                    messages.error(request, f'Only {available_quantity} units of {item.property.property_name} are available.')
                    return redirect('borrow_batch_request_detail', batch_id=batch_id)
                
                item.status = 'approved'
                item.approved_quantity = approved_quantity
                if remarks:
                    item.remarks = remarks
                item.save()
                
                # Reserve the approved quantity
                item.property.reserved_quantity += approved_quantity
                item.property.save()
                
                # Create notification
                Notification.objects.create(
                    user=batch_request.user,
                    message=f"Item '{item.property.property_name}' in your batch borrow request #{batch_request.id} has been approved for {approved_quantity} units.",
                    remarks=remarks
                )
                
                messages.success(request, f'Item "{item.property.property_name}" approved successfully.')
                
            elif action == 'reject':
                # If item was previously approved, release the reserved quantity
                if item.status == 'approved' and item.approved_quantity:
                    item.property.reserved_quantity = max(0, item.property.reserved_quantity - item.approved_quantity)
                    item.property.save()
                
                item.status = 'rejected'
                if remarks:
                    item.remarks = remarks
                item.save()
                
                # Create notification
                Notification.objects.create(
                    user=batch_request.user,
                    message=f"Item '{item.property.property_name}' in your batch borrow request #{batch_request.id} has been rejected.",
                    remarks=remarks
                )
                
                messages.success(request, f'Item "{item.property.property_name}" rejected successfully.')
            
            # Update batch status based on individual item statuses
            total_items = batch_request.items.count()
            approved_items = batch_request.items.filter(status='approved').count()
            rejected_items = batch_request.items.filter(status='rejected').count()
            pending_items = batch_request.items.filter(status='pending').count()
            processed_items = approved_items + rejected_items
            
            if processed_items == total_items:
                # All items have been processed (approved or rejected)
                if approved_items > 0:
                    # At least some items were approved
                    batch_request.status = 'for_claiming'
                    if approved_items == total_items:
                        batch_request.approved_date = timezone.now()
                else:
                    # All items were rejected
                    batch_request.status = 'rejected'
            elif approved_items > 0:
                # Some approved, some still pending
                batch_request.status = 'partially_approved'
            else:
                # No items approved yet (either all pending or all rejected)
                if rejected_items > 0 and pending_items > 0:
                    # Some rejected, some pending
                    batch_request.status = 'pending'
                else:
                    # All pending or all rejected
                    batch_request.status = 'pending' if pending_items > 0 else 'rejected'
            
            batch_request.save()
            
            # Send borrow batch completion email ONLY if all items are now processed
            if processed_items == total_items:
                # All items have been processed, send completion email
                from .utils import send_borrow_batch_request_completion_email
                approved_items_qs = batch_request.items.filter(status='approved')
                rejected_items_qs = batch_request.items.filter(status='rejected')
                send_borrow_batch_request_completion_email(batch_request, approved_items_qs, rejected_items_qs)
        
        return redirect('borrow_batch_request_detail', batch_id=batch_id)
    
    from .permissions import has_admin_permission
    context = {
        'batch': batch_request,
        'items': batch_request.items.all().order_by('property__property_name'),
        'can_void_request': has_admin_permission(request.user, 'void_request'),
    }
    
    return render(request, 'app/borrow_batch_request_detail.html', context)

@permission_required('app.view_admin_module')
@login_required
@require_POST
def claim_borrow_batch_items(request, batch_id):
    """
    Handle claiming of all approved items in a batch borrow request.
    This will deduct stock and mark items as active.
    Returns JSON response for AJAX calls from list view.
    """
    batch_request = get_object_or_404(BorrowRequestBatch, id=batch_id)
    
    # Check if this is an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    # Only allow claiming if batch is in for_claiming status
    if batch_request.status != 'for_claiming':
        error_msg = 'This request is not available for claiming.'
        if is_ajax:
            return JsonResponse({'success': False, 'message': error_msg})
        messages.error(request, error_msg)
        return redirect_with_tab(request, 'user_borrow_requests')
    
    # Get all approved items that haven't been claimed yet
    approved_items = batch_request.items.filter(status='approved', claimed_date__isnull=True)
    
    if not approved_items.exists():
        error_msg = 'No approved items available for claiming.'
        if is_ajax:
            return JsonResponse({'success': False, 'message': error_msg})
        messages.error(request, error_msg)
        return redirect_with_tab(request, 'user_borrow_requests')
    
    # Check stock availability for all items before processing
    insufficient_stock_items = []
    for item in approved_items:
        available_quantity = item.property.quantity or 0
        approved_qty = item.approved_quantity or item.quantity or 0
        
        if available_quantity < approved_qty:
            insufficient_stock_items.append(f"{item.property.property_name} (need: {approved_qty}, available: {available_quantity})")
    
    if insufficient_stock_items:
        error_msg = f"Insufficient stock for: {', '.join(insufficient_stock_items)}"
        if is_ajax:
            return JsonResponse({'success': False, 'message': error_msg})
        messages.error(request, error_msg)
        return redirect_with_tab(request, 'user_borrow_requests')
    
    # Process all approved items
    claimed_items = []
    
    for item in approved_items:
        approved_qty = item.approved_quantity or item.quantity or 0
        old_quantity = item.property.quantity
        
        # Deduct from property quantity
        item.property.quantity -= approved_qty
        item.property.save()
        
        # Create PropertyHistory entry for deduction with responsible person and reason
        PropertyHistory.objects.create(
            property=item.property,
            user=request.user,
            action='quantity_update',
            field_name='quantity',
            old_value=str(old_quantity),
            new_value=str(item.property.quantity),
            remarks=f"Deducted {approved_qty} units for borrow request. Borrower: {batch_request.user.get_full_name() or batch_request.user.username}. Return date: {batch_request.latest_return_date}. Purpose: {batch_request.purpose[:100]}"
        )
        
        # Update the approved_quantity if it was None
        if item.approved_quantity is None:
            item.approved_quantity = item.quantity
        
        # Mark item as active and claimed
        item.status = 'active'
        item.claimed_date = timezone.now()
        item.save()
        
        claimed_items.append(item)
        
        # Log activity
        ActivityLog.log_activity(
            user=request.user,
            action='claim',
            model_name='BorrowRequestItem',
            object_repr=f"Batch #{batch_id} - {item.property.property_name}",
            description=f"Claimed {approved_qty} units of {item.property.property_name} from batch borrow request #{batch_id}"
        )
    
    # Update batch status and dates
    batch_request.status = 'active'
    batch_request.claimed_date = timezone.now()
    batch_request.claimed_by = request.user
    batch_request.save()
    
    # Create notification for the requester
    Notification.objects.create(
        user=batch_request.user,
        message=f"Your borrow request #{batch_request.id} is now active and items have been claimed.",
        remarks=f"Total items claimed: {len(claimed_items)}. Please return by: {batch_request.latest_return_date}"
    )
    
    # Log batch activation
    ActivityLog.log_activity(
        user=request.user,
        action='activate',
        model_name='BorrowRequestBatch',
        object_repr=f"Batch #{batch_id}",
        description=f"Activated batch borrow request #{batch_id} with {len(claimed_items)} items claimed"
    )
    
    success_msg = f'Successfully claimed {len(claimed_items)} items.'
    
    # Return JSON response for AJAX, or redirect for normal form submission
    if is_ajax:
        return JsonResponse({'success': True, 'message': success_msg})
    
    messages.success(request, success_msg)
    return redirect_with_tab(request, 'user_borrow_requests')

@permission_required('app.view_admin_module')
@login_required
@require_POST
def return_borrow_batch_items(request, batch_id):
    """
    Handle returning of all active items in a batch borrow request.
    This will increase stock and mark items as returned.
    Returns JSON response for AJAX calls from list view.
    """
    batch_request = get_object_or_404(BorrowRequestBatch, id=batch_id)
    
    # Check if this is an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    # Only allow returning if batch is in active or overdue status
    if batch_request.status not in ['active', 'overdue']:
        error_msg = 'This request is not available for returning.'
        if is_ajax:
            return JsonResponse({'success': False, 'message': error_msg})
        messages.error(request, error_msg)
        return redirect_with_tab(request, 'user_borrow_requests')
    
    # Get all active items
    active_items = batch_request.items.filter(status__in=['active', 'overdue'])
    
    if not active_items.exists():
        error_msg = 'No active items available for returning.'
        if is_ajax:
            return JsonResponse({'success': False, 'message': error_msg})
        messages.error(request, error_msg)
        return redirect_with_tab(request, 'user_borrow_requests')
    
    # Process all active items
    returned_items = []
    
    for item in active_items:
        approved_qty = item.approved_quantity or item.quantity or 0
        old_quantity = item.property.quantity
        
        # Add back to property quantity
        item.property.quantity += approved_qty
        item.property.save()
        
        # Create PropertyHistory entry for return with responsible person
        PropertyHistory.objects.create(
            property=item.property,
            user=request.user,
            action='quantity_update',
            field_name='quantity',
            old_value=str(old_quantity),
            new_value=str(item.property.quantity),
            remarks=f"Returned {approved_qty} units from borrow. Borrower: {batch_request.user.get_full_name() or batch_request.user.username}. Processed by: {request.user.get_full_name() or request.user.username}"
        )
        
        # Mark item as returned
        item.status = 'returned'
        item.actual_return_date = timezone.now().date()
        item.save()
        
        returned_items.append(item)
        
        # Log activity
        ActivityLog.log_activity(
            user=request.user,
            action='return',
            model_name='BorrowRequestItem',
            object_repr=f"Batch #{batch_id} - {item.property.property_name}",
            description=f"Returned {approved_qty} units of {item.property.property_name} from batch borrow request #{batch_id}"
        )
    
    # Update batch status and dates
    batch_request.status = 'returned'
    batch_request.returned_date = timezone.now()
    batch_request.completed_date = timezone.now()
    batch_request.save()
    
    # If this borrow batch was created from a reservation, update the reservation status
    source_reservations = batch_request.source_reservation_batch.all()
    if source_reservations.exists():
        for source_reservation in source_reservations:
            # Mark the reservation items as completed
            source_reservation.items.filter(status='active').update(status='completed')
            # Update the reservation batch status to completed if all items are done
            if not source_reservation.items.filter(status__in=['pending', 'approved']).exists():
                source_reservation.status = 'completed'
                source_reservation.save()
    
    # Create notification for the requester
    Notification.objects.create(
        user=batch_request.user,
        message=f"Your borrow request #{batch_request.id} has been marked as returned.",
        remarks="Thank you for returning all items."
    )
    
    # Log batch completion
    ActivityLog.log_activity(
        user=request.user,
        action='return',
        model_name='BorrowRequestBatch',
        object_repr=f"Batch #{batch_id}",
        description=f"Completed batch borrow request #{batch_id} with {len(returned_items)} items returned"
    )
    
    success_msg = f'Successfully returned {len(returned_items)} items.'
    
    # Return JSON response for AJAX, or redirect for normal form submission
    if is_ajax:
        return JsonResponse({'success': True, 'message': success_msg})
    
    messages.success(request, success_msg)
    return redirect_with_tab(request, 'user_borrow_requests')


# Resource Allocation Dashboard View (Unified for all request types)
class ResourceAllocationDashboardView(PermissionRequiredMixin, TemplateView):
    """
    Unified dashboard for viewing all resource allocations across:
    - Borrow Requests
    - Reservations  
    - Supply Requests
    
    Accessible only to admin users.
    """
    permission_required = 'app.view_admin_module'
    template_name = 'app/resource_allocation_dashboard.html'
    
    def get_context_data(self, **kwargs):
        from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
        from app.models import BorrowRequestItem, ReservationItem, SupplyRequestItem, SupplyRequestBatch
        
        context = super().get_context_data(**kwargs)
        
        # Get the current tab from request - default to borrow
        current_tab = self.request.GET.get('tab', 'borrow')
        
        # Get search and filter parameters
        search_query = self.request.GET.get('search', '').strip()
        department_filter = self.request.GET.get('department', '')
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')
        status_filter = self.request.GET.get('status', '')
        
        # Parse dates if provided
        date_from_obj = None
        date_to_obj = None
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            except ValueError:
                pass
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        paginate_by = 15
        
        # Initialize empty allocations - only fetch data for active tab
        borrow_page_obj = None
        reservation_page_obj = None
        supply_page_obj = None
        
        # =========================
        # LAZY LOAD: Only fetch current tab's data
        # =========================
        if current_tab == 'borrow':
            # Exclude items from voided or cancelled batch requests
            borrow_items = BorrowRequestItem.objects.filter(
                Q(status='approved') | Q(status='active')
            ).exclude(
                Q(batch_request__status='voided') | Q(batch_request__status='cancelled')
            ).select_related('batch_request__user', 'batch_request__user__userprofile', 'property').order_by('-batch_request__request_date')
            
            # Apply filters for borrow
            if search_query:
                borrow_items = borrow_items.filter(
                    Q(batch_request__user__first_name__icontains=search_query) |
                    Q(batch_request__user__last_name__icontains=search_query) |
                    Q(batch_request__user__username__icontains=search_query) |
                    Q(property__property_name__icontains=search_query) |
                    Q(property__property_number__icontains=search_query) |
                    Q(batch_request__purpose__icontains=search_query)
                ).distinct()
            
            if department_filter:
                borrow_items = borrow_items.filter(batch_request__user__userprofile__department__name=department_filter)
            
            if date_from_obj:
                borrow_items = borrow_items.filter(batch_request__request_date__date__gte=date_from_obj)
            
            if date_to_obj:
                borrow_items = borrow_items.filter(batch_request__request_date__date__lte=date_to_obj)
            
            if status_filter:
                borrow_items = borrow_items.filter(status=status_filter)
            
            # Format borrow allocations
            borrow_allocations = []
            for item in borrow_items:
                borrow_allocations.append({
                    'type': 'Borrow Request',
                    'type_class': 'borrow',
                    'request_id': item.batch_request.id,
                    'user': item.batch_request.user,
                    'property': item.property,
                    'quantity': item.approved_quantity or item.quantity,
                    'status': item.get_status_display(),
                    'status_value': item.status,
                    'request_date': item.batch_request.request_date,
                    'return_date': item.return_date,
                    'purpose': item.batch_request.purpose,
                    'remarks': item.remarks or '',
                    'detail_url': 'borrow_batch_request_detail',
                })
            
            paginator_borrow = Paginator(borrow_allocations, paginate_by)
            page_borrow = self.request.GET.get('borrow_page', 1)
            try:
                borrow_page_obj = paginator_borrow.page(page_borrow)
            except (PageNotAnInteger, EmptyPage):
                borrow_page_obj = paginator_borrow.page(1)
        
        elif current_tab == 'reservation':
            reservation_items = ReservationItem.objects.filter(
                Q(status='approved') | Q(status='active')
            ).select_related('batch_request__user', 'batch_request__user__userprofile', 'property').order_by('-batch_request__request_date')
            
            # Apply filters for reservations
            if search_query:
                reservation_items = reservation_items.filter(
                    Q(batch_request__user__first_name__icontains=search_query) |
                    Q(batch_request__user__last_name__icontains=search_query) |
                    Q(batch_request__user__username__icontains=search_query) |
                    Q(property__property_name__icontains=search_query) |
                    Q(property__property_number__icontains=search_query) |
                    Q(batch_request__purpose__icontains=search_query)
                ).distinct()
            
            if department_filter:
                reservation_items = reservation_items.filter(batch_request__user__userprofile__department__name=department_filter)
            
            if date_from_obj:
                reservation_items = reservation_items.filter(batch_request__request_date__date__gte=date_from_obj)
            
            if date_to_obj:
                reservation_items = reservation_items.filter(batch_request__request_date__date__lte=date_to_obj)
            
            if status_filter:
                reservation_items = reservation_items.filter(status=status_filter)
            
            # Format reservation allocations
            reservation_allocations = []
            for item in reservation_items:
                reservation_allocations.append({
                    'type': 'Reservation',
                    'type_class': 'reservation',
                    'request_id': item.batch_request.id,
                    'user': item.batch_request.user,
                    'property': item.property,
                    'quantity': item.quantity,
                    'status': item.get_status_display(),
                    'status_value': item.status,
                    'request_date': item.batch_request.request_date,
                    'needed_date': item.needed_date,
                    'return_date': item.return_date,
                    'purpose': item.batch_request.purpose,
                    'remarks': item.remarks or '',
                    'detail_url': 'reservation_batch_detail',
                })
            
            paginator_reservation = Paginator(reservation_allocations, paginate_by)
            page_reservation = self.request.GET.get('reservation_page', 1)
            try:
                reservation_page_obj = paginator_reservation.page(page_reservation)
            except (PageNotAnInteger, EmptyPage):
                reservation_page_obj = paginator_reservation.page(1)
        
        elif current_tab == 'supply':
            # Only show approved supply items (once approved, they're allocated)
            # Exclude items from voided or cancelled batch requests
            supply_items = SupplyRequestItem.objects.filter(
                status='approved'
            ).exclude(
                Q(batch_request__status='voided') | Q(batch_request__status='cancelled')
            ).select_related('batch_request__user', 'batch_request__user__userprofile', 'supply').order_by('-batch_request__request_date')
            
            # Apply filters for supply
            if search_query:
                supply_items = supply_items.filter(
                    Q(batch_request__user__first_name__icontains=search_query) |
                    Q(batch_request__user__last_name__icontains=search_query) |
                    Q(batch_request__user__username__icontains=search_query) |
                    Q(supply__supply_name__icontains=search_query) |
                    Q(supply__barcode__icontains=search_query) |
                    Q(batch_request__purpose__icontains=search_query)
                ).distinct()
            
            if department_filter:
                supply_items = supply_items.filter(batch_request__user__userprofile__department__name=department_filter)
            
            if date_from_obj:
                supply_items = supply_items.filter(batch_request__request_date__date__gte=date_from_obj)
            
            if date_to_obj:
                supply_items = supply_items.filter(batch_request__request_date__date__lte=date_to_obj)
            
            if status_filter and status_filter == 'approved':
                supply_items = supply_items.filter(status='approved')
            
            # Format supply allocations
            supply_allocations = []
            for item in supply_items:
                supply_allocations.append({
                    'type': 'Supply Request',
                    'type_class': 'supply',
                    'request_id': item.batch_request.id,
                    'user': item.batch_request.user,
                    'supply': item.supply,
                    'quantity': item.approved_quantity or item.quantity,
                    'status': item.get_status_display(),
                    'status_value': item.status,
                    'request_date': item.batch_request.request_date,
                    'purpose': item.batch_request.purpose,
                    'remarks': item.remarks or '',
                    'detail_url': 'batch_request_detail',
                })
            
            paginator_supply = Paginator(supply_allocations, paginate_by)
            page_supply = self.request.GET.get('supply_page', 1)
            try:
                supply_page_obj = paginator_supply.page(page_supply)
            except (PageNotAnInteger, EmptyPage):
                supply_page_obj = paginator_supply.page(1)
        
        # Build URL parameters string for pagination
        url_params = ''
        if search_query:
            url_params += f'&search={search_query}'
        if department_filter:
            url_params += f'&department={department_filter}'
        if date_from:
            url_params += f'&date_from={date_from}'
        if date_to:
            url_params += f'&date_to={date_to}'
        if status_filter:
            url_params += f'&status={status_filter}'
        
        context.update({
            'current_tab': current_tab,
            'borrow_allocations': borrow_page_obj,
            'reservation_allocations': reservation_page_obj,
            'supply_allocations': supply_page_obj,
            'search_query': search_query,
            'department_filter': department_filter,
            'date_from': date_from,
            'date_to': date_to,
            'status_filter': status_filter,
            'departments': Department.objects.all(),
            'url_params': url_params,
            'allocation_statuses': ['approved', 'active'],
        })
        
        return context


class UserBorrowRequestBatchListView(LoginRequiredMixin, ListView):
    """View for displaying batch borrow requests - shows all for admin, user's own for regular users"""
    model = BorrowRequestBatch
    template_name = 'app/borrow_batch.html'
    context_object_name = 'batch_requests'
    paginate_by = 10

    def get_queryset(self):
        # Check for overdue batches, near-overdue items, and expired items before getting the queryset
        BorrowRequestBatch.check_overdue_batches()
        BorrowRequestBatch.check_near_overdue_items()
        BorrowRequestBatch.check_expired_batches()
        
        queryset = BorrowRequestBatch.objects \
            .select_related('user', 'user__userprofile', 'claimed_by') \
            .prefetch_related('items__property') \
            .order_by('request_date')
        
        # If user has admin permissions, show all requests
        # If user is regular user, show only their requests
        if self.request.user.has_perm('app.view_admin_module'):
            # Admin users see all requests
            return queryset
        else:
            # Regular users see only their own requests
            return queryset.filter(user=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get the current tab from request
        current_tab = self.request.GET.get('tab', 'pending')
        
        # Get search and filter parameters
        search_query = self.request.GET.get('search', '').strip()
        department_filter = self.request.GET.get('department', '')
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')
        
        # Base queryset with related data
        base_queryset = BorrowRequestBatch.objects.select_related('user', 'user__userprofile', 'user__userprofile__department').prefetch_related('items__property').order_by('request_date')
        
        # Apply user filtering based on permissions
        if self.request.user.has_perm('app.view_admin_module'):
            # Admin users see all requests
            pass  # No additional filtering needed
        else:
            # Regular users see only their own requests
            base_queryset = base_queryset.filter(user=self.request.user)
        
        # Apply search filter
        if search_query:
            base_queryset = base_queryset.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(user__username__icontains=search_query) |
                Q(purpose__icontains=search_query) |
                Q(items__property__property_name__icontains=search_query)
            ).distinct()
        
        # Apply department filter
        if department_filter:
            base_queryset = base_queryset.filter(user__userprofile__department__name=department_filter)
        
        # Apply date filters
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                base_queryset = base_queryset.filter(request_date__date__gte=date_from_obj)
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                base_queryset = base_queryset.filter(request_date__date__lte=date_to_obj)
            except ValueError:
                pass
        
        # Get filtered requests by status
        # Pending and partially_approved statuses
        pending_requests = base_queryset.filter(Q(status='pending') | Q(status='partially_approved'))
        for_claiming_requests = base_queryset.filter(status='for_claiming')
        active_requests = base_queryset.filter(status='active')
        returned_requests = base_queryset.filter(status='returned')
        overdue_requests = base_queryset.filter(status='overdue')
        rejected_requests = base_queryset.filter(status='rejected')
        expired_requests = base_queryset.filter(status='expired')
        voided_requests = base_queryset.filter(status='voided').order_by('-request_date')
        
        # Pagination for each tab
        paginator_pending = Paginator(pending_requests, self.paginate_by)
        paginator_for_claiming = Paginator(for_claiming_requests, self.paginate_by)
        paginator_active = Paginator(active_requests, self.paginate_by)
        paginator_returned = Paginator(returned_requests, self.paginate_by)
        paginator_overdue = Paginator(overdue_requests, self.paginate_by)
        paginator_rejected = Paginator(rejected_requests, self.paginate_by)
        paginator_expired = Paginator(expired_requests, self.paginate_by)
        paginator_voided = Paginator(voided_requests, self.paginate_by)
        
        # Get page numbers for each tab
        page_pending = self.request.GET.get('pending_page', 1)
        page_for_claiming = self.request.GET.get('for_claiming_page', 1)
        page_active = self.request.GET.get('active_page', 1)
        page_returned = self.request.GET.get('returned_page', 1)
        page_overdue = self.request.GET.get('overdue_page', 1)
        page_rejected = self.request.GET.get('rejected_page', 1)
        page_expired = self.request.GET.get('expired_page', 1)
        page_voided = self.request.GET.get('voided_page', 1)
        
        # Build URL parameters string for pagination
        url_params = ''
        if search_query:
            url_params += f'&search={search_query}'
        if department_filter:
            url_params += f'&department={department_filter}'
        if date_from:
            url_params += f'&date_from={date_from}'
        if date_to:
            url_params += f'&date_to={date_to}'
        
        context.update({
            'current_tab': current_tab,
            'pending_requests': paginator_pending.get_page(page_pending),
            'for_claiming_requests': paginator_for_claiming.get_page(page_for_claiming),
            'active_requests': paginator_active.get_page(page_active),
            'returned_requests': paginator_returned.get_page(page_returned),
            'overdue_requests': paginator_overdue.get_page(page_overdue),
            'rejected_requests': paginator_rejected.get_page(page_rejected),
            'expired_requests': paginator_expired.get_page(page_expired),
            'voided_requests': paginator_voided.get_page(page_voided),
            'search_query': search_query,
            'department_filter': department_filter,
            'date_from': date_from,
            'date_to': date_to,
            'departments': Department.objects.all(),
            'url_params': url_params,
        })
        
        return context


# Individual Borrow Item Management Views for Batch Requests
@permission_required('app.view_admin_module')
@login_required
@require_POST
@admin_permission_required('approve_borrow_request')
def approve_borrow_item(request, batch_id, item_id):
    """Approve an individual item in a batch borrow request"""
    batch_request = get_object_or_404(BorrowRequestBatch, id=batch_id)
    item = get_object_or_404(BorrowRequestItem, id=item_id, batch_request=batch_request)
    
    # Get approved quantity from form
    approved_quantity = request.POST.get('approved_quantity', item.quantity)
    try:
        approved_quantity = int(approved_quantity)
        if approved_quantity <= 0 or approved_quantity > item.quantity:
            messages.error(request, f'Invalid approved quantity. Must be between 1 and {item.quantity}.')
            return redirect('borrow_batch_request_detail', batch_id=batch_id)
    except (ValueError, TypeError):
        approved_quantity = item.quantity
    
    # Check if enough property units are available
    # Use available_quantity which accounts for already reserved/approved items
    available_quantity = item.property.available_quantity
    if approved_quantity > available_quantity:
        messages.error(request, f'Only {available_quantity} units of {item.property.property_name} are available (including approved reservations and borrows).')
        return redirect('borrow_batch_request_detail', batch_id=batch_id)
    
    # Mark item as approved
    item.approved = True
    item.status = 'approved'
    item.approved_quantity = approved_quantity
    remarks = request.POST.get('remarks', '')
    if remarks:
        item.remarks = remarks
    item.save()
    
    # Check if all items in the batch are processed
    total_items = batch_request.items.count()
    approved_items = batch_request.items.filter(approved=True).count()
    rejected_items = batch_request.items.filter(approved=False, status='rejected').count()
    processed_items = approved_items + rejected_items
    
    # Update batch status
    if processed_items == total_items:
        # All items have been processed (approved or rejected)
        if approved_items > 0:
            batch_request.status = 'for_claiming'
            batch_request.approved_date = timezone.now()
        else:
            batch_request.status = 'rejected'
    elif approved_items > 0:
        batch_request.status = 'partially_approved'
    
    batch_request.save()
    
    # Send borrow batch completion email ONLY if all items are now processed
    if processed_items == total_items:
        # All items have been processed, send completion email
        from .utils import send_borrow_batch_request_completion_email
        approved_items_qs = batch_request.items.filter(status='approved')
        rejected_items_qs = batch_request.items.filter(status='rejected')
        send_borrow_batch_request_completion_email(batch_request, approved_items_qs, rejected_items_qs)
    
    # Create notification for user
    Notification.objects.create(
        user=batch_request.user,
        message=f"Item '{item.property.property_name}' in your batch borrow request #{batch_request.id} has been approved for {item.approved_quantity} units.",
        remarks=remarks
    )
    
    # Log activity
    ActivityLog.log_activity(
        user=request.user,
        action='approve',
        model_name='BorrowRequestItem',
        object_repr=f"Batch #{batch_id} - {item.property.property_name}",
        description=f"Approved {item.approved_quantity} out of {item.quantity} units of {item.property.property_name} in batch borrow request #{batch_id}"
    )
    
    return redirect('borrow_batch_request_detail', batch_id=batch_id)


@permission_required('app.view_admin_module')
@login_required
@require_POST
@admin_permission_required('approve_borrow_request')
def reject_borrow_item(request, batch_id, item_id):
    """Reject an individual item in a batch borrow request"""
    batch_request = get_object_or_404(BorrowRequestBatch, id=batch_id)
    item = get_object_or_404(BorrowRequestItem, id=item_id, batch_request=batch_request)
    
    # Mark item as not approved and add remarks
    item.approved = False
    item.status = 'rejected'
    remarks = request.POST.get('remarks', '')
    if remarks:
        item.remarks = remarks
    item.save()
    
    # Check if all items in the batch are processed
    total_items = batch_request.items.count()
    approved_items = batch_request.items.filter(approved=True).count()
    rejected_items = batch_request.items.filter(approved=False, status='rejected').count()
    processed_items = approved_items + rejected_items
    
    # Update batch status
    if processed_items == total_items:
        # All items have been processed (approved or rejected)
        if approved_items > 0:
            batch_request.status = 'for_claiming'
            batch_request.approved_date = timezone.now()
        else:
            batch_request.status = 'rejected'
    elif approved_items > 0:
        batch_request.status = 'partially_approved'
    
    batch_request.save()
    
    # Send borrow batch completion email ONLY if all items are now processed
    if processed_items == total_items:
        # All items have been processed, send completion email
        from .utils import send_borrow_batch_request_completion_email
        approved_items_qs = batch_request.items.filter(status='approved')
        rejected_items_qs = batch_request.items.filter(status='rejected')
        send_borrow_batch_request_completion_email(batch_request, approved_items_qs, rejected_items_qs)
    
    # Create notification for user
    Notification.objects.create(
        user=batch_request.user,
        message=f"Item '{item.property.property_name}' in your batch borrow request #{batch_request.id} has been rejected.",
        remarks=remarks
    )
    
    # Log activity
    ActivityLog.log_activity(
        user=request.user,
        action='reject',
        model_name='BorrowRequestItem',
        object_repr=f"Batch #{batch_id} - {item.property.property_name}",
        description=f"Rejected {item.quantity} units of {item.property.property_name} in batch borrow request #{batch_id}"
    )
    
    return redirect('borrow_batch_request_detail', batch_id=batch_id)


@permission_required('app.view_admin_module')
@login_required
@require_POST
def claim_individual_borrow_item(request, batch_id, item_id):
    """
    Handle claiming of an individual approved borrow item.
    This will deduct stock and mark the specific item as active.
    """
    batch_request = get_object_or_404(BorrowRequestBatch, id=batch_id)
    item = get_object_or_404(BorrowRequestItem, id=item_id, batch_request=batch_request)
    
    # Only allow claiming if item is approved and not yet claimed
    if item.status != 'approved' or item.claimed_date is not None:
        messages.error(request, 'This item is not available for claiming.')
        return redirect('borrow_batch_request_detail', batch_id=batch_id)
    
    # Check stock availability
    available_quantity = item.property.quantity or 0
    approved_qty = item.approved_quantity or item.quantity or 0
    
    if available_quantity < approved_qty:
        messages.error(request, f"Insufficient stock for {item.property.property_name}. Available: {available_quantity}, needed: {approved_qty}")
        return redirect('borrow_batch_request_detail', batch_id=batch_id)
    
    # Store old quantity for history
    old_quantity = item.property.quantity
    
    # Deduct from property stock
    item.property.quantity -= approved_qty
    item.property.save()
    
    # Create PropertyHistory entry for deduction with responsible person and reason
    PropertyHistory.objects.create(
        property=item.property,
        user=request.user,
        action='quantity_update',
        field_name='quantity',
        old_value=str(old_quantity),
        new_value=str(item.property.quantity),
        remarks=f"Deducted {approved_qty} units for borrow request. Borrower: {batch_request.user.get_full_name() or batch_request.user.username}. Return date: {item.return_date}. Claimed by: {request.user.get_full_name() or request.user.username}"
    )
    
    # Update the approved_quantity if it was None
    if item.approved_quantity is None:
        item.approved_quantity = item.quantity
    
    # Mark item as active and claimed
    item.status = 'active'
    item.claimed_date = timezone.now()
    item.save()
    
    # Check if all approved items in the batch are now active
    remaining_approved_items = batch_request.items.filter(status='approved', claimed_date__isnull=True)
    if not remaining_approved_items.exists():
        # All approved items have been claimed, mark batch as active
        batch_request.status = 'active'
        if not batch_request.claimed_date:
            batch_request.claimed_date = timezone.now()
        if not batch_request.claimed_by:
            batch_request.claimed_by = request.user
        batch_request.save()
        
        # Notify user
        Notification.objects.create(
            user=batch_request.user,
            message=f"Your borrow request #{batch_request.id} is now active.",
            remarks="All approved items have been claimed. Please return by the specified dates."
        )
    
    # Log activity
    ActivityLog.log_activity(
        user=request.user,
        action='claim',
        model_name='BorrowRequestItem',
        object_repr=f"Batch #{batch_id} - {item.property.property_name}",
        description=f"Claimed {item.approved_quantity or item.quantity} units of {item.property.property_name} from batch borrow request #{batch_id}"
    )
    
    return redirect('borrow_batch_request_detail', batch_id=batch_id)


@permission_required('app.view_admin_module')
@login_required
@require_POST
def return_individual_borrow_item(request, batch_id, item_id):
    """
    Handle returning of an individual active borrow item.
    This will increase stock and mark the specific item as returned.
    """
    batch_request = get_object_or_404(BorrowRequestBatch, id=batch_id)
    item = get_object_or_404(BorrowRequestItem, id=item_id, batch_request=batch_request)
    
    # Only allow returning if item is active or overdue
    if item.status not in ['active', 'overdue']:
        messages.error(request, 'This item is not available for returning.')
        return redirect('borrow_batch_request_detail', batch_id=batch_id)
    
    # Add back to property stock
    approved_qty = item.approved_quantity or item.quantity or 0
    old_quantity = item.property.quantity
    item.property.quantity += approved_qty
    item.property.save()
    
    # Create PropertyHistory entry for return with responsible person
    PropertyHistory.objects.create(
        property=item.property,
        user=request.user,
        action='quantity_update',
        field_name='quantity',
        old_value=str(old_quantity),
        new_value=str(item.property.quantity),
        remarks=f"Returned {approved_qty} units from borrow. Borrower: {batch_request.user.get_full_name() or batch_request.user.username}. Processed by: {request.user.get_full_name() or request.user.username}"
    )
    
    # Mark item as returned
    item.status = 'returned'
    item.actual_return_date = timezone.now().date()
    item.save()
    
    # Check if all active items in the batch are now returned
    remaining_active_items = batch_request.items.filter(status__in=['active', 'overdue'])
    if not remaining_active_items.exists():
        # All items have been returned, mark batch as returned
        batch_request.status = 'returned'
        batch_request.returned_date = timezone.now()
        batch_request.completed_date = timezone.now()
        batch_request.save()
        
        # If this borrow batch was created from a reservation, update the reservation status
        source_reservations = batch_request.source_reservation_batch.all()
        if source_reservations.exists():
            for source_reservation in source_reservations:
                # Mark the reservation items as completed
                source_reservation.items.filter(status='active').update(status='completed')
                # Update the reservation batch status to completed if all items are done
                if not source_reservation.items.filter(status__in=['pending', 'approved']).exists():
                    source_reservation.status = 'completed'
                    source_reservation.save()
        
        # Notify user
        Notification.objects.create(
            user=batch_request.user,
            message=f"Your borrow request #{batch_request.id} has been completed.",
            remarks="All items have been successfully returned."
        )
    
    # Log activity
    ActivityLog.log_activity(
        user=request.user,
        action='return',
        model_name='BorrowRequestItem',
        object_repr=f"Batch #{batch_id} - {item.property.property_name}",
        description=f"Returned {approved_qty} units of {item.property.property_name} from batch borrow request #{batch_id}"
    )
    
    return redirect('borrow_batch_request_detail', batch_id=batch_id)


class AdminProfileView(PermissionRequiredMixin, View):
    permission_required = 'auth.view_user'
    template_name = 'app/admin_profile.html'
    
    def get(self, request):
        from .forms import AdminProfileUpdateForm
        form = AdminProfileUpdateForm(user=request.user)
        
        # Get user profile or create one if it doesn't exist
        user_profile, created = UserProfile.objects.get_or_create(user=request.user)
        
        return render(request, self.template_name, {
            'form': form,
            'user_profile': user_profile
        })
    
    def post(self, request):
        from .forms import AdminProfileUpdateForm
        form = AdminProfileUpdateForm(user=request.user, data=request.POST)
        
        # Get user profile or create one if it doesn't exist
        user_profile, created = UserProfile.objects.get_or_create(user=request.user)
        
        if form.is_valid():
            # Update user fields
            request.user.username = form.cleaned_data['username']
            request.user.first_name = form.cleaned_data['first_name']
            request.user.last_name = form.cleaned_data['last_name']
            request.user.email = form.cleaned_data['email']
            
            # Update phone and designation in user profile
            user_profile.phone = form.cleaned_data['phone']
            user_profile.designation = form.cleaned_data['designation']
            user_profile.save()
            
            request.user.save()
            
            # Log profile update activity
            ActivityLog.log_activity(
                user=request.user,
                action='update',
                model_name='UserProfile',
                object_repr=f"{request.user.get_full_name() or request.user.username}",
                description="Admin updated their profile information"
            )
            
            messages.success(request, 'Profile updated successfully!')
            return redirect('admin_profile')
        
        return render(request, self.template_name, {
            'form': form,
            'user_profile': user_profile
        })

@require_http_methods(["GET"])
def get_latest_batch_request(request):
    """Get the latest batch request ID for a user"""
    from django.contrib.auth.models import User
    from django.http import JsonResponse
    
    username = request.GET.get('username')
    if not username:
        return JsonResponse({'error': 'Username required'}, status=400)
    
    try:
        user = User.objects.get(username=username)
        latest_batch = BorrowRequestBatch.objects.filter(user=user).order_by('-request_date').first()
        
        if latest_batch:
            return JsonResponse({'batch_id': latest_batch.id, 'user': username})
        else:
            return JsonResponse({'error': 'No batch requests found'}, status=404)
    except User.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_latest_supply_request(request):
    """Get the latest supply request ID for a user"""
    from django.contrib.auth.models import User
    from django.http import JsonResponse
    
    username = request.GET.get('username')
    if not username:
        return JsonResponse({'error': 'Username required'}, status=400)
    
    try:
        user = User.objects.get(username=username)
        # Try batch supply requests first
        latest_batch = SupplyRequestBatch.objects.filter(user=user).order_by('-request_date').first()
        
        if latest_batch:
            return JsonResponse({'batch_id': latest_batch.id, 'user': username, 'type': 'batch'})
        else:
            # Try individual supply requests
            latest_supply = SupplyRequest.objects.filter(user=user).order_by('-request_date').first()
            if latest_supply:
                return JsonResponse({'supply_id': latest_supply.id, 'user': username, 'type': 'individual'})
            else:
                return JsonResponse({'error': 'No supply requests found'}, status=404)
    except User.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_latest_damage_report(request):
    """Get the latest damage report ID for a user"""
    from django.contrib.auth.models import User
    from django.http import JsonResponse
    
    username = request.GET.get('username')
    if not username:
        return JsonResponse({'error': 'Username required'}, status=400)
    
    try:
        user = User.objects.get(username=username)
        latest_report = DamageReport.objects.filter(user=user).order_by('-date_reported').first()
        
        if latest_report:
            return JsonResponse({'report_id': latest_report.id, 'user': username})
        else:
            return JsonResponse({'error': 'No damage reports found'}, status=404)
    except User.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_latest_reservation(request):
    """Get the latest reservation ID for a user"""
    from django.contrib.auth.models import User
    from django.http import JsonResponse
    
    username = request.GET.get('username')
    if not username:
        return JsonResponse({'error': 'Username required'}, status=400)
    
    try:
        user = User.objects.get(username=username)
        latest_reservation = Reservation.objects.filter(user=user).order_by('-reservation_date').first()
        
        if latest_reservation:
            return JsonResponse({'reservation_id': latest_reservation.id, 'user': username})
        else:
            return JsonResponse({'error': 'No reservations found'}, status=404)
    except User.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Requisition and Issue Slip PDF Generation Views
@login_required
def download_requisition_slip(request, batch_id):
    """
    Download the requisition and issue slip PDF for a supply request batch.
    Available for both admin users and the requester.
    """
    batch_request = get_object_or_404(SupplyRequestBatch, id=batch_id)
    
    # Check permissions: admin users can download any slip, regular users can only download their own
    if not (request.user.userprofile.role == 'ADMIN' or batch_request.user == request.user):
        messages.error(request, 'You do not have permission to access this requisition slip.')
        return redirect('user_supply_requests')
    
    # Only generate slip for approved, partially approved, for_claiming, or completed requests
    if batch_request.status not in ['approved', 'partially_approved', 'for_claiming', 'completed']:
        messages.error(request, 'Requisition slip is only available for approved requests.')
        return redirect('batch_request_detail' if request.user.userprofile.role == 'ADMIN' else 'user_supply_requests', batch_id=batch_id)
    
    try:
        from .pdf_utils import download_requisition_slip
        
        # Log the download activity
        ActivityLog.log_activity(
            user=request.user,
            action='view',
            model_name='SupplyRequestBatch',
            object_repr=f"Requisition Slip #{batch_id}",
            description=f"Downloaded requisition slip for batch request #{batch_id}"
        )
        
        return download_requisition_slip(batch_request)
        
    except Exception as e:
        messages.error(request, f'Error generating requisition slip: {str(e)}')
        return redirect('batch_request_detail' if request.user.userprofile.role == 'ADMIN' else 'user_supply_requests', batch_id=batch_id)


@login_required
def view_requisition_slip(request, batch_id):
    """
    View the requisition and issue slip PDF in browser for a supply request batch.
    Available for both admin users and the requester.
    """
    batch_request = get_object_or_404(SupplyRequestBatch, id=batch_id)
    
    # Check permissions: admin users can view any slip, regular users can only view their own
    if not (request.user.userprofile.role == 'ADMIN' or batch_request.user == request.user):
        messages.error(request, 'You do not have permission to access this requisition slip.')
        return redirect('user_supply_requests')
    
    # Only generate slip for approved, partially approved, for_claiming, or completed requests
    if batch_request.status not in ['approved', 'partially_approved', 'for_claiming', 'completed']:
        messages.error(request, 'Requisition slip is only available for approved requests.')
        return redirect('batch_request_detail' if request.user.userprofile.role == 'ADMIN' else 'user_supply_requests', batch_id=batch_id)
    
    try:
        from .pdf_utils import view_requisition_slip
        
        # Log the view activity
        ActivityLog.log_activity(
            user=request.user,
            action='view',
            model_name='SupplyRequestBatch',
            object_repr=f"Requisition Slip #{batch_id}",
            description=f"Viewed requisition slip for batch request #{batch_id}"
        )
        
        return view_requisition_slip(batch_request)
        
    except Exception as e:
        messages.error(request, f'Error generating requisition slip: {str(e)}')
        return redirect('batch_request_detail' if request.user.userprofile.role == 'ADMIN' else 'user_supply_requests', batch_id=batch_id)


# Barcode Inventory Management Views
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin

class ManagePropertyInventoryView(PermissionRequiredMixin, TemplateView):
    template_name = 'app/manage_property_inventory.html'
    permission_required = 'app.view_admin_module'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class ManageSupplyInventoryView(PermissionRequiredMixin, TemplateView):
    template_name = 'app/manage_supply_inventory.html'
    permission_required = 'app.view_admin_module'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


# View to serve damage report images from database
from django.http import HttpResponse

def damage_report_image(request, report_id):
    """
    Serve damage report image stored as binary data in PostgreSQL.
    Returns the image as HTTP response with proper MIME type.
    """
    report = get_object_or_404(DamageReport, pk=report_id)
    
    # Check if report has image and it hasn't been deleted
    if report.image_data and not report.deleted_at:
        response = HttpResponse(report.image_data, content_type=report.image_type)
        response['Content-Disposition'] = f'inline; filename="{report.image_name or f"damage_report_{report_id}.jpg"}"'
        return response
    else:
        # Return 404 if no image or image was deleted
        from django.http import Http404
        raise Http404("Image not found or has been deleted")


def lost_item_image(request, report_id):
    """
    Serve lost item report image stored as binary data in PostgreSQL.
    Returns the image as HTTP response with proper MIME type.
    """
    report = get_object_or_404(LostItem, pk=report_id)
    
    # Check if report has image and it hasn't been deleted
    if report.image_data and not report.deleted_at:
        response = HttpResponse(report.image_data, content_type=report.image_type)
        response['Content-Disposition'] = f'inline; filename="{report.image_name or f"lost_item_{report_id}.jpg"}"'
        return response
    else:
        # Return 404 if no image or image was deleted
        from django.http import Http404
        raise Http404("Image not found or has been deleted")


@login_required
@user_passes_test(lambda u: u.userprofile.role == 'ADMIN', login_url='user_dashboard')
@require_http_methods(["POST"])
def delete_damage_report_image(request, report_id):
    """
    Delete a damage report image (admin only).
    Marks the image as deleted with audit trail.
    """
    from django.utils import timezone
    from django.contrib import messages
    from django.http import JsonResponse
    
    report = get_object_or_404(DamageReport, pk=report_id)
    
    # Check if report has an image to delete
    if not report.has_image:
        return JsonResponse({
            'success': False,
            'message': 'No image to delete or image already deleted.'
        }, status=400)
    
    # Soft delete: Mark as deleted but keep the data for audit
    report.deleted_at = timezone.now()
    report.deleted_by = request.user
    report.save()
    
    # Log activity
    ActivityLog.log_activity(
        user=request.user,
        action='delete',
        model_name='DamageReport Image',
        object_repr=f"Report #{report.id} - {report.item.property_name}",
        description=f"Deleted image '{report.image_name}' from damage report #{report.id}"
    )
    
    # Notify the user who submitted the report
    Notification.objects.create(
        user=report.user,
        message=f"The image attached to your damage report #{report.id} for {report.item.property_name} has been removed by an administrator.",
        remarks=f"Image removed by {request.user.username} on {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    )
    
    return JsonResponse({
        'success': True,
        'message': f'Image deleted successfully. Report #{report.id} remains in the system.'
    })


@login_required
@user_passes_test(lambda u: u.userprofile.role == 'ADMIN', login_url='user_dashboard')
@require_http_methods(["POST"])
def bulk_delete_damage_report_images(request):
    """
    Bulk delete damage report images (admin only).
    Accepts a list of report IDs via POST.
    """
    from django.utils import timezone
    from django.http import JsonResponse
    import json
    
    try:
        data = json.loads(request.body)
        report_ids = data.get('report_ids', [])
        
        if not report_ids:
            return JsonResponse({
                'success': False,
                'message': 'No reports selected.'
            }, status=400)
        
        # Get reports that have images
        reports = DamageReport.objects.filter(
            id__in=report_ids,
            image_data__isnull=False,
            deleted_at__isnull=True
        )
        
        deleted_count = 0
        for report in reports:
            report.deleted_at = timezone.now()
            report.deleted_by = request.user
            report.save()
            
            # Log activity
            ActivityLog.log_activity(
                user=request.user,
                action='delete',
                model_name='DamageReport Image',
                object_repr=f"Report #{report.id} - {report.item.property_name}",
                description=f"Bulk deleted image '{report.image_name}' from damage report #{report.id}"
            )
            
            # Notify the user
            Notification.objects.create(
                user=report.user,
                message=f"The image attached to your damage report #{report.id} for {report.item.property_name} has been removed by an administrator.",
                remarks=f"Bulk deletion by {request.user.username} on {timezone.now().strftime('%Y-%m-%d %H:%M')}"
            )
            
            deleted_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'{deleted_count} image(s) deleted successfully.',
            'deleted_count': deleted_count
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid request data.'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


@login_required
@permission_required('app.view_admin_module', raise_exception=True)
def generate_completed_supply_requests_pdf(request):
    """Generate PDF report for completed supply requests with filters"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from django.db.models import Q
        from io import BytesIO
        from datetime import datetime
    except ImportError as e:
        return HttpResponse(f"Error importing required libraries: {str(e)}", status=500)
    
    # Get filter parameters
    user_filter = request.GET.get('user', '')
    department_filter = request.GET.get('department', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Build queryset for completed requests
    completed_requests = SupplyRequestBatch.objects.filter(
        status='completed'
    ).select_related(
        'user', 
        'user__userprofile', 
        'user__userprofile__department'
    ).prefetch_related(
        'items__supply'
    ).order_by('-completed_date')
    
    # Apply filters
    if user_filter:
        completed_requests = completed_requests.filter(user__id=user_filter)
    
    if department_filter:
        completed_requests = completed_requests.filter(user__userprofile__department__id=department_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            completed_requests = completed_requests.filter(completed_date__date__gte=date_from_obj.date())
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            completed_requests = completed_requests.filter(completed_date__date__lte=date_to_obj.date())
        except ValueError:
            pass
    
    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    filename = f'Completed_Supply_Requests_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create the PDF object with LANDSCAPE orientation
    from reportlab.lib.pagesizes import landscape
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        rightMargin=0.5*inch, 
        leftMargin=0.5*inch,
        topMargin=0.5*inch, 
        bottomMargin=0.5*inch
    )
    
    # Container for the 'Flowable' objects
    story = []
    
    # Define styles
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#152d64'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#152d64'),
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_LEFT
    )
    
    # Title
    story.append(Paragraph("COMPLETED SUPPLY REQUESTS REPORT", title_style))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", normal_style))
    story.append(Spacer(1, 20))
    
    # Filter information
    filter_info = []
    if user_filter:
        try:
            from django.contrib.auth.models import User
            user = User.objects.get(id=user_filter)
            user_name = f"{user.first_name} {user.last_name}".strip() or user.username
            filter_info.append(f"User: {user_name}")
        except:
            pass
    
    if department_filter:
        try:
            from .models import Department
            dept = Department.objects.get(id=department_filter)
            filter_info.append(f"Department: {dept.name}")
        except:
            pass
    
    if date_from:
        filter_info.append(f"From: {date_from}")
    
    if date_to:
        filter_info.append(f"To: {date_to}")
    
    if filter_info:
        story.append(Paragraph("<b>Filters Applied:</b> " + " | ".join(filter_info), normal_style))
        story.append(Spacer(1, 15))
    
    total_count = completed_requests.count()
    story.append(Paragraph(f"<b>Total Requests:</b> {total_count}", normal_style))
    story.append(Spacer(1, 15))
    
    # Define text wrapping style for table cells
    wrap_style = ParagraphStyle(
        'WrapText',
        parent=styles['Normal'],
        fontSize=7,
        alignment=TA_LEFT,
        spaceAfter=0
    )
    
    # Check if there are no requests
    if total_count == 0:
        story.append(Paragraph("<b>No completed supply requests found with the selected filters.</b>", normal_style))
        story.append(Spacer(1, 20))
        story.append(Paragraph("Try adjusting your filter criteria or check back later.", normal_style))
    else:
        # Create a single consolidated table with all items from all completed requests
        story.append(Paragraph("<b>All Completed Supply Requests</b>", heading_style))
        story.append(Spacer(1, 10))
        
        # Prepare table data with all items
        table_data = [
            ["Req#", "Requester", "Dept", "Purpose", "Item Name", "Req\nQty", "App\nQty", "Request\nDate", "Completed\nDate", "Released By", "Remarks"]
        ]
        
        for batch_request in completed_requests:
            user_name = f"{batch_request.user.first_name} {batch_request.user.last_name}".strip() or batch_request.user.username
            dept_name = batch_request.user.userprofile.department.name if batch_request.user.userprofile.department else "N/A"
            
            purpose_display = (batch_request.purpose[:25] + "...") if batch_request.purpose and len(batch_request.purpose) > 25 else (batch_request.purpose or "N/A")
            
            claimed_by_name = "N/A"
            if batch_request.claimed_by:
                claimed_by_name = f"{batch_request.claimed_by.first_name} {batch_request.claimed_by.last_name}".strip() or batch_request.claimed_by.username
            
            request_date_str = batch_request.request_date.strftime('%m/%d/%Y')
            completed_date_str = batch_request.completed_date.strftime('%m/%d/%Y') if batch_request.completed_date else "N/A"
            
            # Add a row for each item in the request
            items = batch_request.items.all().order_by('supply__supply_name')
            if items.count() == 0:
                # If no items, still show the request
                table_data.append([
                    f"#{batch_request.id:03d}",
                    Paragraph(user_name, wrap_style),  # Wrap in Paragraph for text wrapping
                    dept_name,
                    purpose_display,
                    "No items",
                    "-",
                    "-",
                    request_date_str,
                    completed_date_str,
                    Paragraph(claimed_by_name, wrap_style),  # Wrap in Paragraph for text wrapping
                    "-"
                ])
            else:
                for idx, item in enumerate(items):
                    remarks_display = "-"
                    if item.remarks:
                        remarks_display = item.remarks[:30] + "..." if len(item.remarks) > 30 else item.remarks
                    
                    # For first item, show all request details
                    if idx == 0:
                        table_data.append([
                            f"#{batch_request.id:03d}",
                            Paragraph(user_name, wrap_style),  # Wrap in Paragraph for text wrapping
                            dept_name,
                            purpose_display,
                            item.supply.supply_name,
                            str(item.quantity),
                            str(item.approved_quantity) if item.approved_quantity else "-",
                            request_date_str,
                            completed_date_str,
                            Paragraph(claimed_by_name, wrap_style),  # Wrap in Paragraph for text wrapping
                            remarks_display
                        ])
                    else:
                        # For subsequent items, only show item details
                        table_data.append([
                            "",  # Empty Req#
                            "",  # Empty Requester
                            "",  # Empty Department
                            "",  # Empty Purpose
                            item.supply.supply_name,
                            str(item.quantity),
                            str(item.approved_quantity) if item.approved_quantity else "-",
                            "",  # Empty Request Date
                            "",  # Empty Completed Date
                            "",  # Empty Released By
                            remarks_display
                        ])
        
        # Create table with appropriate column widths for landscape
        main_table = Table(table_data, colWidths=[
            0.45*inch,  # Req#
            1.5*inch,   # Requester (increased width for long names)
            0.85*inch,  # Department
            1.1*inch,   # Purpose
            1.9*inch,   # Item Name
            0.55*inch,  # Requested Qty
            0.55*inch,  # Approved Qty
            0.7*inch,   # Request Date
            0.75*inch,  # Completed Date
            1.0*inch,   # Released By
            1.2*inch    # Remarks (slightly reduced to compensate)
        ])
        
        main_table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#152d64')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7.5),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            
            # Data rows styling
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Req# center
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Requester left
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),    # Department left
            ('ALIGN', (3, 1), (3, -1), 'LEFT'),    # Purpose left
            ('ALIGN', (4, 1), (4, -1), 'LEFT'),    # Item name left
            ('ALIGN', (5, 1), (6, -1), 'CENTER'),  # Quantities center
            ('ALIGN', (7, 1), (8, -1), 'CENTER'),  # Dates center
            ('ALIGN', (9, 1), (9, -1), 'LEFT'),    # Released By left
            ('ALIGN', (10, 1), (10, -1), 'LEFT'),  # Remarks left
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Padding
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#152d64')),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        
        story.append(main_table)
    
    # Build PDF
    try:
        doc.build(story)
        
        # Get the value of the BytesIO buffer and write it to the response
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return HttpResponse(f"Error generating PDF: {str(e)}\n\nDetails:\n{error_details}", status=500, content_type='text/plain')


def generate_items_tally_report_pdf(request):
    """Generate PDF report for items tallied by supply name (aggregated across all requests)"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from django.db.models import Q, Sum, Count
        from io import BytesIO
        from datetime import datetime
    except ImportError as e:
        return HttpResponse(f"Error importing required libraries: {str(e)}", status=500)
    
    # Get filter parameters
    user_filter = request.GET.get('user', '')
    department_filter = request.GET.get('department', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status_filter = request.GET.get('status', '')
    
    # Build queryset for completed requests to aggregate items
    completed_requests = SupplyRequestBatch.objects.filter(
        status='completed'
    ).select_related(
        'user', 
        'user__userprofile', 
        'user__userprofile__department'
    ).prefetch_related(
        'items__supply'
    )
    
    # Apply filters
    if user_filter:
        completed_requests = completed_requests.filter(user__id=user_filter)
    
    if department_filter:
        completed_requests = completed_requests.filter(user__userprofile__department__id=department_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            completed_requests = completed_requests.filter(completed_date__date__gte=date_from_obj.date())
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            completed_requests = completed_requests.filter(completed_date__date__lte=date_to_obj.date())
        except ValueError:
            pass
    
    if status_filter:
        completed_requests = completed_requests.filter(items__status=status_filter)
    
    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    filename = f'Items_Tally_Report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create the PDF object with LANDSCAPE orientation
    from reportlab.lib.pagesizes import landscape
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        rightMargin=0.5*inch, 
        leftMargin=0.5*inch,
        topMargin=0.5*inch, 
        bottomMargin=0.5*inch
    )
    
    # Container for the 'Flowable' objects
    story = []
    
    # Define styles
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#152d64'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#152d64'),
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_LEFT
    )
    
    wrap_style = ParagraphStyle(
        'WrapText',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_LEFT,
        spaceAfter=0
    )
    
    # Title
    story.append(Paragraph("ITEMS TALLY REPORT", title_style))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", normal_style))
    story.append(Spacer(1, 20))
    
    # Filter information
    filter_info = []
    if user_filter:
        try:
            from django.contrib.auth.models import User
            user = User.objects.get(id=user_filter)
            user_name = f"{user.first_name} {user.last_name}".strip() or user.username
            filter_info.append(f"User: {user_name}")
        except:
            pass
    
    if department_filter:
        try:
            from .models import Department
            dept = Department.objects.get(id=department_filter)
            filter_info.append(f"Department: {dept.name}")
        except:
            pass
    
    if date_from:
        filter_info.append(f"From: {date_from}")
    
    if date_to:
        filter_info.append(f"To: {date_to}")
    
    if status_filter:
        status_display = dict(SupplyRequestItem.STATUS_CHOICES).get(status_filter, status_filter)
        filter_info.append(f"Status: {status_display}")
    
    if filter_info:
        story.append(Paragraph("<b>Filters Applied:</b> " + " | ".join(filter_info), normal_style))
        story.append(Spacer(1, 15))
    
    # Aggregate items by supply name
    items_tally = {}
    total_items_requested = 0
    total_items_approved = 0
    
    for batch_request in completed_requests:
        for item in batch_request.items.all():
            supply_name = item.supply.supply_name
            
            if supply_name not in items_tally:
                items_tally[supply_name] = {
                    'total_requested': 0,
                    'total_approved': 0,
                    'request_count': 0,
                    'unit': item.supply.unit if hasattr(item.supply, 'unit') and item.supply.unit else 'pcs'
                }
            
            items_tally[supply_name]['total_requested'] += item.quantity
            items_tally[supply_name]['total_approved'] += (item.approved_quantity or 0)
            items_tally[supply_name]['request_count'] += 1
            
            total_items_requested += item.quantity
            total_items_approved += (item.approved_quantity or 0)
    
    total_item_types = len(items_tally)
    story.append(Paragraph(f"<b>Total Item Types:</b> {total_item_types} | <b>Total Quantity Requested:</b> {total_items_requested} | <b>Total Quantity Approved:</b> {total_items_approved}", normal_style))
    story.append(Spacer(1, 15))
    
    if total_item_types == 0:
        story.append(Paragraph("<b>No items found with the selected filters.</b>", normal_style))
        story.append(Spacer(1, 20))
        story.append(Paragraph("Try adjusting your filter criteria or check back later.", normal_style))
    else:
        # Create a table with items tally
        story.append(Paragraph("<b>Items Summary</b>", heading_style))
        story.append(Spacer(1, 10))
        
        # Prepare table data
        table_data = [
            ["Item Name", "Unit", "Total\nRequested", "Total\nApproved", "Times\nRequested"]
        ]
        
        # Sort items by name
        for supply_name in sorted(items_tally.keys()):
            item_data = items_tally[supply_name]
            table_data.append([
                Paragraph(supply_name, wrap_style),
                item_data['unit'],
                str(item_data['total_requested']),
                str(item_data['total_approved']),
                str(item_data['request_count'])
            ])
        
        # Create table with appropriate column widths for landscape
        main_table = Table(table_data, colWidths=[
            3.0*inch,   # Item Name (slightly reduced for new unit column)
            0.8*inch,   # Unit
            1.1*inch,   # Total Requested
            1.1*inch,   # Total Approved
            1.1*inch    # Times Requested
        ])
        
        main_table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#152d64')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            
            # Data rows styling
            ('FONTSIZE', (0, 1), (-1, -1), 7.5),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),     # Item name left
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),   # Unit center
            ('ALIGN', (2, 1), (-1, -1), 'CENTER'),  # Quantities and count center
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Padding
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#152d64')),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        
        story.append(main_table)
    
    # Build PDF
    try:
        doc.build(story)
        
        # Get the value of the BytesIO buffer and write it to the response
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return HttpResponse(f"Error generating PDF: {str(e)}\n\nDetails:\n{error_details}", status=500, content_type='text/plain')


@login_required
@permission_required('app.view_admin_module', raise_exception=True)
def supply_approved_tally(request):
    """
    Admin page to view the approved supplies tally.
    Shows all given supplies per item with their approved quantities.
    Includes search, department, user, category, and date filters with pagination.
    Includes analytics dashboard for supply allocation by department and category.
    """
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.db.models import Sum, Count, Q
    
    # Get filter parameters
    search_query = request.GET.get('search', '').strip()
    department_filter = request.GET.get('department', '')
    user_filter = request.GET.get('user', '')
    category_filter = request.GET.get('category', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    date_range_filter = request.GET.get('date_range', '')  # For analytics filters: 30days, 90days, 1year
    
    # Build base queryset for approved items (only completed/distributed items)
    approved_items_query = SupplyRequestItem.objects.filter(
        status='completed',
        batch_request__status__in=['approved', 'for_claiming', 'completed']
    ).select_related('supply', 'supply__category', 'batch_request', 'batch_request__user', 'batch_request__user__userprofile', 'batch_request__user__userprofile__department')
    
    # Apply filters
    if department_filter:
        approved_items_query = approved_items_query.filter(batch_request__user__userprofile__department_id=department_filter)
    
    if user_filter:
        approved_items_query = approved_items_query.filter(batch_request__user_id=user_filter)
    
    if category_filter:
        approved_items_query = approved_items_query.filter(supply__category_id=category_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            approved_items_query = approved_items_query.filter(batch_request__request_date__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            approved_items_query = approved_items_query.filter(batch_request__request_date__date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Aggregate supplies from approved/completed batches
    approved_items_tally = {}
    approved_items = approved_items_query
    
    # Aggregate by supply
    for item in approved_items:
        supply_name = item.supply.supply_name
        supply_id = item.supply.id
        
        if supply_name not in approved_items_tally:
            approved_items_tally[supply_name] = {
                'supply_id': supply_id,
                'supply_name': supply_name,
                'unit': item.supply.unit or 'pcs',
                'total_approved_quantity': 0,
                'num_requests': 0,
                'batch_requests': []
            }
        
        approved_quantity = item.approved_quantity or item.quantity
        approved_items_tally[supply_name]['total_approved_quantity'] += approved_quantity
        approved_items_tally[supply_name]['num_requests'] += 1
        
        # Track which batches this item is from
        if item.batch_request.id not in [br['batch_id'] for br in approved_items_tally[supply_name]['batch_requests']]:
            approved_items_tally[supply_name]['batch_requests'].append({
                'batch_id': item.batch_request.id,
                'batch_status': item.batch_request.status,
                'requestor': item.batch_request.user.get_full_name() or item.batch_request.user.username
            })
    
    # Apply search filter to tally data
    if search_query:
        approved_items_tally = {
            k: v for k, v in approved_items_tally.items() 
            if search_query.lower() in k.lower() or search_query.lower() in str(v['supply_id']).lower()
        }
    
    # Sort by supply name
    sorted_tally = sorted(approved_items_tally.items(), key=lambda x: x[0])
    tally_data = [tally_info for _, tally_info in sorted_tally]
    
    # Pagination
    paginator = Paginator(tally_data, 10)  # 10 items per page
    page = request.GET.get('page', 1)
    
    try:
        tally_page = paginator.page(page)
    except PageNotAnInteger:
        tally_page = paginator.page(1)
    except EmptyPage:
        tally_page = paginator.page(paginator.num_pages)
    
    # Build URL parameters for pagination
    url_params = []
    if search_query:
        url_params.append(f'search={search_query}')
    if department_filter:
        url_params.append(f'department={department_filter}')
    if user_filter:
        url_params.append(f'user={user_filter}')
    if category_filter:
        url_params.append(f'category={category_filter}')
    if date_from:
        url_params.append(f'date_from={date_from}')
    if date_to:
        url_params.append(f'date_to={date_to}')
    
    base_url_params = '&'.join(url_params)
    
    # Calculate summary
    total_items = len(tally_data)
    total_quantity = sum(item['total_approved_quantity'] for item in tally_data)
    total_requests = sum(item['num_requests'] for item in tally_data)
    
    # Get departments for filter dropdown
    departments = Department.objects.all().order_by('name')
    
    # Get categories for filter dropdown
    categories = SupplyCategory.objects.all().order_by('name')
    
    # Get users who have made approved requests for filter dropdown
    users_with_approved = User.objects.filter(
        supplyrequestbatch__status__in=['approved', 'for_claiming', 'completed']
    ).distinct().order_by('first_name', 'last_name')
    
    # If department filter is selected, get only users from that department
    users_in_department = User.objects.none()
    if department_filter:
        users_in_department = User.objects.filter(
            userprofile__department_id=department_filter,
            supplyrequestbatch__status__in=['approved', 'for_claiming', 'completed']
        ).distinct().order_by('first_name', 'last_name')
    
    # Build a dictionary of users by department for JS filtering
    users_by_department = {}
    for dept in departments:
        dept_users = User.objects.filter(
            userprofile__department=dept,
            supplyrequestbatch__status__in=['approved', 'for_claiming', 'completed']
        ).distinct().order_by('first_name', 'last_name')
        users_by_department[dept.id] = [
            {'id': u.id, 'name': u.get_full_name() or u.username}
            for u in dept_users
        ]
    
    # ===== ANALYTICS DATA AGGREGATION =====
    # Re-fetch items for analytics (need fresh queryset)
    approved_items_for_analytics = approved_items_query
    
    # Aggregate supply allocations by department
    analytics_by_department = {}
    analytics_by_category = {}
    analytics_by_item_and_dept = {}  # Track per-item department breakdown
    analytics_by_category_and_dept = {}  # Track category-department breakdown
    
    for item in approved_items_for_analytics:
        approved_qty = item.approved_quantity or item.quantity
        dept = item.batch_request.user.userprofile.department if item.batch_request.user.userprofile else None
        category = item.supply.category
        supply_name = item.supply.supply_name
        supply_id = item.supply.id
        category_name = category.name if category else 'Uncategorized'
        
        # Aggregate by department
        dept_name = dept.name if dept else 'Others'
        if dept_name not in analytics_by_department:
            analytics_by_department[dept_name] = {'quantity': 0, 'id': dept.id if dept else None}
        analytics_by_department[dept_name]['quantity'] += approved_qty
        
        # Aggregate by category
        if category:
            if category_name not in analytics_by_category:
                analytics_by_category[category_name] = {'quantity': 0, 'id': category.id}
            analytics_by_category[category_name]['quantity'] += approved_qty
        
        # Aggregate by item and department (for item selector)
        item_key = f"{supply_id}_{supply_name}"
        if item_key not in analytics_by_item_and_dept:
            analytics_by_item_and_dept[item_key] = {
                'supply_id': supply_id,
                'supply_name': supply_name,
                'departments': {}
            }
        
        if dept_name not in analytics_by_item_and_dept[item_key]['departments']:
            analytics_by_item_and_dept[item_key]['departments'][dept_name] = {
                'quantity': 0,
                'dept_id': dept.id if dept else None
            }
        analytics_by_item_and_dept[item_key]['departments'][dept_name]['quantity'] += approved_qty
        
        # Aggregate by category and department (for category filter)
        if category_name not in analytics_by_category_and_dept:
            analytics_by_category_and_dept[category_name] = {}
        
        if dept_name not in analytics_by_category_and_dept[category_name]:
            analytics_by_category_and_dept[category_name][dept_name] = {
                'quantity': 0,
                'dept_id': dept.id if dept else None
            }
        analytics_by_category_and_dept[category_name][dept_name]['quantity'] += approved_qty
    
    # ===== ITEM ANALYTICS (UNFILTERED BY DEPARTMENT) =====
    # Build separate analytics for item selector that ignores department filter
    # Show only COMPLETED items (released and distributed to departments)
    all_approved_items = SupplyRequestItem.objects.filter(
        status='completed',
        batch_request__status__in=['approved', 'for_claiming', 'completed']
    ).select_related('supply', 'supply__category', 'batch_request', 'batch_request__user', 'batch_request__user__userprofile', 'batch_request__user__userprofile__department')
    
    # Apply only user, category, and date filters (NOT department filter)
    if user_filter:
        all_approved_items = all_approved_items.filter(batch_request__user_id=user_filter)
    
    if category_filter:
        all_approved_items = all_approved_items.filter(supply__category_id=category_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            all_approved_items = all_approved_items.filter(batch_request__request_date__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            all_approved_items = all_approved_items.filter(batch_request__request_date__date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Aggregate items by department (without department filter)
    analytics_by_item_and_dept_unfiltered = {}
    
    for item in all_approved_items:
        approved_qty = item.approved_quantity or item.quantity
        dept = item.batch_request.user.userprofile.department if item.batch_request.user.userprofile else None
        supply_name = item.supply.supply_name
        supply_id = item.supply.id
        
        # Aggregate by item and department
        item_key = f"{supply_id}_{supply_name}"
        if item_key not in analytics_by_item_and_dept_unfiltered:
            analytics_by_item_and_dept_unfiltered[item_key] = {
                'supply_id': supply_id,
                'supply_name': supply_name,
                'departments': {}
            }
        
        dept_name = dept.name if dept else 'Others'
        if dept_name not in analytics_by_item_and_dept_unfiltered[item_key]['departments']:
            analytics_by_item_and_dept_unfiltered[item_key]['departments'][dept_name] = {
                'quantity': 0,
                'dept_id': dept.id if dept else None
            }
        analytics_by_item_and_dept_unfiltered[item_key]['departments'][dept_name]['quantity'] += approved_qty
    
    # Prepare chart data
    dept_labels = list(analytics_by_department.keys())
    dept_data = [analytics_by_department[d]['quantity'] for d in dept_labels]
    
    cat_labels = list(analytics_by_category.keys())
    cat_data = [analytics_by_category[c]['quantity'] for c in cat_labels]
    
    # Prepare item-department data structure for dynamic chart (FILTERED - uses department filter)
    item_dept_data = {}
    for item_key, item_data in analytics_by_item_and_dept.items():
        item_name = item_data['supply_name']
        dept_breakdown = {}
        for dept_name, dept_info in item_data['departments'].items():
            dept_breakdown[dept_name] = dept_info['quantity']
        item_dept_data[item_name] = dept_breakdown
    
    # Prepare item-department data for item selector (UNFILTERED - ignores department filter)
    item_dept_data_unfiltered = {}
    for item_key, item_data in analytics_by_item_and_dept_unfiltered.items():
        item_name = item_data['supply_name']
        dept_breakdown = {}
        for dept_name, dept_info in item_data['departments'].items():
            dept_breakdown[dept_name] = dept_info['quantity']
        item_dept_data_unfiltered[item_name] = dept_breakdown
    
    # Prepare category-department data structure for category filter
    category_dept_data = {}
    for category_name, dept_breakdown in analytics_by_category_and_dept.items():
        dept_labels_for_cat = list(dept_breakdown.keys())
        dept_values_for_cat = [dept_breakdown[d]['quantity'] for d in dept_labels_for_cat]
        category_dept_data[category_name] = {
            'labels': dept_labels_for_cat,
            'data': dept_values_for_cat
        }
    
    import json
    users_by_department_json = json.dumps(users_by_department)
    dept_chart_data = json.dumps({'labels': dept_labels, 'data': dept_data})
    cat_chart_data = json.dumps({'labels': cat_labels, 'data': cat_data})
    item_dept_analytics_json = json.dumps(item_dept_data)
    item_dept_analytics_unfiltered_json = json.dumps(item_dept_data_unfiltered)
    category_dept_analytics_json = json.dumps(category_dept_data)
    
    context = {
        'page_obj': tally_page,
        'paginator': paginator,
        'total_items': total_items,
        'total_quantity': total_quantity,
        'total_requests': total_requests,
        'search_query': search_query,
        'department_filter': department_filter,
        'user_filter': user_filter,
        'category_filter': category_filter,
        'date_from': date_from,
        'date_to': date_to,
        'date_range_filter': date_range_filter,
        'departments': departments,
        'categories': categories,
        'users_with_approved': users_in_department if department_filter else users_with_approved,
        'users_by_department_json': users_by_department_json,
        'url_params': '&' + base_url_params if base_url_params else '',
        'dept_chart_data': dept_chart_data,
        'cat_chart_data': cat_chart_data,
        'item_dept_analytics_json': item_dept_analytics_json,
        'item_dept_analytics_unfiltered_json': item_dept_analytics_unfiltered_json,
        'category_dept_analytics_json': category_dept_analytics_json,
    }
    
    return render(request, 'app/supply_approved_tally.html', context)


@login_required
def sample_admin(request):
    """Sample page demonstrating modern sidebar implementation"""
    return render(request, 'app/sample_admin.html')

