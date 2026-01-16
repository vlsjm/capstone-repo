import barcode
from barcode.writer import ImageWriter
from io import BytesIO
import base64
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from django.utils import timezone
from django.core.files.base import ContentFile
import logging
import requests
import os
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

def generate_barcode(code):
    """
    Generate a Code128 barcode and return it as a base64 encoded string.
    Deprecated: Use generate_barcode_image() for new implementations.
    """
    # Create barcode instance
    code128 = barcode.get_barcode_class('code128')
    
    # Configure barcode writer options for 1.5 inch width
    # 1.5 inch = 144 pixels at 96 DPI (standard screen DPI)
    # Module width controls the width of individual bars
    writer_options = {
        'module_width': 0.3,  # Width of individual bars in mm (larger = wider barcode)
        'module_height': 5.0,  # Height of bars in mm
        'quiet_zone': 3.0,  # Quiet zone on sides in mm
        'font_size': 1,  # Font size for text
        'text_distance': 5.0,  # Distance between barcode and text in mm (increased for balanced spacing)
        'dpi': 300,  # Higher DPI for better print quality
    }
    
    # Generate the barcode
    rv = BytesIO()
    code128(code, writer=ImageWriter()).write(rv, options=writer_options)
    
    # Convert to base64
    image_base64 = base64.b64encode(rv.getvalue()).decode()
    return f"data:image/png;base64,{image_base64}"


def generate_barcode_image(code):
    """
    Generate a Code128 barcode and return it as a ContentFile for ImageField.
    Returns a tuple: (filename, ContentFile)
    """
    # Create barcode instance
    code128 = barcode.get_barcode_class('code128')
    
    # Pad short codes to minimum 12 characters for consistent barcode width
    # This ensures all barcodes have similar bar thickness
    barcode_text = code
    if len(code) < 12:
        # Pad with spaces on the right to make all barcodes similar length
        barcode_text = code.ljust(12)
    
    # Configure barcode writer options for 1.5 inch width
    # 1.5 inch = 38.1 mm
    # Module width controls the width of individual bars
    writer_options = {
        'module_width': 0.3,  # Width of individual bars in mm
        'module_height': 15.0,  # Height of bars in mm (increased by 5mm)
        'quiet_zone': 3.0,  # Quiet zone on sides in mm
        'font_size': 10,  # Font size for text
        'text_distance': 5.0,  # Distance between barcode and text in mm
        'dpi': 300,  # Higher DPI for better print quality
    }
    
    # Generate the barcode with padded text
    rv = BytesIO()
    code128(barcode_text, writer=ImageWriter()).write(rv, options=writer_options)
    
    # Return as ContentFile with original filename (not padded)
    filename = f"{code}.png"
    return filename, ContentFile(rv.getvalue())
    return filename, ContentFile(rv.getvalue())



def send_batch_request_completion_email(batch_request, approved_items, rejected_items):
    """
    Send an email notification when a batch request is fully processed and ready for claiming.
    
    Args:
        batch_request: The SupplyRequestBatch instance
        approved_items: QuerySet of approved items
        rejected_items: QuerySet of rejected items
    """
    try:
        user = batch_request.user
        
        # Skip if user doesn't have an email
        if not user.email:
            logger.warning(f"User {user.username} doesn't have an email address. Skipping batch completion email notification.")
            return False
            
        # Prepare context for email template
        # Build absolute URL for dashboard
        site_url = os.getenv('SITE_URL', 'http://localhost:8000')
        dashboard_url = f"{site_url}/userpanel/dashboard/"
        
        context = {
            'user': user,
            'batch_request': batch_request,
            'approved_items': approved_items,
            'rejected_items': rejected_items,
            'has_approved_items': approved_items.exists(),
            'has_rejected_items': rejected_items.exists(),
            'total_approved': approved_items.count(),
            'total_rejected': rejected_items.count(),
            'total_items': batch_request.items.count(),
            'dashboard_url': dashboard_url,
            'current_year': timezone.now().year,
        }
        
        # Render email template
        html_message = render_to_string('app/email/batch_request_completed.html', context)
        plain_message = strip_tags(html_message)
        
        # Create subject based on batch status
        if approved_items.exists() and not rejected_items.exists():
            subject = f"Batch Request #{batch_request.id} - All Items Approved"
        elif approved_items.exists() and rejected_items.exists():
            subject = f"Batch Request #{batch_request.id} - Partially Approved"
        else:
            subject = f"Batch Request #{batch_request.id} - Request Rejected"
        
        # Send email
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.EMAIL_HOST_USER,
            recipient_list=[user.email],
            html_message=html_message,
            fail_silently=False,
        )
        
        logger.info(f"Batch request completion email sent to {user.email} for batch #{batch_request.id}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send batch completion email to {user.email}: {str(e)}")
        return False 

def send_borrow_batch_request_completion_email(batch_request, approved_items, rejected_items):
    """
    Send an email notification when a borrow batch request is fully processed and ready for claiming.
    
    Args:
        batch_request: The BorrowRequestBatch instance
        approved_items: QuerySet of approved items
        rejected_items: QuerySet of rejected items
    """
    try:
        user = batch_request.user
        
        # Skip if user doesn't have an email
        if not user.email:
            logger.warning(f"User {user.username} doesn't have an email address. Skipping borrow batch completion email notification.")
            return False
            
        # Prepare context for email template
        # Build absolute URL for dashboard
        site_url = os.getenv('SITE_URL', 'http://localhost:8000')
        dashboard_url = f"{site_url}/userpanel/dashboard/"
        
        context = {
            'user': user,
            'batch_request': batch_request,
            'approved_items': approved_items,
            'rejected_items': rejected_items,
            'has_approved_items': approved_items.exists(),
            'has_rejected_items': rejected_items.exists(),
            'total_approved': approved_items.count(),
            'total_rejected': rejected_items.count(),
            'total_items': batch_request.items.count(),
            'dashboard_url': dashboard_url,
            'current_year': timezone.now().year,
        }
        
        # Render email template
        html_message = render_to_string('app/email/borrow_batch_request_completed.html', context)
        plain_message = strip_tags(html_message)
        
        # Create subject based on batch status
        if approved_items.exists() and not rejected_items.exists():
            subject = f"Borrow Request #{batch_request.id} - All Items Approved"
        elif approved_items.exists() and rejected_items.exists():
            subject = f"Borrow Request #{batch_request.id} - Partially Approved"
        else:
            subject = f"Borrow Request #{batch_request.id} - Request Rejected"
        
        # Send email
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.EMAIL_HOST_USER,
            recipient_list=[user.email],
            html_message=html_message,
            fail_silently=False,
        )
        
        logger.info(f"Borrow batch request completion email sent to {user.email} for batch #{batch_request.id}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send borrow batch completion email to {user.email}: {str(e)}")
        return False 


def send_sms_alert(phone_number, message, sms_provider=0):
    """
    Send an SMS alert using the specified SMS provider.
    
    Args:
        phone_number (str): Recipient's phone number
        message (str): Message content
        sms_provider (int): SMS Provider (0 or 1) - default: 0
    
    Returns:
        tuple: (success: bool, response_data: dict or str)
    """
    try:
        # Get API token from environment or settings
        api_token = os.getenv('SMS_API_TOKEN') or getattr(settings, 'SMS_API_TOKEN', None)
        
        if not api_token:
            logger.warning("SMS API TOKEN not configured. Skipping SMS alert.")
            return False, "SMS API TOKEN not configured"
        
        # Get SMS API endpoint from environment or settings (configurable)
        # Default: iProgtech SMS API
        sms_url = os.getenv('SMS_API_ENDPOINT') or getattr(settings, 'SMS_API_ENDPOINT', 
                                                            'https://sms.iprogtech.com/api/v1/sms_messages')
        
        # Prepare request parameters (URL encoded for API)
        params = {
            'api_token': api_token,
            'phone_number': phone_number,
            'message': message
        }
        
        # Send SMS request
        response = requests.post(sms_url, params=params, timeout=10)
        
        if response.status_code == 200:
            logger.info(f"SMS alert sent successfully to {phone_number}")
            return True, response.json()
        else:
            logger.error(f"Failed to send SMS to {phone_number}. Status: {response.status_code}")
            return False, response.text
            
    except requests.exceptions.RequestException as e:
        logger.error(f"SMS request failed for {phone_number}: {str(e)}")
        return False, str(e)
    except Exception as e:
        logger.error(f"Unexpected error sending SMS to {phone_number}: {str(e)}")
        return False, str(e)


def send_overdue_borrow_sms(borrow_request):
    """
    Send an SMS reminder for an overdue borrow request.
    
    Args:
        borrow_request: The BorrowRequest instance
    
    Returns:
        bool: True if SMS was sent successfully, False otherwise
    """
    try:
        user = borrow_request.user
        
        # Get user phone number
        phone_number = None
        try:
            user_profile = user.userprofile
            phone_number = user_profile.phone
        except:
            logger.warning(f"User {user.username} doesn't have a phone number. Skipping SMS alert.")
            return False
        
        if not phone_number:
            logger.warning(f"User {user.username} has no phone number on file. Skipping SMS alert.")
            return False
        
        # Calculate days overdue
        from datetime import date
        days_overdue = (date.today() - borrow_request.return_date).days
        
        # Create SMS message
        message = (
            f"Hello {user.first_name or user.username},\n\n"
            f"This is a reminder that your borrow of {borrow_request.property.property_name} "
            f"(Qty: {borrow_request.quantity}) is OVERDUE.\n\n"
            f"Original return date: {borrow_request.return_date}\n"
            f"Days overdue: {days_overdue}\n\n"
            f"Please return the item at your earliest convenience.\n\n"
            f"Thank you,\nResource Hive Team"
        )
        
        # Send SMS
        success, response = send_sms_alert(phone_number, message)
        
        if success:
            logger.info(f"Overdue reminder SMS sent to {user.username} ({phone_number})")
        else:
            logger.error(f"Failed to send overdue reminder SMS to {user.username}: {response}")
        
        return success
        
    except Exception as e:
        logger.error(f"Error sending overdue SMS reminder: {str(e)}")
        return False


def calculate_reminder_trigger_date(borrow_date, return_date):
    """
    Calculate when to send a near-overdue reminder based on borrow duration.
    
    Uses a hybrid approach:
    - For SHORT-TERM borrows (1-2 days): Send reminder X hours before return date
    - For LONGER borrows: Calculate reminder based on percentage of borrow period remaining
    - Cap between MIN and MAX days to avoid too-early or too-late reminders
    
    Args:
        borrow_date: The date when the item was borrowed
        return_date: The date when the item should be returned
    
    Returns:
        datetime.datetime: The timezone-aware datetime on which the reminder should be sent
    """
    from datetime import timedelta, datetime, time
    from django.utils import timezone as tz
    from zoneinfo import ZoneInfo
    
    try:
        # Get the configured timezone
        local_tz = ZoneInfo(settings.TIME_ZONE)
        
        # Calculate total days borrowed
        days_borrowed = (return_date - borrow_date).days
        
        # Special handling for short-term borrows (1-2 days)
        if days_borrowed <= settings.BORROW_SHORT_TERM_DAYS:
            # For short-term borrows, send reminder X hours before return date
            # Assuming return date is end of day (23:59), subtract the notice hours
            return_datetime = datetime.combine(return_date, time(23, 59))
            # Make it timezone-aware
            return_datetime = return_datetime.replace(tzinfo=local_tz)
            reminder_trigger_datetime = return_datetime - timedelta(hours=settings.BORROW_SHORT_TERM_NOTICE_HOURS)
            
            logger.debug(
                f"SHORT-TERM borrow ({days_borrowed} days): "
                f"Reminder will trigger {settings.BORROW_SHORT_TERM_NOTICE_HOURS} hours before return "
                f"({reminder_trigger_datetime})"
            )
            
            return reminder_trigger_datetime
        
        # For longer borrows, use the percentage-based calculation
        # Calculate reminder days before return based on percentage of borrow period
        reminder_days_before = days_borrowed * settings.BORROW_REMINDER_PERCENTAGE
        
        # Apply min/max caps
        reminder_days_before = max(
            settings.BORROW_REMINDER_MIN_DAYS,
            min(settings.BORROW_REMINDER_MAX_DAYS, reminder_days_before)
        )
        
        # Calculate the trigger date (at start of day for longer borrows)
        reminder_trigger_date = return_date - timedelta(days=reminder_days_before)
        reminder_trigger_datetime = datetime.combine(reminder_trigger_date, time(0, 0))
        # Make it timezone-aware
        reminder_trigger_datetime = reminder_trigger_datetime.replace(tzinfo=local_tz)
        
        logger.debug(
            f"LONG-TERM borrow ({days_borrowed} days): "
            f"Reminder {reminder_days_before:.1f} days before return ({reminder_trigger_datetime.date()})"
        )
        
        return reminder_trigger_datetime
        
    except Exception as e:
        logger.error(f"Error calculating reminder trigger date: {str(e)}")
        return datetime.combine(return_date, time(0, 0)).replace(tzinfo=ZoneInfo(settings.TIME_ZONE))  # Fallback to return date if error



def send_near_overdue_borrow_email(borrow_request_item):
    """
    Send an email reminder for a borrow request item that is approaching its return date.
    
    Args:
        borrow_request_item: The BorrowRequestItem instance
    
    Returns:
        bool: True if email was sent successfully, False otherwise
    """
    try:
        from datetime import date, datetime, time
        
        user = borrow_request_item.batch_request.user
        
        # Skip if user doesn't have an email
        if not user.email:
            logger.warning(
                f"User {user.username} doesn't have an email address. "
                f"Skipping near-overdue reminder."
            )
            return False
        
        # Calculate days until return
        today = date.today()
        days_until_return = (borrow_request_item.return_date - today).days
        
        # Calculate hours until return for short-term borrows
        now = timezone.now()
        return_datetime = datetime.combine(borrow_request_item.return_date, time(23, 59))
        hours_until_return = (return_datetime - now).total_seconds() / 3600
        
        # Determine if this is a short-term borrow
        borrow_duration = (borrow_request_item.return_date - borrow_request_item.batch_request.request_date.date()).days
        is_short_term = borrow_duration <= settings.BORROW_SHORT_TERM_DAYS
        
        # Prepare context for email template
        # Build absolute URL for dashboard
        site_url = os.getenv('SITE_URL', 'http://localhost:8000')
        dashboard_url = f"{site_url}/userpanel/dashboard/"
        
        context = {
            'user': user,
            'batch_request': borrow_request_item.batch_request,
            'item': borrow_request_item,
            'property_name': borrow_request_item.property.property_name,
            'quantity': borrow_request_item.quantity,
            'return_date': borrow_request_item.return_date,
            'days_until_return': days_until_return,
            'hours_until_return': int(hours_until_return),
            'is_short_term': is_short_term,
            'dashboard_url': dashboard_url,
            'current_year': timezone.now().year,
        }
        
        # Render email template
        html_message = render_to_string('app/email/borrow_near_overdue_reminder.html', context)
        plain_message = strip_tags(html_message)
        
        # Create subject - show hours for short-term, days for longer borrows
        if is_short_term and hours_until_return < 24:
            subject = f"Reminder: Your borrowed item '{borrow_request_item.property.property_name}' is due in {int(hours_until_return)} hour(s)"
        else:
            subject = f"Reminder: Your borrowed item '{borrow_request_item.property.property_name}' is due in {days_until_return} day(s)"
        
        # Send email
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.EMAIL_HOST_USER,
            recipient_list=[user.email],
            html_message=html_message,
            fail_silently=False,
        )
        
        logger.info(
            f"Near-overdue reminder email sent to {user.email} for item "
            f"{borrow_request_item.property.property_name} (due {borrow_request_item.return_date}) "
            f"- {days_until_return} days / {int(hours_until_return)} hours remaining"
        )
        return True
        
    except Exception as e:
        logger.error(f"Failed to send near-overdue reminder email: {str(e)}")
        return False


def send_reservation_expired_email(reservation_batch):
    """
    Send an email notification when a reservation batch expires.
    
    Args:
        reservation_batch: The ReservationBatch instance that has expired
    """
    try:
        user = reservation_batch.user
        
        # Skip if user doesn't have an email
        if not user.email:
            logger.warning(f"User {user.username} doesn't have an email address. Skipping reservation expiration email notification.")
            return False
            
        # Prepare context for email template
        site_url = os.getenv('SITE_URL', 'http://localhost:8000')
        dashboard_url = f"{site_url}/userpanel/dashboard/"
        
        # Get items in the batch
        items = reservation_batch.items.all()
        
        context = {
            'user': user,
            'batch_id': reservation_batch.id,
            'batch': reservation_batch,
            'items': items,
            'total_items': items.count(),
            'latest_return_date': reservation_batch.latest_return_date,
            'dashboard_url': dashboard_url,
            'current_year': timezone.now().year,
        }
        
        # Render email template
        html_message = render_to_string('app/email/reservation_expired.html', context)
        plain_message = strip_tags(html_message)
        
        subject = f"Reservation Batch #{reservation_batch.id} - Request Expired"
        
        # Send email
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.EMAIL_HOST_USER,
            recipient_list=[user.email],
            html_message=html_message,
            fail_silently=False,
        )
        
        logger.info(f"Reservation expiration email sent to {user.email} for batch #{reservation_batch.id}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send reservation expiration email to {user.email}: {str(e)}")
        return False

def parse_ppmp_excel(ppmp_instance):
    """
    Parse PPMP Excel file and create PPMPItem instances.
    Looks for the sheet named '3- 2025 PPMP FORM' and extracts all data.
    
    Args:
        ppmp_instance: PPMP model instance with uploaded file
    
    Returns:
        tuple: (success: bool, message: str, items_count: int)
    """
    from .models import PPMPItem
    
    try:
        # Load the workbook
        wb = load_workbook(ppmp_instance.file.path, data_only=True)
        
        # Find the correct sheet - looking for "3- 2025 PPMP FORM" or similar
        target_sheet = None
        for sheet_name in wb.sheetnames:
            if 'PPMP FORM' in sheet_name.upper() or '3-' in sheet_name:
                target_sheet = wb[sheet_name]
                break
        
        if not target_sheet:
            return False, "Could not find '3- 2025 PPMP FORM' sheet in the Excel file", 0
        
        items_created = 0
        max_rows = 2000  # Safety limit to prevent parsing too many rows
        
        # Start reading from a reasonable row (skip headers)
        # Based on the image, data starts around row 23
        start_row = 23
        
        for row_num, row in enumerate(target_sheet.iter_rows(min_row=start_row), start=start_row):
            # Safety check: stop if we've reached max rows
            if row_num > start_row + max_rows:
                logger.warning(f"Reached maximum row limit ({max_rows}). Stopping parse.")
                break
                
            # Get cell values with safety checks
            def get_cell_value(cell):
                return cell.value if cell and cell.value is not None else ''
            
            # Extract data from each column (based on the Excel structure)
            # A=0(Index), B=1(PS-DBM/NON-PS), C=2(MOOE), D=3(Sub Cat A), E=4(Sub Cat B), 
            # F=5(Sub Cat C), G=6(Code), H=7(UACS), I=8(Description), J=9(U/M), K=10(Unit Price),
            # L=11(Q1), M=12(Q2), N=13(Q3), O=14(Q4), P=15(TOTAL QUANTITY), Q=16(TOTAL AMOUNT)
            unit = get_cell_value(row[9]) if len(row) > 9 else ''  # Column J - U/M (Book, Pad, Box, etc.)
            mode = get_cell_value(row[2]) if len(row) > 2 else ''  # Column C - MOOE
            supplies_materials = get_cell_value(row[3]) if len(row) > 3 else ''  # Column D - Sub Category A
            office_supplies = get_cell_value(row[4]) if len(row) > 4 else ''  # Column E - Sub Category B
            common_supplies = get_cell_value(row[5]) if len(row) > 5 else ''  # Column F - Sub Category C
            
            # UACS code is in column H (index 7)
            description = get_cell_value(row[7]) if len(row) > 7 else ''  # Column H - UACS CODE (5020301000)
            
            # Item name/description is in column I (index 8)
            unit_measure = get_cell_value(row[8]) if len(row) > 8 else ''  # Column I - Item description
            
            # Skip rows without item name or with very short names (likely headers/footers)
            if not unit_measure or str(unit_measure).strip() == '' or len(str(unit_measure).strip()) < 3:
                continue
            
            # Skip rows that look like headers or section breaks (contains common header keywords)
            item_name_lower = str(unit_measure).lower()
            skip_keywords = ['total', 'sub-total', 'subtotal', 'grand total', 'page', 'continued', 
                           'item description', 'description', 'ppmp', 'end of', 'section']
            if any(keyword in item_name_lower for keyword in skip_keywords):
                continue
            
            # Unit price, quantity, and total amount columns
            # Column K (index 10) - Unit Price
            # Column P (index 15) - TOTAL QUANTITY (sum of Q1+Q2+Q3+Q4)
            unit_price_val = get_cell_value(row[10]) if len(row) > 10 else 0  # Column K
            quantity_val = get_cell_value(row[15]) if len(row) > 15 else 0  # Column P - TOTAL QUANTITY
            
            # Convert to proper types
            try:
                unit_price = float(unit_price_val) if unit_price_val else 0
            except (ValueError, TypeError):
                unit_price = 0
            
            try:
                quantity = int(quantity_val) if quantity_val else 0
            except (ValueError, TypeError):
                quantity = 0
            
            # Calculate total_amount as unit_price * quantity
            total_amount = unit_price * quantity
            
            # Skip items with zero or negative quantity (likely hidden or placeholder rows)
            if quantity <= 0:
                continue
            
            # Create PPMPItem
            # Note: description = UACS code, unit_measure = actual item name
            PPMPItem.objects.create(
                ppmp=ppmp_instance,
                unit=str(unit),
                mode=str(mode),
                supplies_materials_expense=str(supplies_materials),
                office_supplies_expense=str(office_supplies),
                common_office_supplies=str(common_supplies),
                description=str(description),  # UACS code (e.g., 5020301000)
                unit_measure=str(unit_measure),  # Item name (e.g., "CLIP, backfold, 19mm")
                unit_price=unit_price,
                quantity=quantity,
                total_amount=total_amount,
                row_number=row_num
            )
            
            items_created += 1
        
        wb.close()
        
        if items_created == 0:
            return False, "No valid items found in the Excel file", 0
        
        return True, f"Successfully imported {items_created} items from PPMP", items_created
        
    except FileNotFoundError:
        return False, "Excel file not found", 0
    except Exception as e:
        logger.error(f"Error parsing PPMP Excel: {str(e)}")
        return False, f"Error parsing Excel file: {str(e)}", 0


def check_ppmp_match(supply_name, department, year=None):
    """
    Check if a supply request matches any items in the department's PPMP.
    
    Args:
        supply_name: Name of the supply being requested
        department: Department instance
        year: Year to check (defaults to current year)
    
    Returns:
        dict: {
            'match_found': bool,
            'ppmp': PPMP instance or None,
            'matched_items': list of PPMPItem instances,
            'message': str
        }
    """
    from .models import PPMP, PPMPItem
    import re
    
    if year is None:
        year = timezone.now().year
    
    try:
        # Get PPMP for this department and year
        ppmp = PPMP.objects.get(department=department, year=year)
        
        # Normalize the supply name for better matching
        normalized_supply_name = re.sub(r'\s+', ' ', supply_name.strip().lower())
        
        # Search for matching items in unit_measure field (which contains the actual item name)
        # First try exact partial match
        matched_items = PPMPItem.objects.filter(
            ppmp=ppmp,
            unit_measure__icontains=supply_name
        )
        
        # If no match, try normalized matching (more flexible)
        if not matched_items.exists():
            all_items = PPMPItem.objects.filter(ppmp=ppmp)
            matched_list = []
            for item in all_items:
                # Search in unit_measure (the actual item name) not description (UACS code)
                normalized_item_name = re.sub(r'\s+', ' ', str(item.unit_measure).strip().lower())
                # Check if either contains the other or significant overlap
                if normalized_supply_name in normalized_item_name or normalized_item_name in normalized_supply_name:
                    matched_list.append(item)
                # Also try word-by-word matching for better accuracy
                elif len(normalized_supply_name) > 10:  # Only for longer names
                    supply_words = set(normalized_supply_name.split())
                    item_words = set(normalized_item_name.split())
                    # If 70% of words match, consider it a match
                    common_words = supply_words.intersection(item_words)
                    if len(common_words) / len(supply_words) >= 0.7:
                        matched_list.append(item)
            
            if matched_list:
                matched_items = matched_list
        else:
            matched_items = list(matched_items)
        
        if matched_items and len(matched_items) > 0:
            return {
                'match_found': True,
                'ppmp': ppmp,
                'matched_items': matched_items,
                'message': f"✓ Found {len(matched_items)} matching item(s) in {department.name}'s PPMP {year}"
            }
        else:
            return {
                'match_found': False,
                'ppmp': ppmp,
                'matched_items': [],
                'message': f"⚠ No matching items found in {department.name}'s PPMP {year}"
            }
    
    except PPMP.DoesNotExist:
        return {
            'match_found': False,
            'ppmp': None,
            'matched_items': [],
            'message': f"⚠ No PPMP found for {department.name} for year {year}"
        }
    except Exception as e:
        logger.error(f"Error checking PPMP match: {str(e)}")
        return {
            'match_found': False,
            'ppmp': None,
            'matched_items': [],
            'message': f"Error checking PPMP: {str(e)}"
        }

def send_borrow_request_expired_email(borrow_batch):
    """
    Send an email notification when a borrow request batch expires.
    
    Args:
        borrow_batch: The BorrowRequestBatch instance that has expired
    """
    try:
        user = borrow_batch.user
        
        # Skip if user doesn't have an email
        if not user.email:
            logger.warning(f"User {user.username} doesn't have an email address. Skipping borrow request expiration email notification.")
            return False
            
        # Prepare context for email template
        site_url = os.getenv('SITE_URL', 'http://localhost:8000')
        dashboard_url = f"{site_url}/userpanel/dashboard/"
        
        # Get items in the batch
        items = borrow_batch.items.all()
        
        context = {
            'user': user,
            'batch_id': borrow_batch.id,
            'batch': borrow_batch,
            'items': items,
            'total_items': items.count(),
            'latest_return_date': borrow_batch.latest_return_date,
            'dashboard_url': dashboard_url,
            'current_year': timezone.now().year,
        }
        
        # Render email template
        html_message = render_to_string('app/email/borrow_request_expired.html', context)
        plain_message = strip_tags(html_message)
        
        subject = f"Borrow Request #{borrow_batch.id} - Request Expired"
        
        # Send email
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.EMAIL_HOST_USER,
            recipient_list=[user.email],
            html_message=html_message,
            fail_silently=False,
        )
        
        logger.info(f"Borrow request expiration email sent to {user.email} for batch #{borrow_batch.id}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send borrow request expiration email to {user.email}: {str(e)}")
        return False