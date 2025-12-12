# Additional views for user requests summary
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import TemplateView
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from app.models import SupplyRequestBatch, BorrowRequestBatch, ReservationBatch, SupplyCategory


class UserRequestsSummaryView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """View for users to see a summary of all their requested items, filterable by year with PDF export"""
    template_name = 'userpanel/user_requests_summary.html'
    permission_required = 'app.view_user_module'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get separate filters for tally and items
        tally_year_filter = self.request.GET.get('tally_year', '')
        tally_month_filter = self.request.GET.get('tally_month', '')
        tally_category_filter = self.request.GET.get('tally_category', '')
        tally_search = self.request.GET.get('tally_search', '').strip()
        items_year_filter = self.request.GET.get('items_year', '')
        items_month_filter = self.request.GET.get('items_month', '')
        
        # Initialize items summary
        items_summary = []
        
        # Don't include supply requests in items table - they're shown in tally above
        # Get all borrow requests
        borrow_batches = BorrowRequestBatch.objects.filter(user=self.request.user)
        if items_year_filter:
            borrow_batches = borrow_batches.filter(request_date__year=items_year_filter)
        if items_month_filter:
            borrow_batches = borrow_batches.filter(request_date__month=items_month_filter)
        
        for batch in borrow_batches:
            for item in batch.items.all():
                items_summary.append({
                    'type': 'Borrow',
                    'item_name': item.property.property_name,
                    'category': item.property.category.name if item.property.category else 'Uncategorized',
                    'quantity': item.quantity,
                    'approved_quantity': item.approved_quantity or 0,
                    'status': item.get_status_display(),
                    'request_date': batch.request_date,
                    'return_date': item.return_date,
                    'purpose': batch.purpose,
                    'request_id': f"BB-{batch.id}"
                })
        
        # Get all reservations
        reservation_batches = ReservationBatch.objects.filter(user=self.request.user)
        if items_year_filter:
            reservation_batches = reservation_batches.filter(request_date__year=items_year_filter)
        if items_month_filter:
            reservation_batches = reservation_batches.filter(request_date__month=items_month_filter)
        
        for batch in reservation_batches:
            for item in batch.items.all():
                items_summary.append({
                    'type': 'Reservation',
                    'item_name': item.property.property_name,
                    'category': item.property.category.name if item.property.category else 'Uncategorized',
                    'quantity': item.quantity,
                    'approved_quantity': item.quantity if item.status == 'approved' else 0,
                    'status': item.get_status_display(),
                    'request_date': batch.request_date,
                    'needed_date': item.needed_date,
                    'return_date': item.return_date,
                    'purpose': batch.purpose,
                    'request_id': f"RB-{batch.id}"
                })
        
        # Sort by date (most recent first)
        items_summary.sort(key=lambda x: x['request_date'], reverse=True)
        
        # Get available years for filter dropdown
        all_years = set()
        for model in [SupplyRequestBatch, BorrowRequestBatch, ReservationBatch]:
            years = model.objects.filter(user=self.request.user).dates('request_date', 'year')
            all_years.update([date.year for date in years])
        
        available_years = sorted(list(all_years), reverse=True)
        
        # Calculate statistics (before pagination) - based on items_year/month filters
        # Note: items_summary now only includes borrow and reservation requests
        total_items = len(items_summary)
        total_requested = sum(item['quantity'] for item in items_summary)
        total_approved = sum(item['approved_quantity'] for item in items_summary)
        
        # Group by type (supply count will be 0 since we excluded supplies)
        supply_count = 0  # Supplies are shown in the tally section above
        borrow_count = len([i for i in items_summary if i['type'] == 'Borrow'])
        reservation_count = len([i for i in items_summary if i['type'] == 'Reservation'])
        
        # Build separate tally data using tally_year_filter and tally_month_filter
        tally_items = []
        
        # Get all supply requests for tally - ONLY COMPLETED/CLAIMED SUPPLIES
        tally_supply_batches = SupplyRequestBatch.objects.filter(user=self.request.user)
        if tally_year_filter:
            tally_supply_batches = tally_supply_batches.filter(request_date__year=tally_year_filter)
        if tally_month_filter:
            tally_supply_batches = tally_supply_batches.filter(request_date__month=tally_month_filter)
        
        for batch in tally_supply_batches:
            # Only include completed supply items (those that have been claimed)
            for item in batch.items.filter(status='completed'):
                category_name = item.supply.category.name if item.supply.category else 'Uncategorized'
                # Apply category filter if specified
                if tally_category_filter and category_name != tally_category_filter:
                    continue
                    
                tally_items.append({
                    'type': 'Supply',
                    'supply_id': item.supply.barcode or f"SUP-{item.supply.id}",
                    'item_name': item.supply.supply_name,
                    'category': category_name,
                    'quantity': item.quantity,
                    'approved_quantity': item.approved_quantity or 0,
                })
        
        # Build item tally (group by item name and type)
        tally_dict = {}
        for item in tally_items:
            # Apply search filter if specified
            if tally_search:
                supply_id_match = tally_search.lower() in item['supply_id'].lower()
                item_name_match = tally_search.lower() in item['item_name'].lower()
                if not (supply_id_match or item_name_match):
                    continue
                    
            key = f"{item['type']}_{item['item_name']}"
            if key not in tally_dict:
                tally_dict[key] = {
                    'type': item['type'],
                    'supply_id': item['supply_id'],
                    'item_name': item['item_name'],
                    'category': item['category'],
                    'total_quantity': 0,
                    'total_approved': 0,
                    'request_count': 0,
                }
            tally_dict[key]['total_quantity'] += item['quantity']
            tally_dict[key]['total_approved'] += item['approved_quantity']
            tally_dict[key]['request_count'] += 1
        
        tally_data = sorted(tally_dict.values(), key=lambda x: x['total_quantity'], reverse=True)
        
        # Pagination for both tally and items
        # Tally pagination
        tally_paginator = Paginator(tally_data, 10)  # 10 items per page
        tally_page = self.request.GET.get('tally_page', 1)
        
        try:
            tally_page_obj = tally_paginator.page(tally_page)
        except PageNotAnInteger:
            tally_page_obj = tally_paginator.page(1)
        except EmptyPage:
            tally_page_obj = tally_paginator.page(tally_paginator.num_pages)
        
        # Items pagination
        paginator = Paginator(items_summary, 10)  # 10 items per page
        page = self.request.GET.get('page', 1)
        
        try:
            items_page = paginator.page(page)
        except PageNotAnInteger:
            items_page = paginator.page(1)
        except EmptyPage:
            items_page = paginator.page(paginator.num_pages)
        
        # Define month names
        month_names = [
            (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
            (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
            (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
        ]
        
        # Get available categories for filter dropdown
        available_categories = SupplyCategory.objects.all().order_by('name')
        
        context.update({
            'items_summary': items_page,
            'items_year_filter': items_year_filter,
            'items_month_filter': items_month_filter,
            'tally_year_filter': tally_year_filter,
            'tally_month_filter': tally_month_filter,
            'tally_category_filter': tally_category_filter,
            'tally_search': tally_search,
            'tally_data': tally_page_obj,
            'available_years': available_years,
            'available_categories': available_categories,
            'month_names': month_names,
            'total_items': total_items,
            'total_requested': total_requested,
            'total_approved': total_approved,
            'supply_count': supply_count,
            'borrow_count': borrow_count,
            'reservation_count': reservation_count,
        })
        
        return context


@login_required
@permission_required('app.view_user_module')
def export_requests_summary_pdf(request):
    """Export user's requests summary as PDF"""
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from io import BytesIO
    from django.utils import timezone
    import traceback
    
    try:
        # Get year and month filters
        year_filter = request.GET.get('year', '')
        month_filter = request.GET.get('month', '')
        
        # Create the HttpResponse object with PDF headers
        response = HttpResponse(content_type='application/pdf')
        
        # Build filename with filters
        filter_text = []
        if year_filter:
            filter_text.append(f"Year_{year_filter}")
        if month_filter:
            month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            filter_text.append(f"Month_{month_names[int(month_filter)]}")
        
        filename = f'Borrow_Reservation_Summary_{"_".join(filter_text) if filter_text else "All_Time"}_{timezone.now().strftime("%Y%m%d")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Create the PDF object using BytesIO buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Container for PDF elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#152d64'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        # Title
        title_text = f"Borrow & Reservation Requests Summary"
        if year_filter or month_filter:
            title_text += " -"
            if year_filter:
                title_text += f" Year {year_filter}"
            if month_filter:
                month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 
                              'July', 'August', 'September', 'October', 'November', 'December']
                title_text += f" {month_names[int(month_filter)]}"
        elements.append(Paragraph(title_text, title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # User info
        user = request.user
        user_name = f"{user.first_name} {user.last_name}" if user.first_name or user.last_name else user.username
        department = user.userprofile.department.name if hasattr(user, 'userprofile') and user.userprofile.department else "N/A"
        
        info_text = f"<b>Requester:</b> {user_name} | <b>Department:</b> {department} | <b>Generated:</b> {timezone.now().strftime('%B %d, %Y %I:%M %p')}"
        info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#495057'))
        elements.append(Paragraph(info_text, info_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Get items summary
        items_summary = []
        
        # Don't include supply requests - they're shown in the tally section
        # Get all borrow requests
        borrow_batches = BorrowRequestBatch.objects.filter(user=request.user)
        if year_filter:
            borrow_batches = borrow_batches.filter(request_date__year=year_filter)
        if month_filter:
            borrow_batches = borrow_batches.filter(request_date__month=month_filter)
        
        for batch in borrow_batches:
            for item in batch.items.all():
                items_summary.append({
                    'type': 'Borrow',
                    'item_name': item.property.property_name,
                    'category': item.property.category.name if item.property.category else 'Uncategorized',
                    'quantity': item.quantity,
                    'approved_quantity': item.approved_quantity or 0,
                    'status': item.get_status_display(),
                    'request_date': batch.request_date,
                    'request_id': f"BB-{batch.id}"
                })
        
        # Get all reservations
        reservation_batches = ReservationBatch.objects.filter(user=request.user)
        if year_filter:
            reservation_batches = reservation_batches.filter(request_date__year=year_filter)
        if month_filter:
            reservation_batches = reservation_batches.filter(request_date__month=month_filter)
        
        for batch in reservation_batches:
            for item in batch.items.all():
                items_summary.append({
                    'type': 'Reservation',
                    'item_name': item.property.property_name,
                    'category': item.property.category.name if item.property.category else 'Uncategorized',
                    'quantity': item.quantity,
                    'approved_quantity': item.quantity if item.status == 'approved' else 0,
                    'status': item.get_status_display(),
                    'request_date': batch.request_date,
                    'request_id': f"RB-{batch.id}"
                })
        
        # Sort by date
        items_summary.sort(key=lambda x: x['request_date'], reverse=True)
        
        # Statistics
        total_items = len(items_summary)
        total_requested = sum(item['quantity'] for item in items_summary)
        total_approved = sum(item['approved_quantity'] for item in items_summary)
        
        stats_text = f"<b>Total Items:</b> {total_items} | <b>Total Quantity Requested:</b> {total_requested} | <b>Total Quantity Approved:</b> {total_approved}"
        elements.append(Paragraph(stats_text, info_style))
        elements.append(Spacer(1, 0.2*inch))
        
        if not items_summary:
            no_data_text = "No requests found for the selected period."
            elements.append(Paragraph(no_data_text, styles['Normal']))
        else:
            # Create table data
            table_data = [
                ["Request ID", "Type", "Item Name", "Category", "Req\nQty", "App\nQty", "Status", "Request Date"]
            ]
            
            for item in items_summary:
                table_data.append([
                    item['request_id'],
                    item['type'],
                    item['item_name'][:40] + '...' if len(item['item_name']) > 40 else item['item_name'],
                    item['category'][:20] + '...' if len(item['category']) > 20 else item['category'],
                    str(item['quantity']),
                    str(item['approved_quantity']),
                    item['status'],
                    item['request_date'].strftime('%m/%d/%Y')
                ])
            
            # Create table with column widths
            col_widths = [0.9*inch, 0.8*inch, 2.2*inch, 1.3*inch, 0.5*inch, 0.5*inch, 1.0*inch, 0.9*inch]
            
            main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
            main_table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#152d64')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                
                # Body styling
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('ALIGN', (4, 1), (5, -1), 'CENTER'),  # Quantity columns centered
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('LEFTPADDING', (0, 1), (-1, -1), 4),
                ('RIGHTPADDING', (0, 1), (-1, -1), 4),
                
                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#152d64')),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ]))
            
            elements.append(main_table)
        
        # Build PDF
        doc.build(elements)
        
        # Get the value of the BytesIO buffer and write it to the response
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response
        
    except Exception as e:
        # Log the error for debugging
        print(f"Error generating PDF: {str(e)}")
        print(traceback.format_exc())
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


@login_required
@permission_required('app.view_user_module')
def export_claimed_supplies_tally_excel(request):
    """Export claimed supplies tally as Excel with filters"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.utils import timezone
    from datetime import datetime
    import traceback
    
    try:
        # Get filters
        year_filter = request.GET.get('year', '')
        month_filter = request.GET.get('month', '')
        category_filter = request.GET.get('category', '')
        start_date_filter = request.GET.get('start_date', '')
        end_date_filter = request.GET.get('end_date', '')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Claimed Supplies Tally"
        
        # Styles
        header_fill = PatternFill(start_color="152d64", end_color="152d64", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="152d64")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_aligned = Alignment(horizontal='center', vertical='center')
        
        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_text = "Claimed Supplies Tally"
        if year_filter or month_filter or category_filter:
            title_text += " - Filtered"
        title_cell.value = title_text
        title_cell.font = title_font
        title_cell.alignment = center_aligned
        
        # User info
        user = request.user
        user_name = f"{user.first_name} {user.last_name}" if user.first_name or user.last_name else user.username
        department = user.userprofile.department.name if hasattr(user, 'userprofile') and user.userprofile.department else "N/A"
        
        ws.merge_cells('A2:E2')
        info_cell = ws['A2']
        info_cell.value = f"Requester: {user_name} | Department: {department} | Generated: {timezone.now().strftime('%B %d, %Y %I:%M %p')}"
        info_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Filter info
        filter_parts = []
        if start_date_filter and end_date_filter:
            try:
                start_date_obj = datetime.strptime(start_date_filter, '%Y-%m-%d')
                end_date_obj = datetime.strptime(end_date_filter, '%Y-%m-%d')
                filter_parts.append(f"Date Range: {start_date_obj.strftime('%B %d, %Y')} - {end_date_obj.strftime('%B %d, %Y')}")
            except:
                pass
        elif start_date_filter:
            try:
                start_date_obj = datetime.strptime(start_date_filter, '%Y-%m-%d')
                filter_parts.append(f"From: {start_date_obj.strftime('%B %d, %Y')}")
            except:
                pass
        elif end_date_filter:
            try:
                end_date_obj = datetime.strptime(end_date_filter, '%Y-%m-%d')
                filter_parts.append(f"Until: {end_date_obj.strftime('%B %d, %Y')}")
            except:
                pass
        elif year_filter:
            filter_parts.append(f"Year: {year_filter}")
            if month_filter:
                month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 
                              'July', 'August', 'September', 'October', 'November', 'December']
                filter_parts.append(f"Month: {month_names[int(month_filter)]}")
        if category_filter:
            filter_parts.append(f"Category: {category_filter}")
        
        if filter_parts:
            ws.merge_cells('A3:E3')
            filter_cell = ws['A3']
            filter_cell.value = f"Filters Applied: {' | '.join(filter_parts)}"
            filter_cell.alignment = Alignment(horizontal='left', vertical='center')
            filter_cell.font = Font(italic=True)
            header_row = 5
        else:
            header_row = 4
        
        # Headers
        headers = ["Supply ID", "Item Name", "Category", "Total Requested", "Total Approved"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_aligned
            cell.border = border
        
        # Get tally data
        tally_items = []
        tally_supply_batches = SupplyRequestBatch.objects.filter(user=request.user)
        
        # Apply date filters - prioritize date range over year/month
        if start_date_filter or end_date_filter:
            if start_date_filter:
                try:
                    start_date = datetime.strptime(start_date_filter, '%Y-%m-%d').date()
                    tally_supply_batches = tally_supply_batches.filter(request_date__gte=start_date)
                except:
                    pass
            if end_date_filter:
                try:
                    end_date = datetime.strptime(end_date_filter, '%Y-%m-%d').date()
                    tally_supply_batches = tally_supply_batches.filter(request_date__lte=end_date)
                except:
                    pass
        else:
            # Use year/month filters if no date range specified
            if year_filter:
                tally_supply_batches = tally_supply_batches.filter(request_date__year=year_filter)
            if month_filter:
                tally_supply_batches = tally_supply_batches.filter(request_date__month=month_filter)
        
        for batch in tally_supply_batches:
            for item in batch.items.filter(status='completed'):
                category_name = item.supply.category.name if item.supply.category else 'Uncategorized'
                if category_filter and category_name != category_filter:
                    continue
                    
                tally_items.append({
                    'supply_id': item.supply.barcode or f"SUP-{item.supply.id}",
                    'item_name': item.supply.supply_name,
                    'category': category_name,
                    'quantity': item.quantity,
                    'approved_quantity': item.approved_quantity or 0,
                })
        
        # Build item tally (group by item name)
        tally_dict = {}
        for item in tally_items:
            key = item['item_name']
            if key not in tally_dict:
                tally_dict[key] = {
                    'supply_id': item['supply_id'],
                    'item_name': item['item_name'],
                    'category': item['category'],
                    'total_quantity': 0,
                    'total_approved': 0,
                }
            tally_dict[key]['total_quantity'] += item['quantity']
            tally_dict[key]['total_approved'] += item['approved_quantity']
        
        tally_data = sorted(tally_dict.values(), key=lambda x: x['total_quantity'], reverse=True)
        
        # Add data rows
        row_num = header_row + 1
        for item in tally_data:
            ws.cell(row=row_num, column=1, value=item['supply_id']).border = border
            ws.cell(row=row_num, column=2, value=item['item_name']).border = border
            ws.cell(row=row_num, column=3, value=item['category']).border = border
            ws.cell(row=row_num, column=4, value=item['total_quantity']).border = border
            ws.cell(row=row_num, column=4).alignment = center_aligned
            ws.cell(row=row_num, column=5, value=item['total_approved']).border = border
            ws.cell(row=row_num, column=5).alignment = center_aligned
            row_num += 1
        
        # Add totals row
        if tally_data:
            row_num += 1
            total_cell = ws.cell(row=row_num, column=1)
            total_cell.value = "TOTAL"
            total_cell.font = Font(bold=True)
            ws.merge_cells(f'A{row_num}:C{row_num}')
            
            total_requested = sum(item['total_quantity'] for item in tally_data)
            total_approved = sum(item['total_approved'] for item in tally_data)
            
            ws.cell(row=row_num, column=4, value=total_requested).font = Font(bold=True)
            ws.cell(row=row_num, column=4).alignment = center_aligned
            ws.cell(row=row_num, column=5, value=total_approved).font = Font(bold=True)
            ws.cell(row=row_num, column=5).alignment = center_aligned
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Build filename
        filter_text = []
        if start_date_filter and end_date_filter:
            try:
                start_date_obj = datetime.strptime(start_date_filter, '%Y-%m-%d')
                end_date_obj = datetime.strptime(end_date_filter, '%Y-%m-%d')
                filter_text.append(f"{start_date_obj.strftime('%Y%m%d')}_to_{end_date_obj.strftime('%Y%m%d')}")
            except:
                pass
        elif start_date_filter:
            try:
                start_date_obj = datetime.strptime(start_date_filter, '%Y-%m-%d')
                filter_text.append(f"From_{start_date_obj.strftime('%Y%m%d')}")
            except:
                pass
        elif end_date_filter:
            try:
                end_date_obj = datetime.strptime(end_date_filter, '%Y-%m-%d')
                filter_text.append(f"Until_{end_date_obj.strftime('%Y%m%d')}")
            except:
                pass
        elif year_filter:
            filter_text.append(f"Year_{year_filter}")
            if month_filter:
                month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                filter_text.append(f"Month_{month_names[int(month_filter)]}")
        if category_filter:
            filter_text.append(f"Cat_{category_filter.replace(' ', '_')}")
        
        filename = f'Claimed_Supplies_Tally_{"_".join(filter_text) if filter_text else "All"}_{timezone.now().strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        print(f"Error generating Excel: {str(e)}")
        print(traceback.format_exc())
        return HttpResponse(f"Error generating Excel: {str(e)}", status=500)
